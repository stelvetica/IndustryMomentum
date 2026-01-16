"""
Excel格式调整工具
功能：
1. 将每个sheet页的A列设置为左对齐
2. 将每个sheet页的A列宽度设置为11
3. 隐藏每个sheet页的19行到52行
4. 比较B9-E9的值，将最大值所在列的8-51行加粗
"""

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from copy import copy


def format_excel(file_path: str):
    """
    调整Excel文件格式
    
    参数:
        file_path: Excel文件路径
    """
    # 加载工作簿
    wb = load_workbook(file_path)
    
    # 遍历所有sheet页
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"正在处理sheet: {sheet_name}")
        
        # 1. 设置A列宽度为11
        ws.column_dimensions['A'].width = 11
        
        # 2. 设置A列左对齐
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=1)  # A列是第1列
            cell.alignment = Alignment(horizontal='left')
        
        # 3. 隐藏19行到52行
        for row in range(19, 53):  # 19到52行（包含52）
            ws.row_dimensions[row].hidden = True
        
        # 4. 比较B9-E9的值，将最大值所在列的8-51行加粗
        # B=2, C=3, D=4, E=5
        values = {}
        for col in range(2, 6):  # B到E列
            cell_value = ws.cell(row=9, column=col).value
            if cell_value is not None:
                try:
                    values[col] = float(cell_value)
                except (ValueError, TypeError):
                    values[col] = float('-inf')
            else:
                values[col] = float('-inf')
        
        if values:
            max_col = max(values, key=values.get)
            print(f"  B9-E9最大值在第{max_col}列 (值={values[max_col]})")
            
            # 将该列的8-51行加粗
            for row in range(8, 52):  # 8到51行
                cell = ws.cell(row=row, column=max_col)
                # 保留原有字体属性，只修改bold
                if cell.font:
                    new_font = copy(cell.font)
                    new_font = Font(
                        name=cell.font.name,
                        size=cell.font.size,
                        bold=True,
                        italic=cell.font.italic,
                        color=cell.font.color
                    )
                else:
                    new_font = Font(bold=True)
                cell.font = new_font
    
    # 保存文件
    wb.save(file_path)
    print(f"文件已保存: {file_path}")


if __name__ == '__main__':
    # 处理默认文件
    format_excel('factors_analysis_report.xlsx')