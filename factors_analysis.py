"""
因子分析模块
负责批量因子分析、IC/IR 计算、分层回测、统计指标计算
输出到Excel，每个因子一个sheet页，包含因子说明

支持多周期回测：20日、60日、120日、240日
超额收益基准：申万一级行业等权收益
"""
import pandas as pd
import numpy as np
import inspect
from scipy import stats
from copy import copy

# 导入数据加载模块和因子模块
import data_loader
from data_loader import (
    load_price_df, load_high_df, load_low_df,
    load_turnover_df, load_volume_df, DEFAULT_CACHE_FILE,
    load_constituent_df, load_stock_price_df, load_stock_mv_df, load_industry_code_df,
    load_barra_factor_returns,
    DEFAULT_CONSTITUENT_FILE, DEFAULT_STOCK_PRICE_FILE, DEFAULT_STOCK_MV_FILE, DEFAULT_INDUSTRY_CODE_FILE
)
import factor_

# 默认参数
DEFAULT_REBALANCE_FREQ = 20   # 调仓频率/预测窗口 (天) - 仅在非月度调仓时使用
DEFAULT_MONTHLY_REBALANCE = True  # 默认使用月度调仓（每月最后一个交易日）
N_LAYERS = 5                  # 分层回测层数
WINDOWS = [20, 60, 120, 240]  # 多周期回测窗口（因子计算的回溯窗口）

# 统一因子配置表
# 每个因子包含：
#   - func: 因子计算函数
#   - base_warmup: 固定预热期（交易日）
#   - window_multiplier: 窗口倍数，预热期 = base_warmup + window * window_multiplier
#   - windows: 该因子适用的窗口列表，None 表示使用默认的 WINDOWS
#   - requires_constituent: 是否需要成分股数据
#
# 按照 factor_.py 中的定义顺序排列
FACTOR_CONFIG = {
    # ===== 基础动量因子（只需要 window 天）=====
    'momentum': {
        'func': factor_.momentum,
        'base_warmup': 0,
        'window_multiplier': 1,
        'windows': None,
        'requires_constituent': False,
        'description': '传统动量因子（区间收益率）'
    },
    'momentum_zscore': {
        'func': factor_.momentum_zscore,
        'base_warmup': 0,
        'window_multiplier': 1,
        'windows': None,
        'requires_constituent': False,
        'description': '标准化动量因子（横截面Z-score）'
    },
    'momentum_sharpe': {
        'func': factor_.momentum_sharpe,
        'base_warmup': 0,
        'window_multiplier': 1,
        'windows': None,
        'requires_constituent': False,
        'description': '夏普动量因子（风险调整后的动量）'
    },
    'momentum_calmar_ratio': {
        'func': factor_.momentum_calmar_ratio,
        'base_warmup': 0,
        'window_multiplier': 1,
        'windows': None,
        'requires_constituent': False,
        'description': 'Calmar比率因子（最大回撤调整）'
    },
    'momentum_rank_zscore': {
        'func': factor_.momentum_rank_zscore,
        'base_warmup': 0,
        'window_multiplier': 1,
        'windows': None,
        'requires_constituent': False,
        'description': 'Rank标准化动量因子（排名标准化）'
    },
    
    # ===== 平稳动量因子 =====
    'momentum_turnover_adj': {
        'func': factor_.momentum_turnover_adj,
        'base_warmup': 0,
        'window_multiplier': 1,
        'windows': None,
        'requires_constituent': False,
        'description': '换手率调整动量因子（量价背离）'
    },
    'momentum_price_volume_icir': {
        'func': factor_.momentum_price_volume_icir,
        'base_warmup': 240,
        'window_multiplier': 0,
        'windows': [20],
        'requires_constituent': False,
        'description': '量价清洗ICIR加权动量（需要amount_df）'
    },
    'momentum_rebound_with_crowding_filter': {
        'func': factor_.momentum_rebound_with_crowding_filter,
        'base_warmup': 0,
        'window_multiplier': 1,
        'windows': None,
        'requires_constituent': False,
        'description': '反弹动量因子（综合动量+拥挤度过滤）'
    },
    'momentum_amplitude_cut': {
        'func': factor_.momentum_amplitude_cut,
        'base_warmup': 0,
        'window_multiplier': 1,
        'windows': None,
        'requires_constituent': False,
        'description': '振幅切割动量（需要high_df, low_df）'
    },
    
    # ===== 特质收益动量因子 =====
    'momentum_pure_liquidity_stripped': {
        'func': factor_.momentum_pure_liquidity_stripped,
        'base_warmup': 240,
        'window_multiplier': 1,
        'windows': None,
        'requires_constituent': False,
        'description': '剥离流动性提纯动量因子'
    },
    'momentum_residual': {
        'func': factor_.momentum_residual,
        'base_warmup': 0,
        'window_multiplier': 1,
        'windows': None,
        'requires_constituent': False,
        'description': '行业残差动量因子（需要Barra因子数据）'
    },
    
    # ===== 行业间相关性动量因子 =====
    'momentum_cross_industry_lasso': {
        'func': factor_.momentum_cross_industry_lasso,
        'base_warmup': 200,
        'window_multiplier': 1,
        'windows': None,
        'requires_constituent': False,
        'description': 'Lasso因子（行业间相关性）'
    },
    
    # ===== 行业内关系动量因子（需要成分股数据）=====
    'momentum_industry_component': {
        'func': factor_.momentum_industry_component,
        'base_warmup': 0,
        'window_multiplier': 1,
        'windows': [240],
        'requires_constituent': True,
        'description': '行业成分股动量因子（一致性上涨）'
    },
    'momentum_pca': {
        'func': factor_.momentum_pca,
        'base_warmup': 60,
        'window_multiplier': 1,
        'windows': [20],
        'requires_constituent': True,
        'description': 'PcaMom技术面内生动量因子'
    },
    'momentum_lead_lag_enhanced': {
        'func': factor_.momentum_lead_lag_enhanced,
        'base_warmup': 0,
        'window_multiplier': 1,
        'windows': None,
        'requires_constituent': True,
        'description': '龙头领先特征修正后的动量增强因子'
    },
}

# 兼容性：保留 FACTOR_REGISTRY 和 FACTOR_WARMUP_CONFIG 作为视图
FACTOR_REGISTRY = {name: config['func'] for name, config in FACTOR_CONFIG.items()}
FACTOR_WARMUP_CONFIG = {
    name: {
        'base_warmup': config['base_warmup'],
        'window_multiplier': config['window_multiplier'],
        'windows': config['windows']
    }
    for name, config in FACTOR_CONFIG.items()
}

class DataContainer:
    """
    数据容器类，统一管理所有数据
    一次性加载所有数据，避免重复读取文件
    """
    def __init__(self, file_path=DEFAULT_CACHE_FILE, start_date=None, end_date=None, 
                 exclude_incomplete_month=True, load_constituent=False):
        """
        初始化数据容器，加载所有数据
        
        参数:
        file_path: str, 数据文件路径
        start_date: str, 数据开始日期 (YYYY-MM-DD)
        end_date: str, 数据结束日期 (YYYY-MM-DD)
        exclude_incomplete_month: bool, 是否排除最后一个不完整的月份（默认True）
        load_constituent: bool, 是否加载成分股数据（默认False，需要时设为True）
        """
        print("正在加载所有数据...")
        self.file_path = file_path
        self.exclude_incomplete_month = exclude_incomplete_month
        self.load_constituent = load_constituent
        
        # 加载行业指数数据
        self.prices_df = load_price_df(file_path)
        self.high_df = load_high_df(file_path)
        self.low_df = load_low_df(file_path)
        self.turnover_df = load_turnover_df(file_path)
        self.volume_df = load_volume_df(file_path)
        
        # amount_df 使用 volume_df 作为代理（成交量可近似代表成交额趋势）
        self.amount_df = self.volume_df

        # 成分股相关数据（按需加载）
        self.constituent_df = None
        self.stock_price_df = None
        self.stock_mv_df = None
        self.industry_code_df = None

        # Barra因子数据（用于残差动量因子）
        try:
            self.barra_factor_returns_df = load_barra_factor_returns()
        except FileNotFoundError:
            self.barra_factor_returns_df = None
            print("  - 警告: Barra因子数据文件不存在，momentum_residual因子将不可用")

        if load_constituent:
            self._load_constituent_data()

        # 根据日期范围筛选数据
        if start_date:
            self.prices_df = self.prices_df[self.prices_df.index >= start_date]
            self.high_df = self.high_df[self.high_df.index >= start_date]
            self.low_df = self.low_df[self.low_df.index >= start_date]
            self.turnover_df = self.turnover_df[self.turnover_df.index >= start_date]
            self.volume_df = self.volume_df[self.volume_df.index >= start_date]
            self.amount_df = self.amount_df[self.amount_df.index >= start_date]
            # 成分股数据也需要筛选
            if self.stock_price_df is not None:
                self.stock_price_df = self.stock_price_df[self.stock_price_df.index >= start_date]
            if self.stock_mv_df is not None:
                self.stock_mv_df = self.stock_mv_df[self.stock_mv_df.index >= start_date]
            # Barra因子数据筛选
            if self.barra_factor_returns_df is not None:
                self.barra_factor_returns_df = self.barra_factor_returns_df[self.barra_factor_returns_df.index >= start_date]
        if end_date:
            self.prices_df = self.prices_df[self.prices_df.index <= end_date]
            self.high_df = self.high_df[self.high_df.index <= end_date]
            self.low_df = self.low_df[self.low_df.index <= end_date]
            self.turnover_df = self.turnover_df[self.turnover_df.index <= end_date]
            self.volume_df = self.volume_df[self.volume_df.index <= end_date]
            self.amount_df = self.amount_df[self.amount_df.index <= end_date]
            if self.stock_price_df is not None:
                self.stock_price_df = self.stock_price_df[self.stock_price_df.index <= end_date]
            if self.stock_mv_df is not None:
                self.stock_mv_df = self.stock_mv_df[self.stock_mv_df.index <= end_date]
            # Barra因子数据筛选
            if self.barra_factor_returns_df is not None:
                self.barra_factor_returns_df = self.barra_factor_returns_df[self.barra_factor_returns_df.index <= end_date]
        
        # 记录原始数据的最后日期
        self.original_last_date = self.prices_df.index[-1] if len(self.prices_df) > 0 else None
        self.last_complete_month_end = None
        
        # 排除最后一个不完整的月份（用于月度调仓）
        if exclude_incomplete_month and len(self.prices_df) > 0:
            self._exclude_incomplete_month()
        
        print(f"数据加载完成:")
        print(f"  - 价格数据: {self.prices_df.shape}")
        print(f"  - 最高价数据: {self.high_df.shape}")
        print(f"  - 最低价数据: {self.low_df.shape}")
        print(f"  - 换手率数据: {self.turnover_df.shape}")
        print(f"  - 成交量数据: {self.volume_df.shape}")
        print(f"  - 日期范围: {self.prices_df.index[0].date()} 至 {self.prices_df.index[-1].date()}")
        
        if load_constituent:
            print(f"  - 成分股数据: 已加载")
            if self.stock_price_df is not None:
                print(f"    - 个股价格: {self.stock_price_df.shape}")
            if self.stock_mv_df is not None:
                print(f"    - 个股市值: {self.stock_mv_df.shape}")
        
        if self.last_complete_month_end and self.original_last_date:
            if self.last_complete_month_end != self.original_last_date:
                print(f"  - 注意: 原始数据截止到 {self.original_last_date.date()}，已排除不完整月份")
                print(f"  - 回测数据截止到 {self.last_complete_month_end.date()}（最后一个完整月末）")
    
    def _load_constituent_data(self):
        """
        加载成分股相关数据
        """
        print("  正在加载成分股数据...")
        try:
            self.constituent_df = load_constituent_df()
            self.stock_price_df = load_stock_price_df()
            self.stock_mv_df = load_stock_mv_df()
            self.industry_code_df = load_industry_code_df()
        except FileNotFoundError as e:
            print(f"  警告: 成分股数据加载失败 - {e}")
            print(f"  需要成分股数据的因子将无法计算")
    
    def has_constituent_data(self):
        """
        检查是否已加载成分股数据
        """
        return (self.constituent_df is not None and 
                self.stock_price_df is not None and 
                self.stock_mv_df is not None and 
                self.industry_code_df is not None)
    
    def _exclude_incomplete_month(self):
        """
        排除最后一个不完整的月份
        """
        from factors_analysis import get_last_complete_month_end
        
        # 获取最后一个完整月份的月末日期
        self.last_complete_month_end = get_last_complete_month_end(self.prices_df.index)
        
        if self.last_complete_month_end is None:
            print("警告: 无法确定完整月份，使用全部数据")
            return
        
        # 如果最后一个完整月末与数据最后日期相同，说明数据是完整的
        if self.last_complete_month_end == self.prices_df.index[-1]:
            return
        
        # 截断数据到最后一个完整月末
        # 注意：这里不截断原始数据，因为因子计算可能需要更多历史数据
        # 截断操作在回测时进行


def get_factor_docstring(factor_name):
    """
    获取因子函数的docstring作为因子说明
    
    参数:
    factor_name: str, 因子名称
    
    返回:
    str: 因子说明文档
    """
    if factor_name not in FACTOR_CONFIG:
        return "未找到因子说明"
    
    func = FACTOR_CONFIG[factor_name]['func']
    docstring = inspect.getdoc(func)
    return docstring if docstring else "无说明文档"


def get_factor_config(factor_name):
    """
    获取因子的完整配置信息
    
    参数:
    factor_name: str, 因子名称
    
    返回:
    dict: 因子配置，包含 func, base_warmup, window_multiplier, windows, requires_constituent, description
    """
    if factor_name not in FACTOR_CONFIG:
        return None
    return FACTOR_CONFIG[factor_name]


def print_factor_config_summary():
    """
    打印因子配置摘要
    """
    print("=" * 80)
    print("因子配置摘要")
    print("=" * 80)
    print(f"{'因子名称':<40} {'需要成分股':<12} {'预热期配置':<20}")
    print("-" * 80)
    for name, config in FACTOR_CONFIG.items():
        requires = "是" if config.get('requires_constituent', False) else "否"
        warmup_info = f"base={config['base_warmup']}, mult={config['window_multiplier']}"
        print(f"{name:<40} {requires:<12} {warmup_info:<20}")
    print("=" * 80)
    print(f"总计: {len(FACTOR_CONFIG)} 个因子")
    constituent_count = sum(1 for c in FACTOR_CONFIG.values() if c.get('requires_constituent', False))
    print(f"需要成分股数据的因子: {constituent_count} 个")


def calculate_factor_warmup_period(factor_name, window):
    """
    计算单个因子在指定窗口下的预热期（交易日数）
    
    参数:
    factor_name: str, 因子名称
    window: int, 因子计算窗口
    
    返回:
    int: 预热期（交易日数）
    """
    if factor_name not in FACTOR_WARMUP_CONFIG:
        # 默认只需要 window 天
        return window
    
    config = FACTOR_WARMUP_CONFIG[factor_name]
    base_warmup = config.get('base_warmup', 0)
    window_multiplier = config.get('window_multiplier', 1)
    
    return base_warmup + window * window_multiplier


def get_factor_windows(factor_name, default_windows=WINDOWS):
    """
    获取因子适用的窗口列表
    
    参数:
    factor_name: str, 因子名称
    default_windows: list, 默认窗口列表
    
    返回:
    list: 该因子适用的窗口列表
    """
    if factor_name not in FACTOR_WARMUP_CONFIG:
        return default_windows
    
    config = FACTOR_WARMUP_CONFIG[factor_name]
    factor_windows = config.get('windows', None)
    
    if factor_windows is None:
        return default_windows
    
    return factor_windows


def calculate_factor_unified_start_date(data: 'DataContainer', factor_name, windows,
                                         rebalance_freq=DEFAULT_REBALANCE_FREQ,
                                         monthly_rebalance=DEFAULT_MONTHLY_REBALANCE):
    """
    计算单个因子在所有窗口下的统一回测起始日期
    
    选择该因子在所有窗口中预热期最长的那个，然后找到该日期之后的第一个月末作为统一起始日期
    
    参数:
    data: DataContainer, 数据容器
    factor_name: str, 因子名称
    windows: list, 窗口列表
    rebalance_freq: int, 调仓频率
    monthly_rebalance: bool, 是否按月调仓
    
    返回:
    pd.Timestamp: 该因子的统一回测起始日期（月末日期）
    """
    all_dates = data.prices_df.index
    
    # 计算该因子在所有窗口下的最大预热期
    max_warmup = 0
    max_warmup_window = None
    
    for window in windows:
        warmup = calculate_factor_warmup_period(factor_name, window)
        if warmup > max_warmup:
            max_warmup = warmup
            max_warmup_window = window
    
    # 确保有足够的数据
    if max_warmup >= len(all_dates):
        raise ValueError(f"因子 {factor_name} 数据不足：需要至少 {max_warmup} 个交易日，但只有 {len(all_dates)} 个")
    
    # 找到预热期结束后的第一个日期
    warmup_end_date = all_dates[max_warmup]
    
    if monthly_rebalance:
        # 找到预热期结束后的第一个月末日期
        monthly_dates = get_monthly_rebalance_dates(all_dates)
        # 找到大于等于 warmup_end_date 的第一个月末
        valid_monthly_dates = monthly_dates[monthly_dates >= warmup_end_date]
        
        if len(valid_monthly_dates) == 0:
            raise ValueError(f"因子 {factor_name} 预热期结束后没有有效的月末日期")
        
        unified_start_date = valid_monthly_dates[0]
    else:
        unified_start_date = warmup_end_date
    
    return unified_start_date, max_warmup, max_warmup_window


def compute_factor(factor_name, data: DataContainer, window, rebalance_freq=DEFAULT_REBALANCE_FREQ):
    """
    计算单个因子值
    
    参数:
    factor_name: str, 因子名称
    data: DataContainer, 数据容器（包含所有需要的数据）
    window: int, 回溯窗口
    rebalance_freq: int, 调仓频率
    
    返回:
    pd.DataFrame: 因子值
    """
    if factor_name not in FACTOR_REGISTRY:
        raise ValueError(f"未知因子名称: {factor_name}. 可选: {list(FACTOR_REGISTRY.keys())}")
    
    # 检查是否需要成分股数据
    factor_config = FACTOR_CONFIG.get(factor_name, {})
    if factor_config.get('requires_constituent', False) and not data.has_constituent_data():
        raise ValueError(f"因子 '{factor_name}' 需要成分股数据，但 DataContainer 未加载成分股数据。"
                        f"请使用 DataContainer(load_constituent=True) 初始化。")
    
    factor_func = FACTOR_REGISTRY[factor_name]
    sig = inspect.signature(factor_func)
    param_names = list(sig.parameters.keys())
    
    # 计算基准收益率（用于需要的因子）
    benchmark_returns = calculate_benchmark_returns(data.prices_df, rebalance_freq)

    # 计算行业日收益率（用于残差动量因子）
    industry_returns_df = data.prices_df.pct_change()

    # 构建参数字典 - 包含所有可能需要的参数
    available_params = {
        # 数据参数
        'prices_df': data.prices_df,
        'high_df': data.high_df,
        'low_df': data.low_df,
        'turnover_df': data.turnover_df,
        'volume_df': data.volume_df,
        'amount_df': data.amount_df,  # 使用volume_df作为代理
        # 成分股数据参数
        'constituent_df': data.constituent_df,
        'stock_price_df': data.stock_price_df,
        'stock_mv_df': data.stock_mv_df,
        'industry_code_df': data.industry_code_df,
        'industry_code_file': DEFAULT_INDUSTRY_CODE_FILE,
        # 窗口参数（所有因子统一使用 window）
        'window': window,
        'rebalance_freq': rebalance_freq,
        # 其他参数
        'zscore_window': 240,
        'smooth_window': 3,
        'min_industries': 15,
        'train_periods': None,
        'benchmark_returns': benchmark_returns,
        # PCA因子参数
        'pca_window': 60,
        'lag': 20,
        # 残差动量因子参数
        'industry_returns_df': industry_returns_df,
        'barra_factor_returns_df': data.barra_factor_returns_df,
    }
    
    # 根据函数签名自动选择参数
    call_kwargs = {}
    for param_name in param_names:
        if param_name in available_params:
            call_kwargs[param_name] = available_params[param_name]
        else:
            param = sig.parameters[param_name]
            if param.default is inspect.Parameter.empty:
                raise ValueError(f"因子 '{factor_name}' 需要参数 '{param_name}'，但未定义。")
    
    return factor_func(**call_kwargs)


def calculate_benchmark_returns(prices_df, rebalance_freq):
    """
    计算基准收益率（申万一级行业等权）
    
    参数:
    prices_df: pd.DataFrame, 价格数据
    rebalance_freq: int, 调仓频率
    
    返回:
    pd.Series: 基准收益率序列（每期）
    """
    # 计算每个行业的期收益率
    period_returns = prices_df.pct_change(rebalance_freq)
    # 等权平均作为基准
    benchmark_returns = period_returns.mean(axis=1)
    return benchmark_returns


def get_monthly_rebalance_dates(all_trade_dates):
    """
    根据给定的所有交易日期，生成每月最后一个交易日作为调仓日期。
    
    参数:
    all_trade_dates: pd.DatetimeIndex, 所有的交易日期序列
    
    返回:
    pd.DatetimeIndex: 每月最后一个交易日序列
    """
    # 确保日期是DatetimeIndex类型
    if not isinstance(all_trade_dates, pd.DatetimeIndex):
        all_trade_dates = pd.to_datetime(all_trade_dates)

    # 按照年月分组，取每个月的最后一个日期
    # 由于all_trade_dates已经是交易日序列，直接取每个月最后一个日期就是该月的最后一个交易日
    monthly_last_dates = all_trade_dates.to_series().groupby([all_trade_dates.year, all_trade_dates.month]).apply(lambda x: x.iloc[-1])
    return pd.DatetimeIndex(monthly_last_dates)


def get_last_complete_month_end(all_trade_dates):
    """
    获取最后一个完整月份的月末日期。
    
    判断逻辑：如果数据的最后一个日期不是该月的最后一个交易日，
    则认为该月不完整，返回上一个月的月末日期。
    
    参数:
    all_trade_dates: pd.DatetimeIndex, 所有的交易日期序列
    
    返回:
    pd.Timestamp: 最后一个完整月份的月末日期
    """
    if not isinstance(all_trade_dates, pd.DatetimeIndex):
        all_trade_dates = pd.to_datetime(all_trade_dates)
    
    if len(all_trade_dates) == 0:
        return None
    
    # 获取所有月末日期
    monthly_ends = get_monthly_rebalance_dates(all_trade_dates)
    
    # 数据的最后一个日期
    last_date = all_trade_dates[-1]
    
    # 检查最后一个日期是否是月末
    # 如果最后一个日期在月末列表中，说明该月是完整的
    if last_date in monthly_ends.values:
        return last_date
    else:
        # 最后一个月不完整，返回倒数第二个月末（如果存在）
        # 找到所有小于 last_date 的月末日期
        complete_month_ends = monthly_ends[monthly_ends < last_date]
        if len(complete_month_ends) > 0:
            return complete_month_ends[-1]
        else:
            return None


def is_month_complete(all_trade_dates, check_date):
    """
    检查指定月份是否完整（即该月的最后一个交易日是否在数据中）
    
    参数:
    all_trade_dates: pd.DatetimeIndex, 所有的交易日期序列
    check_date: pd.Timestamp, 要检查的日期
    
    返回:
    bool: 该月是否完整
    """
    if not isinstance(all_trade_dates, pd.DatetimeIndex):
        all_trade_dates = pd.to_datetime(all_trade_dates)
    
    # 获取该月的所有交易日
    year, month = check_date.year, check_date.month
    month_dates = all_trade_dates[(all_trade_dates.year == year) & (all_trade_dates.month == month)]
    
    if len(month_dates) == 0:
        return False
    
    # 获取该月在数据中的最后一个交易日
    last_trade_day_in_data = month_dates[-1]
    
    # 检查这是否是该月的最后一个交易日
    # 方法：检查下一个交易日是否在下个月
    last_date_idx = all_trade_dates.get_loc(last_trade_day_in_data)
    
    # 如果是数据的最后一天，需要判断是否是月末
    if last_date_idx == len(all_trade_dates) - 1:
        # 检查日期是否接近月末（最后5个自然日内）
        import calendar
        _, last_day = calendar.monthrange(year, month)
        return last_trade_day_in_data.day >= last_day - 5
    else:
        # 检查下一个交易日是否在下个月
        next_trade_day = all_trade_dates[last_date_idx + 1]
        return next_trade_day.month != month or next_trade_day.year != year


def calculate_monthly_forward_returns(prices_df, all_trade_dates):
    """
    计算月度未来收益率（从当前月末到下一个月末的收益率）
    
    参数:
    prices_df: pd.DataFrame, 价格数据
    all_trade_dates: pd.DatetimeIndex, 所有交易日期
    
    返回:
    pd.DataFrame: 月度未来收益率 (index=日期, columns=行业)
    """
    # 获取每月最后一个交易日
    monthly_dates = get_monthly_rebalance_dates(all_trade_dates)
    
    # 计算每个月末到下一个月末的收益率
    forward_returns = pd.DataFrame(index=prices_df.index, columns=prices_df.columns, dtype=float)
    
    for i, date in enumerate(monthly_dates[:-1]):
        next_date = monthly_dates[i + 1]
        if date in prices_df.index and next_date in prices_df.index:
            # 计算从当前月末到下一个月末的收益率
            ret = (prices_df.loc[next_date] / prices_df.loc[date]) - 1
            forward_returns.loc[date] = ret
    
    return forward_returns


def calc_rank_ic(factor_series, return_series):
    """
    计算Rank IC (Spearman相关系数)
    
    参数:
        factor_series: pd.Series, 因子值序列
        return_series: pd.Series, 收益率序列
    
    返回:
        float, Rank IC值
    """
    valid_mask = factor_series.notna() & return_series.notna()
    if valid_mask.sum() < 3:
        return np.nan
    
    factor_valid = factor_series[valid_mask]
    return_valid = return_series[valid_mask]
    
    ic, _ = stats.spearmanr(factor_valid, return_valid)
    return ic


def calculate_ic_ir(factor_df, forward_returns_df, rebalance_freq=DEFAULT_REBALANCE_FREQ, monthly_rebalance=DEFAULT_MONTHLY_REBALANCE, unified_start_date=None):
    """
    计算因子的 IC 和 IR
    
    参数:
        factor_df: pd.DataFrame, 因子值 (index=日期, columns=行业)
        forward_returns_df: pd.DataFrame, 未来收益率 (index=日期, columns=行业)
        rebalance_freq: int, 调仓频率（交易日），默认20。如果 monthly_rebalance 为 True，此参数无效。
        monthly_rebalance: bool, 是否按月调仓，默认为 True (每月最后一个交易日调仓)
        unified_start_date: pd.Timestamp, 统一的回测起始日期（如果指定，只计算该日期之后的IC）
    
    返回:
        tuple: (ic_series, ic_cumsum, ic_mean, ic_std, icir, ic_win_rate, ic_abs_mean)
    """
    all_dates = factor_df.index
    
    if monthly_rebalance:
        # 按月调仓
        rebalance_dates = get_monthly_rebalance_dates(all_dates)
    else:
        # 按固定频率调仓
        rebalance_indices = list(range(0, len(all_dates), rebalance_freq))
        rebalance_dates = all_dates[rebalance_indices]
    
    # 如果指定了统一起始日期，过滤调仓日期
    if unified_start_date is not None:
        rebalance_dates = rebalance_dates[rebalance_dates >= unified_start_date]
    
    # 计算每个调仓日的IC
    ic_list = []
    ic_dates = []
    
    for date in rebalance_dates:
        if date not in factor_df.index or date not in forward_returns_df.index:
            continue
        
        ic = calc_rank_ic(factor_df.loc[date], forward_returns_df.loc[date])
        
        if not np.isnan(ic):
            ic_list.append(ic)
            ic_dates.append(date)
    
    # 构建IC时间序列
    ic_series = pd.Series(ic_list, index=ic_dates)
    
    # 计算IC累积序列
    ic_cumsum = ic_series.cumsum()
    
    # 计算统计指标
    ic_mean = ic_series.mean() if len(ic_series) > 0 else np.nan
    ic_std = ic_series.std() if len(ic_series) > 0 else np.nan
    icir = ic_mean / ic_std if ic_std > 0 else np.nan
    ic_win_rate = (ic_series > 0).sum() / len(ic_series) if len(ic_series) > 0 else np.nan
    ic_abs_mean = ic_series.abs().mean() if len(ic_series) > 0 else np.nan
    
    return ic_series, ic_cumsum, ic_mean, ic_std, icir, ic_win_rate, ic_abs_mean


def get_latest_month_holdings(factor_df, prices_df, window, n_layers=N_LAYERS, monthly_rebalance=DEFAULT_MONTHLY_REBALANCE):
    """
    获取最新月份的分层持仓（包括不完整月份）
    用于输出最新的选股结果，即使该月份还没有完整的收益数据
    
    参数:
    factor_df: pd.DataFrame, 因子值 (index=日期, columns=行业)
    prices_df: pd.DataFrame, 价格数据
    window: int, 回溯窗口
    n_layers: int, 分层数
    monthly_rebalance: bool, 是否按月调仓
    
    返回:
    dict: {layer_idx: {date: [行业列表]}} 最新月份的持仓
    """
    all_dates = factor_df.index
    
    if not monthly_rebalance:
        return {}
    
    # 获取所有月末日期（包括不完整月份）
    all_month_ends = get_monthly_rebalance_dates(all_dates)
    
    if len(all_month_ends) == 0:
        return {}
    
    # 获取最新的月末日期
    latest_month_end = all_month_ends[-1]
    
    # 确保该日期在因子数据中存在
    if latest_month_end not in factor_df.index:
        return {}
    
    # 获取该日期的因子值
    fac = factor_df.loc[latest_month_end]
    valid_mask = ~fac.isna()
    fac = fac[valid_mask]
    
    if len(fac) < n_layers * 2:
        return {}
    
    # 按因子值排序并分层
    sorted_assets = fac.sort_values(ascending=True)
    n_assets = len(sorted_assets)
    layer_size = n_assets // n_layers
    
    latest_holdings = {i: {} for i in range(n_layers)}
    
    for layer in range(n_layers):
        start_idx = layer * layer_size
        end_idx = start_idx + layer_size if layer < n_layers - 1 else n_assets
        layer_assets = sorted_assets.index[start_idx:end_idx].tolist()
        latest_holdings[layer][latest_month_end] = layer_assets
    
    return latest_holdings


def stratified_backtest(factor_df, prices_df, window, rebalance_freq=DEFAULT_REBALANCE_FREQ, n_layers=N_LAYERS, monthly_rebalance=DEFAULT_MONTHLY_REBALANCE, last_complete_month_end=None, unified_start_date=None):
    """
    分层回测：根据因子值将资产分成n层，计算每层的收益率累计净值
    同时返回每期各层选中的行业
    
    参数:
    factor_df: pd.DataFrame, 因子值 (index=日期, columns=行业)
    prices_df: pd.DataFrame, 价格数据
    window: int, 回溯窗口（因子计算的回溯窗口）
    rebalance_freq: int, 调仓频率（交易日），默认20。如果 monthly_rebalance 为 True，此参数无效。
    n_layers: int, 分层数
    monthly_rebalance: bool, 是否按月调仓，默认为 True (每月最后一个交易日调仓)
    last_complete_month_end: pd.Timestamp, 最后一个完整月份的月末日期（用于排除不完整月份的收益计算）
    unified_start_date: pd.Timestamp, 统一的回测起始日期（如果指定，从该日期开始回测）
    
    返回:
    tuple: (nav_df, layer_returns, layer_holdings_history)
    """
    daily_returns = prices_df.pct_change()
    
    all_dates_factor_df = factor_df.index

    if monthly_rebalance:
        # 按月调仓，获取所有月末日期
        rebalance_dates_raw = get_monthly_rebalance_dates(all_dates_factor_df)
        
        # 如果指定了统一起始日期，从该日期开始
        if unified_start_date is not None:
            rebalance_dates_raw = rebalance_dates_raw[rebalance_dates_raw >= unified_start_date]
        
        # 如果指定了最后一个完整月末，排除之后的调仓日期
        if last_complete_month_end is not None:
            rebalance_dates_raw = rebalance_dates_raw[rebalance_dates_raw <= last_complete_month_end]
    else:
        # 按固定频率调仓
        # 如果指定了统一起始日期，从该日期开始
        if unified_start_date is not None:
            start_idx = all_dates_factor_df.get_loc(unified_start_date) if unified_start_date in all_dates_factor_df else 0
            effective_start_dates = all_dates_factor_df[start_idx:]
        else:
            # 默认从 window 位置开始
            start_date_idx = min(window, len(all_dates_factor_df) - 1)
            effective_start_dates = all_dates_factor_df[start_date_idx:]
        
        rebalance_indices = list(range(0, len(effective_start_dates), rebalance_freq))
        rebalance_dates_raw = effective_start_dates[rebalance_indices]
    
    # 确保调仓日期在 prices_df 中存在
    valid_dates = [d for d in rebalance_dates_raw if d in prices_df.index]

    layer_nav = {i: [1.0] for i in range(n_layers)}
    nav_dates = [valid_dates[0]] if valid_dates else []
    layer_holdings_history = {i: {} for i in range(n_layers)}  # 记录每期每层持仓
    
    for i, date in enumerate(valid_dates[:-1]):
        next_date = valid_dates[i + 1]
        
        fac = factor_df.loc[date]
        valid_mask = ~fac.isna()
        fac = fac[valid_mask]
        
        if len(fac) < n_layers * 2:
            for layer in range(n_layers):
                layer_nav[layer].append(layer_nav[layer][-1])
            nav_dates.append(next_date)
            continue
        
        # 按因子值排序并分层（升序：G1=因子值最小，G5=因子值最大）
        sorted_assets = fac.sort_values(ascending=True)
        n_assets = len(sorted_assets)
        layer_size = n_assets // n_layers
        
        holding_period = daily_returns.loc[date:next_date].iloc[1:]
        for layer in range(n_layers):
            start_idx = layer * layer_size
            end_idx = start_idx + layer_size if layer < n_layers - 1 else n_assets
            layer_assets = sorted_assets.index[start_idx:end_idx].tolist()
            
            # 记录持仓
            layer_holdings_history[layer][date] = layer_assets
            
            if layer_assets and len(holding_period) > 0:
                layer_ret = holding_period[layer_assets].mean(axis=1)
                cumulative_ret = (1 + layer_ret).prod() - 1
                layer_nav[layer].append(layer_nav[layer][-1] * (1 + cumulative_ret))
            else:
                layer_nav[layer].append(layer_nav[layer][-1])
        
        nav_dates.append(next_date)
    
    nav_df = pd.DataFrame(layer_nav, index=nav_dates)
    nav_df.columns = [f'G{i+1}' for i in range(n_layers)]
    layer_returns = nav_df.pct_change().dropna()
    
    return nav_df, layer_returns, layer_holdings_history


def calculate_excess_metrics(nav_df, benchmark_nav, rebalance_freq=DEFAULT_REBALANCE_FREQ, monthly_rebalance=DEFAULT_MONTHLY_REBALANCE):
    """
    计算超额收益指标
    
    参数:
    nav_df: pd.DataFrame, 各层净值
    benchmark_nav: pd.Series, 基准净值
    rebalance_freq: int, 调仓频率（仅在非月度调仓时使用）
    monthly_rebalance: bool, 是否按月调仓
    
    返回:
    dict: 各层的超额收益指标
    """
    results = {}
    
    start_date = nav_df.index[0]
    end_date = nav_df.index[-1]
    years = (end_date - start_date).days / 365.25
    # 月度调仓时，每年约12个调仓周期；否则按固定频率计算
    periods_per_year = 12 if monthly_rebalance else 252 / rebalance_freq
    
    # 计算基准收益率序列
    benchmark_returns = benchmark_nav.pct_change().dropna()
    
    for col in nav_df.columns:
        nav = nav_df[col]
        
        # 超额净值 = 策略净值 / 基准净值
        excess_nav = nav / benchmark_nav
        excess_returns = excess_nav.pct_change().dropna()
        
        # 超额累计收益率
        excess_total_return = (excess_nav.iloc[-1] / excess_nav.iloc[0] - 1) * 100
        
        # 超额年化收益率
        excess_total_ratio = excess_nav.iloc[-1] / excess_nav.iloc[0]
        excess_annual_return = (excess_total_ratio ** (1 / years) - 1) * 100 if years > 0 else 0
        
        # 绝对收益指标
        total_return = (nav.iloc[-1] / nav.iloc[0] - 1) * 100
        total_ratio = nav.iloc[-1] / nav.iloc[0]
        annual_return = (total_ratio ** (1 / years) - 1) * 100 if years > 0 else 0
        
        returns = nav.pct_change().dropna()
        volatility = returns.std() * np.sqrt(periods_per_year) * 100
        
        # 夏普比率
        sharpe = (returns.mean() / returns.std() * np.sqrt(periods_per_year)) if returns.std() > 0 else 0
        
        # 最大回撤
        cummax = nav.cummax()
        drawdown = (nav - cummax) / cummax
        max_dd = drawdown.min() * 100
        
        # 多头胜率：策略收益 > 基准收益的期数占比
        aligned_returns = returns.align(benchmark_returns, join='inner')
        strategy_ret = aligned_returns[0]
        bench_ret = aligned_returns[1]
        win_count = (strategy_ret > bench_ret).sum()
        total_periods = len(strategy_ret)
        win_rate = (win_count / total_periods * 100) if total_periods > 0 else 0
        
        results[col] = {
            '累计收益率(%)': total_return,
            '年化收益率(%)': annual_return,
            '超额累计收益率(%)': excess_total_return,
            '超额年化收益率(%)': excess_annual_return,
            '年化波动率(%)': volatility,
            '夏普比率': sharpe,
            '最大回撤(%)': max_dd,
            '多头胜率(%)': win_rate,
        }
    
    return results


def calculate_yearly_returns(nav_df, benchmark_nav, start_year=2017):
    """
    计算每年的收益统计（仅针对G5）
    
    参数:
    nav_df: pd.DataFrame, 各层净值（包含G5列）
    benchmark_nav: pd.Series, 基准净值
    start_year: int, 起始年份
    
    返回:
    pd.DataFrame: 每年的多头收益、超额收益、基准收益
    """
    if 'G5' not in nav_df.columns:
        return pd.DataFrame()
    
    g5_nav = nav_df['G5']
    
    # 获取所有年份
    years = sorted(set(nav_df.index.year))
    years = [y for y in years if y >= start_year]
    
    yearly_data = []
    
    for year in years:
        # 获取该年的数据
        year_mask = nav_df.index.year == year
        year_dates = nav_df.index[year_mask]
        
        if len(year_dates) < 2:
            continue
        
        start_date = year_dates[0]
        end_date = year_dates[-1]
        
        # G5多头收益
        g5_return = (g5_nav.loc[end_date] / g5_nav.loc[start_date] - 1) * 100
        
        # 基准收益
        bench_return = (benchmark_nav.loc[end_date] / benchmark_nav.loc[start_date] - 1) * 100
        
        # 超额收益
        excess_return = g5_return - bench_return
        
        yearly_data.append({
            '年份': year,
            'G5多头收益(%)': round(g5_return, 2),
            '超额收益(%)': round(excess_return, 2),
            '基准收益(%)': round(bench_return, 2),
        })
    
    # 添加全样本统计
    if len(nav_df) >= 2:
        # 寻找大于等于 start_year 的第一个有效日期作为全样本起始日期
        full_sample_start_date_mask = nav_df.index.year >= start_year
        if not full_sample_start_date_mask.any():
            # 如果没有符合条件的年份，则不计算全样本
            return pd.DataFrame(yearly_data)

        first_valid_date_for_full_sample = nav_df.index[full_sample_start_date_mask][0]
        
        # 确保起始日期在g5_nav和benchmark_nav中存在
        if first_valid_date_for_full_sample in g5_nav.index and first_valid_date_for_full_sample in benchmark_nav.index:
            total_g5_return = (g5_nav.iloc[-1] / g5_nav.loc[first_valid_date_for_full_sample] - 1) * 100
            total_bench_return = (benchmark_nav.iloc[-1] / benchmark_nav.loc[first_valid_date_for_full_sample] - 1) * 100
            total_excess_return = total_g5_return - total_bench_return
            
            yearly_data.append({
                '年份': '全样本',
                'G5多头收益(%)': round(total_g5_return, 2),
                '超额收益(%)': round(total_excess_return, 2),
                '基准收益(%)': round(total_bench_return, 2),
            })
        else:
            print(f"警告: 无法在全样本计算中找到起始日期 {first_valid_date_for_full_sample} 的对应数据。")
    
    return pd.DataFrame(yearly_data)


def analyze_single_factor_window(factor_name, data: DataContainer, window, rebalance_freq=DEFAULT_REBALANCE_FREQ, monthly_rebalance=DEFAULT_MONTHLY_REBALANCE, unified_start_date=None):
    """
    分析单个因子在单个窗口下的表现
    
    参数:
    factor_name: str, 因子名称
    data: DataContainer, 数据容器
    window: int, 因子计算的回溯窗口
    rebalance_freq: int, 调仓频率（仅在非月度调仓时使用）
    monthly_rebalance: bool, 是否按月调仓，默认为 True
    unified_start_date: pd.Timestamp, 统一的回测起始日期（如果指定，所有因子从该日期开始回测）
    
    返回:
    dict: 包含IC/IR、分层指标、持仓等信息
    """
    # 计算因子值
    factor_df = compute_factor(factor_name, data, window, rebalance_freq)
    
    # 获取最后一个完整月末日期（用于排除不完整月份的收益计算）
    last_complete_month_end = None
    if monthly_rebalance and hasattr(data, 'last_complete_month_end'):
        last_complete_month_end = data.last_complete_month_end
    
    # 计算未来收益率
    if monthly_rebalance:
        # 月度调仓：计算从当前月末到下一个月末的收益率
        forward_returns_df = calculate_monthly_forward_returns(data.prices_df, data.prices_df.index)
    else:
        # 固定频率调仓：使用固定周期的未来收益率
        forward_returns_df = data.prices_df.pct_change(rebalance_freq).shift(-rebalance_freq)
    
    # 计算IC/IR（包含IC累积序列）
    # 注意：IC计算需要完整的未来收益率，不完整月份的IC会自动被排除（因为forward_returns为NaN）
    ic_series, ic_cumsum, ic_mean, ic_std, icir, ic_win_rate, ic_abs_mean = calculate_ic_ir(
        factor_df, forward_returns_df, rebalance_freq=rebalance_freq, monthly_rebalance=monthly_rebalance,
        unified_start_date=unified_start_date
    )
    
    # 分层回测（传入最后一个完整月末日期，排除不完整月份的收益计算）
    nav_df, layer_returns, layer_holdings = stratified_backtest(
        factor_df, data.prices_df, window, rebalance_freq=rebalance_freq, monthly_rebalance=monthly_rebalance,
        last_complete_month_end=last_complete_month_end, unified_start_date=unified_start_date
    )
    
    # 获取最新月份的持仓（包括不完整月份，用于输出最新选股结果）
    if monthly_rebalance:
        latest_holdings = get_latest_month_holdings(factor_df, data.prices_df, window, N_LAYERS, monthly_rebalance)
        # 合并最新持仓到 layer_holdings（如果最新月份不在回测持仓中）
        for layer_idx, holdings in latest_holdings.items():
            for date, assets in holdings.items():
                if date not in layer_holdings.get(layer_idx, {}):
                    if layer_idx not in layer_holdings:
                        layer_holdings[layer_idx] = {}
                    layer_holdings[layer_idx][date] = assets
    
    # 计算基准净值（等权行业指数）
    # 对齐到调仓日期
    benchmark_nav = pd.Series(index=nav_df.index, dtype=float)
    benchmark_nav.iloc[0] = 1.0
    for i in range(1, len(nav_df.index)):
        prev_date = nav_df.index[i-1]
        curr_date = nav_df.index[i]
        # 计算期间基准收益（所有行业等权平均）
        period_ret = (data.prices_df.loc[curr_date] / data.prices_df.loc[prev_date] - 1).mean()
        benchmark_nav.iloc[i] = benchmark_nav.iloc[i-1] * (1 + period_ret)
    
    # 计算超额指标
    excess_metrics = calculate_excess_metrics(nav_df, benchmark_nav, rebalance_freq, monthly_rebalance)
    
    # 计算每年收益统计（仅针对G5）
    yearly_returns = calculate_yearly_returns(nav_df, benchmark_nav, start_year=2017)
    
    return {
        'ic_mean': ic_mean,
        'icir': icir,
        'ic_win_rate': ic_win_rate,
        'ic_abs_mean': ic_abs_mean,
        'ic_series': ic_series,
        'ic_cumsum': ic_cumsum,
        'nav_df': nav_df,
        'benchmark_nav': benchmark_nav,
        'excess_metrics': excess_metrics,
        'layer_holdings': layer_holdings,
        'yearly_returns': yearly_returns,
    }


def analyze_all_factors(data: DataContainer, windows=WINDOWS, rebalance_freq=DEFAULT_REBALANCE_FREQ, monthly_rebalance=DEFAULT_MONTHLY_REBALANCE, use_unified_start_date=True):
    """
    分析所有因子在所有窗口下的表现
    
    参数:
    data: DataContainer, 数据容器
    windows: list, 窗口列表（因子计算的回溯窗口，作为默认值）
    rebalance_freq: int, 调仓频率（仅在非月度调仓时使用）
    monthly_rebalance: bool, 是否按月调仓，默认为 True (每月最后一个交易日调仓)
    use_unified_start_date: bool, 是否使用统一的回测起始日期，默认为 True
        - True: 每个因子内部不同窗口使用该因子的最大预热期对应的统一起始日期
        - False: 每个窗口使用各自的预热期
    
    返回:
    dict: {factor_name: {window: analysis_result}}
    """
    all_results = {}
    
    # 打印预热期分析
    if use_unified_start_date:
        print("\n" + "=" * 60)
        print("各因子预热期分析（每个因子内部不同窗口使用统一起始日期）")
        print("=" * 60)
    
    for factor_name in FACTOR_REGISTRY.keys():
        # 获取该因子适用的窗口列表
        factor_windows = get_factor_windows(factor_name, windows)
        
        print(f"\n正在分析因子: {factor_name}")
        print(f"  适用窗口: {factor_windows}")
        factor_results = {}
        
        # 为每个因子单独计算其统一起始日期（使用该因子适用的窗口）
        factor_unified_start_date = None
        if use_unified_start_date:
            try:
                factor_unified_start_date, max_warmup, max_warmup_window = calculate_factor_unified_start_date(
                    data, factor_name, factor_windows, rebalance_freq, monthly_rebalance
                )
                print(f"  预热期: {max_warmup}天 (来自{max_warmup_window}日窗口)")
                print(f"  统一起始日期: {factor_unified_start_date.strftime('%Y-%m-%d')}")
            except Exception as e:
                print(f"  计算统一起始日期失败: {e}")
                factor_unified_start_date = None
        
        # 只计算该因子适用的窗口
        for window in factor_windows:
            print(f"  窗口: {window}日...")
            try:
                result = analyze_single_factor_window(
                    factor_name, data, window, rebalance_freq, monthly_rebalance=monthly_rebalance,
                    unified_start_date=factor_unified_start_date
                )
                factor_results[window] = result
            except Exception as e:
                print(f"    错误: {e}")
                factor_results[window] = None
        
        all_results[factor_name] = factor_results
    
    return all_results


def find_best_windows(factor_results, top_n=2):
    """
    找到IC和ICIR最大的窗口

    参数:
    factor_results: dict, {window: analysis_result} 单个因子的所有窗口结果
    top_n: int, 返回前N个最优窗口，默认2

    返回:
    dict: {
        'best_ic_windows': [(window, ic_mean), ...],  # IC最大的窗口列表
        'best_icir_windows': [(window, icir), ...],   # ICIR最大的窗口列表
        'best_combined_windows': [window, ...],       # IC和ICIR综合最优的窗口（取并集）
    }
    """
    ic_values = {}
    icir_values = {}

    for window, result in factor_results.items():
        if result is None:
            continue
        ic_mean = result.get('ic_mean', np.nan)
        icir = result.get('icir', np.nan)

        # 使用绝对值比较（因为负IC也可能有效）
        if not np.isnan(ic_mean):
            ic_values[window] = abs(ic_mean)
        if not np.isnan(icir):
            icir_values[window] = abs(icir)

    # 按绝对值排序，找到最大的窗口
    sorted_ic = sorted(ic_values.items(), key=lambda x: x[1], reverse=True)
    sorted_icir = sorted(icir_values.items(), key=lambda x: x[1], reverse=True)

    # 取前top_n个
    best_ic_windows = sorted_ic[:top_n]
    best_icir_windows = sorted_icir[:top_n]

    # 综合最优窗口（IC和ICIR的并集）
    best_windows_set = set()
    for window, _ in best_ic_windows:
        best_windows_set.add(window)
    for window, _ in best_icir_windows:
        best_windows_set.add(window)

    # 返回原始值（非绝对值）用于显示
    best_ic_with_sign = []
    for window, _ in best_ic_windows:
        original_ic = factor_results[window]['ic_mean']
        best_ic_with_sign.append((window, original_ic))

    best_icir_with_sign = []
    for window, _ in best_icir_windows:
        original_icir = factor_results[window]['icir']
        best_icir_with_sign.append((window, original_icir))

    return {
        'best_ic_windows': best_ic_with_sign,
        'best_icir_windows': best_icir_with_sign,
        'best_combined_windows': sorted(list(best_windows_set)),
    }


def create_best_windows_summary(all_results, top_n=2):
    """
    创建所有因子的最优窗口汇总表

    参数:
    all_results: dict, {factor_name: {window: analysis_result}}
    top_n: int, 每个指标取前N个最优窗口

    返回:
    pd.DataFrame: 最优窗口汇总表
    """
    summary_data = []

    for factor_name, factor_results in all_results.items():
        if not factor_results:
            continue

        best_info = find_best_windows(factor_results, top_n)

        # 格式化最优窗口信息
        ic_windows_str = ', '.join([f"{w}日(IC={v:.4f})" for w, v in best_info['best_ic_windows']])
        icir_windows_str = ', '.join([f"{w}日(ICIR={v:.4f})" for w, v in best_info['best_icir_windows']])
        combined_windows_str = ', '.join([f"{w}日" for w in best_info['best_combined_windows']])

        summary_data.append({
            '因子名称': factor_name,
            '最优IC窗口': ic_windows_str,
            '最优ICIR窗口': icir_windows_str,
            '综合最优窗口': combined_windows_str,
        })

    return pd.DataFrame(summary_data)


def create_factor_summary_df(factor_name, factor_results, windows=WINDOWS):
    """
    创建单个因子的汇总DataFrame
    
    格式：
    - 列名为窗口周期（20, 60, 120, 240）
    - 行名为指标名称
    - G5放在最前面，然后是G4, G3, G2, G1
    - IC和ICIR保留4位小数，其他指标保留2位小数
    
    返回:
    pd.DataFrame: 汇总表格（行为指标，列为窗口）
    """
    # 构建数据字典：{窗口: {指标: 值}}
    data_dict = {}
    
    for window in windows:
        result = factor_results.get(window)
        if result is None:
            continue
        
        window_data = {}
        # IC和ICIR保留4位小数
        window_data['IC均值'] = round(result['ic_mean'], 4) if not np.isnan(result['ic_mean']) else np.nan
        window_data['ICIR'] = round(result['icir'], 4) if not np.isnan(result['icir']) else np.nan
        window_data['IC胜率'] = round(result['ic_win_rate'], 4) if not np.isnan(result['ic_win_rate']) else np.nan
        
        # 按G5, G4, G3, G2, G1顺序添加各层指标
        layer_order = ['G5', 'G4', 'G3', 'G2', 'G1']
        for layer_name in layer_order:
            if layer_name in result['excess_metrics']:
                metrics = result['excess_metrics'][layer_name]
                for metric_name, value in metrics.items():
                    row_name = f'{layer_name}_{metric_name}'
                    window_data[row_name] = round(value, 2)
        
        data_dict[window] = window_data
    
    # 转换为DataFrame并转置（行为指标，列为窗口）
    df = pd.DataFrame(data_dict)
    
    # 确保列按窗口顺序排列
    df = df.reindex(columns=windows)
    
    return df


def clean_industry_name(name):
    """
    清理行业名称，删除"（申万）"后缀
    
    参数:
    name: str, 行业名称
    
    返回:
    str: 清理后的行业名称
    """
    if isinstance(name, str):
        return name.replace('（申万）', '').replace('(申万)', '')
    return name


def format_date_index(df):
    """
    格式化DataFrame的日期索引，只保留日期部分（不含时间）
    
    参数:
    df: pd.DataFrame, 带有日期索引的DataFrame
    
    返回:
    pd.DataFrame: 格式化后的DataFrame
    """
    if isinstance(df.index, pd.DatetimeIndex):
        df.index = df.index.strftime('%Y-%m-%d')
    return df


def create_g5_holdings_df(factor_results, windows=WINDOWS):
    """
    创建G5持仓记录DataFrame（按列输出每个窗口的G5持仓）
    
    格式：
    - 列名为窗口周期（20, 60, 120, 240）
    - 行为日期（降序，最新日期在最前面）
    - 值为持仓行业（逗号分隔，不含"（申万）"后缀）
    
    参数:
    factor_results: dict, 因子分析结果 {window: result}
    windows: list, 窗口列表
    
    返回:
    pd.DataFrame: G5持仓记录（行为日期，列为窗口）
    """
    # 收集所有窗口的G5持仓
    holdings_dict = {}
    all_dates = set()
    
    for window in windows:
        result = factor_results.get(window)
        if result is None or result.get('layer_holdings') is None:
            continue
        
        # G5 是 layer_idx=4
        g5_holdings = result['layer_holdings'].get(4, {})
        
        # 转换为 {日期: 持仓行业字符串}，同时清理行业名称
        window_holdings = {}
        for date, industries in g5_holdings.items():
            # 清理行业名称，删除"（申万）"后缀
            cleaned_industries = [clean_industry_name(ind) for ind in industries] if industries else []
            window_holdings[date] = ', '.join(cleaned_industries) if cleaned_industries else ''
            all_dates.add(date)
        
        holdings_dict[window] = window_holdings
    
    if not all_dates:
        return pd.DataFrame()
    
    # 按日期降序排列（最新日期在最前面）
    sorted_dates = sorted(all_dates, reverse=True)
    
    # 构建DataFrame
    data = {}
    for window in windows:
        if window in holdings_dict:
            data[window] = [holdings_dict[window].get(date, '') for date in sorted_dates]
        else:
            data[window] = [''] * len(sorted_dates)
    
    # 格式化日期索引（只保留日期部分）
    formatted_dates = [d.strftime('%Y-%m-%d') if hasattr(d, 'strftime') else str(d) for d in sorted_dates]
    
    df = pd.DataFrame(data, index=formatted_dates)
    df.index.name = '日期'
    
    return df


def create_ic_cumsum_df(factor_results, windows=WINDOWS):
    """
    创建IC累积序列DataFrame
    
    参数:
    factor_results: dict, 因子分析结果 {window: result}
    windows: list, 窗口列表
    
    返回:
    pd.DataFrame: IC累积序列（行为日期，列为窗口）
    """
    ic_cumsum_dict = {}
    all_dates = set()
    
    for window in windows:
        result = factor_results.get(window)
        if result is None or result.get('ic_cumsum') is None:
            continue
        
        ic_cumsum = result['ic_cumsum']
        ic_cumsum_dict[window] = ic_cumsum
        all_dates.update(ic_cumsum.index)
    
    if not all_dates:
        return pd.DataFrame()
    
    # 按日期升序排列
    sorted_dates = sorted(all_dates)
    
    # 构建DataFrame
    data = {}
    for window in windows:
        if window in ic_cumsum_dict:
            data[f'{window}日'] = [ic_cumsum_dict[window].get(date, np.nan) for date in sorted_dates]
        else:
            data[f'{window}日'] = [np.nan] * len(sorted_dates)
    
    # 格式化日期索引（只保留日期部分）
    formatted_dates = [d.strftime('%Y-%m-%d') if hasattr(d, 'strftime') else str(d) for d in sorted_dates]
    
    df = pd.DataFrame(data, index=formatted_dates)
    df.index.name = '日期'
    
    # 保留四位小数
    df = df.round(4)
    
    # 删除含有空值的行，只保留所有窗口都有值的行
    df = df.dropna()
    
    return df


def create_layer_nav_df(factor_results, windows=WINDOWS):
    """
    创建分层累积净值DataFrame（每个窗口的G1-G5净值）
    
    参数:
    factor_results: dict, 因子分析结果 {window: result}
    windows: list, 窗口列表
    
    返回:
    dict: {window: nav_df} 每个窗口的分层净值DataFrame
    """
    layer_nav_dict = {}
    
    for window in windows:
        result = factor_results.get(window)
        if result is None or result.get('nav_df') is None:
            continue
        
        nav_df = result['nav_df'].copy()
        # 添加基准净值列
        if result.get('benchmark_nav') is not None:
            nav_df['基准'] = result['benchmark_nav']
        
        # 保留四位小数
        nav_df = nav_df.round(4)
        
        # 格式化日期索引（只保留日期部分）
        nav_df = format_date_index(nav_df)
        
        layer_nav_dict[window] = nav_df
    
    return layer_nav_dict


def create_g5_yearly_returns_df(factor_results, windows=WINDOWS):
    """
    创建G5每年收益统计DataFrame（仅针对指定窗口，如G5对应的窗口）
    
    参数:
    factor_results: dict, 因子分析结果 {window: result}
    windows: list, 窗口列表
    
    返回:
    dict: {window: yearly_returns_df}
    """
    yearly_dict = {}
    
    for window in windows:
        result = factor_results.get(window)
        if result is None or result.get('yearly_returns') is None:
            continue
        
        yearly_df = result['yearly_returns']
        if not yearly_df.empty:
            yearly_dict[window] = yearly_df
    
    return yearly_dict


def get_data_date_range(factor_results, windows=WINDOWS):
    """
    获取数据日期范围
    
    参数:
    factor_results: dict, 因子分析结果 {window: result}
    windows: list, 窗口列表
    
    返回:
    tuple: (start_date_str, end_date_str) 日期范围字符串
    """
    all_dates = []
    
    for window in windows:
        result = factor_results.get(window)
        if result is None or result.get('nav_df') is None:
            continue
        
        nav_df = result['nav_df']
        if len(nav_df) > 0:
            all_dates.extend(nav_df.index.tolist())
    
    if not all_dates:
        return None, None
    
    start_date = min(all_dates)
    end_date = max(all_dates)
    
    # 格式化日期
    start_str = start_date.strftime('%Y-%m-%d') if hasattr(start_date, 'strftime') else str(start_date)
    end_str = end_date.strftime('%Y-%m-%d') if hasattr(end_date, 'strftime') else str(end_date)
    
    return start_str, end_str


def export_to_excel(all_results, output_file='factors_analysis_report.xlsx', windows=WINDOWS):
    """
    将所有因子分析结果导出到Excel
    每个因子一个sheet页

    参数:
    all_results: dict, 所有因子的分析结果
    output_file: str, 输出文件名
    windows: list, 窗口列表
    """
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 首先创建汇总sheet页，展示所有因子的最优窗口
        print("正在导出最优窗口汇总...")
        best_windows_df = create_best_windows_summary(all_results, top_n=2)
        if not best_windows_df.empty:
            best_windows_df.to_excel(writer, sheet_name='最优窗口汇总', index=False)

        for factor_name, factor_results in all_results.items():
            print(f"正在导出因子: {factor_name}")

            # 获取因子说明
            docstring = get_factor_docstring(factor_name)

            # 获取数据日期范围
            start_date, end_date = get_data_date_range(factor_results, windows)

            # 获取该因子的最优窗口信息
            best_info = find_best_windows(factor_results, top_n=2)

            # 创建汇总表
            summary_df = create_factor_summary_df(factor_name, factor_results, windows)

            # 创建G5持仓记录（按列输出每个窗口）
            g5_holdings_df = create_g5_holdings_df(factor_results, windows)

            # 创建IC累积序列
            ic_cumsum_df = create_ic_cumsum_df(factor_results, windows)

            # 创建分层累积净值
            layer_nav_dict = create_layer_nav_df(factor_results, windows)

            # 创建G5每年收益统计
            g5_yearly_dict = create_g5_yearly_returns_df(factor_results, windows)

            # 写入sheet
            sheet_name = factor_name[:31]  # Excel sheet名最长31字符

            # 写入因子说明标题
            start_row = 0
            title_df = pd.DataFrame({f'【因子说明】': [docstring]})
            title_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
            start_row += 3

            # 写入数据日期范围
            if start_date and end_date:
                date_range_df = pd.DataFrame({f'【数据日期范围】': [f'{start_date} 至 {end_date}']})
                date_range_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
                start_row += 3
            else:
                start_row += 1

            # 写入最优窗口信息
            best_ic_str = ', '.join([f"{w}日(IC={v:.4f})" for w, v in best_info['best_ic_windows']])
            best_icir_str = ', '.join([f"{w}日(ICIR={v:.4f})" for w, v in best_info['best_icir_windows']])
            best_combined_str = ', '.join([f"{w}日" for w in best_info['best_combined_windows']])
            best_windows_info = pd.DataFrame({
                '【最优窗口】': [f'最优IC窗口: {best_ic_str}'],
                '': [f'最优ICIR窗口: {best_icir_str}'],
                ' ': [f'综合最优窗口: {best_combined_str}']
            })
            best_windows_info.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
            start_row += 3

            # 写入汇总表（行为指标，列为窗口，需要写入index）
            if not summary_df.empty:
                # 写入汇总表标题
                header_df = pd.DataFrame({f'【因子汇总指标】': ['']})
                header_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
                start_row += 1
                summary_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=True)
                start_row += len(summary_df) + 3

            # 只对最优窗口输出详细的G5每年收益统计
            best_combined_windows = best_info['best_combined_windows']
            for window in best_combined_windows:
                if window in g5_yearly_dict:
                    yearly_df = g5_yearly_dict[window]
                    # 写入标题行（标记为最优窗口）
                    header_df = pd.DataFrame({f'【G5每年收益统计 - {window}日窗口 ★最优】': ['']})
                    header_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
                    start_row += 1
                    yearly_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
                    start_row += len(yearly_df) + 3

            # 写入IC累积序列
            if not ic_cumsum_df.empty:
                # 写入标题行
                header_df = pd.DataFrame({f'【IC累积序列】': ['']})
                header_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
                start_row += 1
                ic_cumsum_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=True)
                start_row += len(ic_cumsum_df) + 3

            # 只对最优窗口输出分层累积净值
            for window in best_combined_windows:
                if window in layer_nav_dict:
                    nav_df = layer_nav_dict[window]
                    # 写入标题行（标记为最优窗口）
                    header_df = pd.DataFrame({f'【分层累积净值 - {window}日窗口 ★最优】': ['']})
                    header_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
                    start_row += 1
                    nav_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=True)
                    start_row += len(nav_df) + 3

            # 只对最优窗口输出G5持仓记录
            if not g5_holdings_df.empty:
                # 筛选最优窗口的列
                best_cols = [col for col in g5_holdings_df.columns if any(f'{w}日' in str(col) for w in best_combined_windows)]
                if best_cols:
                    g5_holdings_best_df = g5_holdings_df[best_cols]
                    # 写入标题行
                    header_df = pd.DataFrame({f'【G5持仓行业 - 最优窗口】': ['']})
                    header_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
                    start_row += 1
                    g5_holdings_best_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=True)
                    start_row += len(g5_holdings_best_df) + 3

    print(f"\n分析报告已导出到: {output_file}")


def list_factors():
    """列出所有可用因子"""
    return list(FACTOR_REGISTRY.keys())


def format_excel_report(file_path: str):
    """
    调整Excel报告文件格式
    
    功能：
    1. 将每个sheet页的A列设置为左对齐
    2. 将每个sheet页的A列宽度设置为11
    3. 隐藏每个sheet页的19行到52行
    4. 比较B9-E9的值，将最大值所在列的8-51行加粗
    
    参数:
        file_path: str, Excel文件路径
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font
    
    # 加载工作簿
    wb = load_workbook(file_path)
    
    # 遍历所有sheet页
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"正在格式化sheet: {sheet_name}")
        
        # 1. 设置A列宽度为11
        ws.column_dimensions['A'].width = 11
        
        # 2. 设置A列左对齐
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=1)  # A列是第1列
            cell.alignment = Alignment(horizontal='left')
        
        # 3. 隐藏19行到52行
        for row in range(19, 53):  # 19到52行（包含52）
            ws.row_dimensions[row].hidden = True
        
        # 4. 比较B9-E9的值，将最大值所在列的8-51行加粗
        # B=2, C=3, D=4, E=5
        values = {}
        for col in range(2, 6):  # B到E列
            cell_value = ws.cell(row=9, column=col).value
            if cell_value is not None:
                try:
                    values[col] = float(cell_value)
                except (ValueError, TypeError):
                    values[col] = float('-inf')
            else:
                values[col] = float('-inf')
        
        if values:
            max_col = max(values, key=values.get)
            print(f"  B9-E9最大值在第{max_col}列 (值={values[max_col]})")
            
            # 将该列的8-51行加粗
            for row in range(8, 52):  # 8到51行
                cell = ws.cell(row=row, column=max_col)
                # 保留原有字体属性，只修改bold
                if cell.font:
                    new_font = copy(cell.font)
                    new_font = Font(
                        name=cell.font.name,
                        size=cell.font.size,
                        bold=True,
                        italic=cell.font.italic,
                        color=cell.font.color
                    )
                else:
                    new_font = Font(bold=True)
                cell.font = new_font
    
    # 保存文件
    wb.save(file_path)
    print(f"Excel格式调整完成: {file_path}")


if __name__ == "__main__":
    print("=" * 60)
    print("因子批量分析")
    print("=" * 60)
    
    # 配置参数
    MONTHLY_REBALANCE = True  # 使用月度调仓（每月最后一个交易日）
    REBALANCE_FREQ = 20  # 调仓频率（天）- 仅在 MONTHLY_REBALANCE=False 时使用
    WINDOWS_TO_TEST = [20, 60, 120, 240]  # 因子计算的回溯窗口
    OUTPUT_FILE = 'factors_analysis_report.xlsx'

    # 固定日期范围（设为None使用全部数据，与单因子测试.py保持一致）
    start_date = None  # 使用全部数据
    end_date = None    # 最新日期
    
    # 加载所有数据（一次性加载）
    print(f"\n正在加载 {start_date} 至 最新 日期的数据...")
    data = DataContainer(DEFAULT_CACHE_FILE, start_date=start_date, end_date=end_date)
    
    # 显示可用因子
    print(f"\n可用因子列表: {list_factors()}")
    
    # 显示调仓方式
    if MONTHLY_REBALANCE:
        print("\n调仓方式: 每月最后一个交易日调仓")
    else:
        print(f"\n调仓方式: 每 {REBALANCE_FREQ} 个交易日调仓")
    
    # 分析所有因子
    print("\n开始分析所有因子...")
    all_results = analyze_all_factors(
        data,
        windows=WINDOWS_TO_TEST,
        rebalance_freq=REBALANCE_FREQ,
        monthly_rebalance=MONTHLY_REBALANCE
    )
    
    # 导出到Excel
    print("\n正在导出到Excel...")
    export_to_excel(all_results, OUTPUT_FILE, WINDOWS_TO_TEST)
    
    # 格式化Excel报告
    print("\n正在格式化Excel报告...")
    format_excel_report(OUTPUT_FILE)
    
    print("\n分析完成！")
