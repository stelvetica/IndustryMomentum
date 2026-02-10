"""
信号因子回测模块 (factor_signal_backtest.py)

用于回测"无因子值"的信号因子，输出为0/1信号而非连续因子值。
与 factor_value_backtest.py 形成对比：
- factor_value_backtest.py: 有因子值的因子，用于IC/ICIR分析和分层回测
- factor_signal_backtest.py: 无因子值的信号因子，用于信号选股回测

包含的信号因子：
1. 泡沫因子 - 基于BSADF泡沫检测的0/1信号
2. 交集因子 - 合成因子Top N与泡沫信号的交集

输出格式：每个因子一个sheet页，包含：
- 策略概览（策略说明、绩效指标、持仓统计）
- 净值序列
- 历史持仓
- 调仓详情
- 信号矩阵

作者: 量化研究
日期: 2025
"""

import pandas as pd
import numpy as np
import warnings
import os
import sys
import time
from datetime import datetime
from typing import Dict, List, Optional, Tuple

warnings.filterwarnings('ignore')

# 添加当前目录到路径
current_dir = os.path.dirname(os.path.abspath(__file__))
if current_dir not in sys.path:
    sys.path.insert(0, current_dir)

# 添加bubble目录到路径
bubble_dir = os.path.join(current_dir, 'bubble')
if bubble_dir not in sys.path:
    sys.path.insert(0, bubble_dir)


# ============================================================================
# 信号因子配置
# ============================================================================

SIGNAL_FACTOR_CONFIG = {
    # 'bubble_factor': {
    #     'name': '行业泡沫动量1-1-BSADF泡沫BOCD变点因子',
    #     'description': '基于BSADF泡沫检测+BOCD变点检测的0/1信号',
    #     'rebalance_type': 'weekly',  # 周度调仓
    #     'logic': '(Signal_BSADF OR Signal_BOCD) AND (Return_t > 0) → 买入信号',
    # },
    'intersection_factor': {
        'name': '交集信号因子-合成因子TopN∩泡沫信号',
        'description': '合成因子Top N行业与泡沫信号=1行业的交集',
        'rebalance_type': 'monthly',  # 月度调仓
        'logic': '合成因子Top6 ∩ 泡沫信号=1 → 最终持仓',
    },
}


# ============================================================================
# 统一回测起始日期计算（使用DataContainer.first_holding_date）
# ============================================================================

def get_unified_start_date_from_data(data: 'DataContainer') -> pd.Timestamp:
    """
    从DataContainer获取统一的回测起始日期

    使用DataContainer中计算好的first_holding_date，确保与factor_value_backtest.py一致
    按照当前数据应该是从20161230开始首次持仓

    参数:
        data: DataContainer, 数据容器

    返回:
        pd.Timestamp: 统一的回测起始日期
    """
    first_holding_date = getattr(data, 'first_holding_date', None)

    if first_holding_date is None:
        raise ValueError("DataContainer中没有first_holding_date属性，请确保backtest_years参数已设置")

    # 转换为Timestamp
    if isinstance(first_holding_date, str):
        first_holding_ts = pd.Timestamp(first_holding_date)
    else:
        first_holding_ts = first_holding_date

    # 找到小于等于first_holding_date的最近月末交易日
    all_dates = data.prices_df.index
    monthly_dates = get_monthly_rebalance_dates(data.prices_df)
    valid_monthly_dates = monthly_dates[monthly_dates <= first_holding_ts]

    if len(valid_monthly_dates) > 0:
        unified_start_date = valid_monthly_dates[-1]
    else:
        # 如果没有找到，使用first_holding_date之后的第一个月末
        valid_monthly_dates = monthly_dates[monthly_dates >= first_holding_ts]
        if len(valid_monthly_dates) > 0:
            unified_start_date = valid_monthly_dates[0]
        else:
            unified_start_date = first_holding_ts

    return unified_start_date


def get_monthly_rebalance_dates(prices_df: pd.DataFrame) -> pd.DatetimeIndex:
    """获取月末调仓日期"""
    trade_dates = prices_df.index
    monthly_last = trade_dates.to_series().groupby(
        [trade_dates.year, trade_dates.month]
    ).apply(lambda x: x.iloc[-1])
    return pd.DatetimeIndex(monthly_last.values)


def get_weekly_rebalance_dates(prices_df: pd.DataFrame) -> pd.DatetimeIndex:
    """获取周末调仓日期（每周五或该周最后一个交易日）"""
    trade_dates = prices_df.index
    weekly_last = trade_dates.to_series().groupby(
        [trade_dates.isocalendar().year, trade_dates.isocalendar().week]
    ).apply(lambda x: x.iloc[-1])
    return pd.DatetimeIndex(weekly_last.values)


# ============================================================================
# 绩效计算函数（使用交易日年化，与factor_value_backtest.py一致）
# ============================================================================

def calculate_performance_stats(strategy_nav: pd.Series,
                                 benchmark_nav: pd.Series,
                                 rebalance_type: str = 'monthly') -> Dict:
    """
    计算绩效统计指标（使用交易日年化）

    参数:
        strategy_nav: pd.Series, 策略净值
        benchmark_nav: pd.Series, 基准净值
        rebalance_type: str, 调仓类型 ('monthly' 或 'weekly')

    返回:
        Dict: 绩效指标
    """
    # 对齐索引
    common_index = strategy_nav.index.intersection(benchmark_nav.index)
    strategy_nav = strategy_nav.loc[common_index]
    benchmark_nav = benchmark_nav.loc[common_index]

    # 计算收益率
    strategy_returns = strategy_nav.pct_change().dropna()
    benchmark_returns = benchmark_nav.pct_change().dropna()
    excess_returns = strategy_returns - benchmark_returns

    # 根据调仓频率计算年化因子
    num_periods = len(strategy_returns)
    if rebalance_type == 'weekly':
        periods_per_year = 52  # 周频
    else:
        periods_per_year = 12  # 月频
    years = num_periods / periods_per_year

    # 总收益
    total_return = strategy_nav.iloc[-1] / strategy_nav.iloc[0] - 1
    benchmark_total_return = benchmark_nav.iloc[-1] / benchmark_nav.iloc[0] - 1
    excess_total_return = total_return - benchmark_total_return

    # 年化收益
    annual_return = (1 + total_return) ** (1 / years) - 1 if years > 0 else 0
    benchmark_annual_return = (1 + benchmark_total_return) ** (1 / years) - 1 if years > 0 else 0
    excess_annual_return = annual_return - benchmark_annual_return

    # 年化波动率（根据调仓频率）
    annual_vol = strategy_returns.std() * np.sqrt(periods_per_year)
    excess_vol = excess_returns.std() * np.sqrt(periods_per_year)

    # 夏普比率
    sharpe = annual_return / annual_vol if annual_vol > 0 else 0

    # 信息比率
    ir = excess_annual_return / excess_vol if excess_vol > 0 else 0

    # 最大回撤
    cummax = strategy_nav.cummax()
    drawdown = (strategy_nav - cummax) / cummax
    max_drawdown = drawdown.min()

    # 超额最大回撤
    excess_nav = strategy_nav / benchmark_nav
    excess_cummax = excess_nav.cummax()
    excess_drawdown = (excess_nav - excess_cummax) / excess_cummax
    excess_max_drawdown = excess_drawdown.min()

    # 调仓胜率：按调仓频率计算每期超额收益为正的比例
    strategy_nav_dt = strategy_nav.copy()
    strategy_nav_dt.index = pd.to_datetime(strategy_nav_dt.index)
    benchmark_nav_dt = benchmark_nav.copy()
    benchmark_nav_dt.index = pd.to_datetime(benchmark_nav_dt.index)

    if rebalance_type == 'weekly':
        # 周频：按周重采样
        strategy_weekly = strategy_nav_dt.resample('W-FRI').last().dropna()
        benchmark_weekly = benchmark_nav_dt.resample('W-FRI').last().dropna()
        strategy_period_returns = strategy_weekly.pct_change().dropna()
        benchmark_period_returns = benchmark_weekly.pct_change().dropna()
    else:
        # 月频：按月重采样
        strategy_monthly = strategy_nav_dt.resample('ME').last().dropna()
        benchmark_monthly = benchmark_nav_dt.resample('ME').last().dropna()
        strategy_period_returns = strategy_monthly.pct_change().dropna()
        benchmark_period_returns = benchmark_monthly.pct_change().dropna()

    # 对齐后计算超额收益
    common_idx = strategy_period_returns.index.intersection(benchmark_period_returns.index)
    excess_period_returns = strategy_period_returns.loc[common_idx] - benchmark_period_returns.loc[common_idx]
    rebalance_win_rate = (excess_period_returns > 0).sum() / len(excess_period_returns) if len(excess_period_returns) > 0 else 0

    return {
        '总收益': total_return,
        '基准总收益': benchmark_total_return,
        '超额总收益': excess_total_return,
        '年化收益': annual_return,
        '基准年化收益': benchmark_annual_return,
        '超额年化收益': excess_annual_return,
        '年化波动率': annual_vol,
        '夏普比率': sharpe,
        '信息比率': ir,
        '最大回撤': max_drawdown,
        '超额最大回撤': excess_max_drawdown,
        '调仓胜率': rebalance_win_rate,
    }


# ============================================================================
# 净值计算函数
# ============================================================================

def calculate_strategy_nav(holdings_history: Dict[pd.Timestamp, List[str]],
                           prices_df: pd.DataFrame,
                           start_date: pd.Timestamp = None) -> pd.Series:
    """
    计算策略净值（按调仓日计算，与factor_value_backtest.py一致）

    参数:
        holdings_history: Dict, 持仓历史 {日期: [行业列表]}
        prices_df: pd.DataFrame, 价格数据
        start_date: pd.Timestamp, 起始日期

    返回:
        pd.Series: 策略净值序列（索引为调仓日）
    """
    rebalance_dates = sorted(holdings_history.keys())
    if start_date is not None:
        rebalance_dates = [d for d in rebalance_dates if d >= start_date]

    if len(rebalance_dates) == 0:
        return pd.Series(dtype=float)

    # 初始化净值序列（索引为调仓日）
    nav = pd.Series(index=rebalance_dates, dtype=float)
    nav.iloc[0] = 1.0

    for i in range(1, len(rebalance_dates)):
        prev_date = rebalance_dates[i - 1]
        curr_date = rebalance_dates[i]
        holdings = holdings_history[prev_date]  # 使用上一期的持仓

        if len(holdings) == 0:
            # 空仓，净值不变
            nav.iloc[i] = nav.iloc[i - 1]
            continue

        # 等权持有
        valid_holdings = [h for h in holdings if h in prices_df.columns]
        if len(valid_holdings) == 0:
            nav.iloc[i] = nav.iloc[i - 1]
            continue

        # 计算期间收益（等权平均）
        if prev_date in prices_df.index and curr_date in prices_df.index:
            period_ret = (prices_df.loc[curr_date, valid_holdings] / prices_df.loc[prev_date, valid_holdings] - 1).mean()
        else:
            period_ret = 0

        nav.iloc[i] = nav.iloc[i - 1] * (1 + period_ret)

    return nav


def calculate_benchmark_nav(prices_df: pd.DataFrame,
                            rebalance_dates: List[pd.Timestamp]) -> pd.Series:
    """
    计算基准净值（行业等权，按调仓日计算，与factor_value_backtest.py一致）

    参数:
        prices_df: pd.DataFrame, 价格数据
        rebalance_dates: List[pd.Timestamp], 调仓日期列表

    返回:
        pd.Series: 基准净值序列（索引为调仓日）
    """
    # 确保调仓日期在价格数据中存在
    valid_dates = [d for d in rebalance_dates if d in prices_df.index]

    if not valid_dates:
        return pd.Series(dtype=float)

    # 初始化基准净值
    benchmark_nav = pd.Series(index=valid_dates, dtype=float)
    benchmark_nav.iloc[0] = 1.0

    # 按调仓日计算基准收益
    for i in range(1, len(valid_dates)):
        prev_date = valid_dates[i - 1]
        curr_date = valid_dates[i]

        # 计算期间基准收益（所有行业等权平均）
        if prev_date in prices_df.index and curr_date in prices_df.index:
            period_ret = (prices_df.loc[curr_date] / prices_df.loc[prev_date] - 1).mean()
        else:
            period_ret = 0

        benchmark_nav.iloc[i] = benchmark_nav.iloc[i - 1] * (1 + period_ret)

    return benchmark_nav


# ============================================================================
# 泡沫因子回测
# ============================================================================

def run_bubble_factor_backtest(prices_df: pd.DataFrame,
                                amount_df: pd.DataFrame,
                                unified_start_date: pd.Timestamp,
                                verbose: bool = True) -> Dict:
    """
    运行泡沫因子回测

    参数:
        prices_df: pd.DataFrame, 价格数据
        amount_df: pd.DataFrame, 成交量数据
        unified_start_date: pd.Timestamp, 统一起始日期
        verbose: bool, 是否打印详细信息

    返回:
        Dict: 回测结果
    """
    from factor_signal import compute_positive_bubble_signal

    if verbose:
        print("计算泡沫信号...")

    # 计算泡沫信号
    bubble_signal = compute_positive_bubble_signal(prices_df, amount_df, verbose=verbose)

    if verbose:
        print(f"泡沫信号计算完成: {bubble_signal.shape[0]}周")

    # 获取周度调仓日期
    weekly_dates = get_weekly_rebalance_dates(prices_df)
    valid_dates = weekly_dates[weekly_dates >= unified_start_date]

    # 构建持仓历史
    holdings_history = {}
    for date in valid_dates:
        # 找到最近的泡沫信号日期
        signal_dates = bubble_signal.index[bubble_signal.index <= date]
        if len(signal_dates) == 0:
            holdings_history[date] = []
            continue
        nearest_signal_date = signal_dates[-1]

        # 获取信号=1的行业
        signals = bubble_signal.loc[nearest_signal_date]
        holdings = signals[signals == 1].index.tolist()
        holdings_history[date] = holdings

    # 计算净值
    nav_series = calculate_strategy_nav(holdings_history, prices_df, unified_start_date)

    # 计算基准净值（按调仓日计算，与factor_value_backtest.py一致）
    rebalance_dates_list = sorted(holdings_history.keys())
    benchmark_nav = calculate_benchmark_nav(prices_df, rebalance_dates_list)

    # 计算统计指标
    stats = calculate_performance_stats(nav_series, benchmark_nav, rebalance_type='weekly')

    # 统计空仓和平均持仓
    empty_count = sum(1 for h in holdings_history.values() if len(h) == 0)
    total_count = len(holdings_history)
    avg_holdings = np.mean([len(h) for h in holdings_history.values()])

    return {
        'holdings_history': holdings_history,
        'nav_series': nav_series,
        'benchmark_nav': benchmark_nav,
        'stats': stats,
        'signal_matrix': bubble_signal,
        'empty_count': empty_count,
        'total_count': total_count,
        'avg_holdings': avg_holdings,
        'rebalance_dates': valid_dates
    }


# ============================================================================
# 交集因子回测
# ============================================================================

def run_intersection_factor_backtest(prices_df: pd.DataFrame,
                                      amount_df: pd.DataFrame,
                                      unified_start_date: pd.Timestamp,
                                      top_n: int = 6,
                                      verbose: bool = True) -> Dict:
    """
    运行交集因子回测

    参数:
        prices_df: pd.DataFrame, 价格数据
        amount_df: pd.DataFrame, 成交量数据
        unified_start_date: pd.Timestamp, 统一起始日期
        top_n: int, 合成因子选取的行业数量
        verbose: bool, 是否打印详细信息

    返回:
        Dict: 回测结果
    """
    from factor_value_backtest import DataContainer
    from factor_signal import compute_positive_bubble_signal
    import factor_value as f

    if verbose:
        print("加载数据...")

    # 加载数据
    data = DataContainer(load_constituent=True)

    if verbose:
        print("计算合成因子...")

    # 计算合成因子
    composite_factor = f.momentum_synthesis_equal(
        data.prices_df,
        barra_factor_returns_df=data.barra_factor_returns_df,
        constituent_df=data.constituent_df,
        stock_price_df=data.stock_price_df,
        stock_mv_df=data.stock_mv_df,
        volume_df=data.volume_df,
        industry_code_df=data.industry_code_df,
        high_df=data.high_df,
        low_df=data.low_df
    )

    if verbose:
        print("计算泡沫信号...")

    # 计算泡沫信号
    bubble_signal = compute_positive_bubble_signal(prices_df, amount_df, verbose=verbose)

    # 获取月度调仓日期
    monthly_dates = get_monthly_rebalance_dates(prices_df)
    valid_dates = monthly_dates[monthly_dates >= unified_start_date]

    # 过滤有效日期
    valid_dates = valid_dates[
        (valid_dates >= composite_factor.index.min()) &
        (valid_dates <= composite_factor.index.max()) &
        (valid_dates >= bubble_signal.index.min())
    ]

    if verbose:
        print(f"回测期间: {valid_dates[0].strftime('%Y-%m-%d')} 至 {valid_dates[-1].strftime('%Y-%m-%d')}")

    # 构建持仓历史
    holdings_history = {}
    detail_history = []  # 调仓详情

    for date in valid_dates:
        # 获取合成因子Top N
        factor_values = composite_factor.loc[date].dropna() if date in composite_factor.index else pd.Series()
        top_industries = factor_values.nlargest(top_n).index.tolist() if len(factor_values) >= top_n else []

        # 获取泡沫信号=1的行业
        signal_dates = bubble_signal.index[bubble_signal.index <= date]
        if len(signal_dates) > 0:
            nearest_signal_date = signal_dates[-1]
            signals = bubble_signal.loc[nearest_signal_date]
            bubble_industries = signals[signals == 1].index.tolist()
        else:
            bubble_industries = []

        # 计算交集
        intersection = list(set(top_industries) & set(bubble_industries))
        holdings_history[date] = intersection

        detail_history.append({
            'date': date,
            'top_industries': top_industries,
            'bubble_industries': bubble_industries,
            'intersection': intersection
        })

    # 计算净值
    nav_series = calculate_strategy_nav(holdings_history, prices_df, unified_start_date)

    # 计算基准净值（按调仓日计算，与factor_value_backtest.py一致）
    rebalance_dates_list = sorted(holdings_history.keys())
    benchmark_nav = calculate_benchmark_nav(prices_df, rebalance_dates_list)

    # 计算统计指标
    stats = calculate_performance_stats(nav_series, benchmark_nav, rebalance_type='monthly')

    # 统计空仓和平均持仓
    empty_count = sum(1 for h in holdings_history.values() if len(h) == 0)
    total_count = len(holdings_history)
    avg_holdings = np.mean([len(h) for h in holdings_history.values()])

    return {
        'holdings_history': holdings_history,
        'nav_series': nav_series,
        'benchmark_nav': benchmark_nav,
        'stats': stats,
        'signal_matrix': bubble_signal,
        'composite_factor': composite_factor,
        'detail_history': detail_history,
        'empty_count': empty_count,
        'total_count': total_count,
        'avg_holdings': avg_holdings,
        'rebalance_dates': valid_dates
    }


# ============================================================================
# Excel导出功能 - 每个因子一个sheet页，所有内容放在同一个sheet
# ============================================================================

def write_factor_to_sheet(writer, sheet_name: str, factor_config: Dict,
                          result: Dict, signal_matrix: pd.DataFrame = None):
    """
    将单个因子的所有回测结果写入一个sheet页

    参数:
        writer: ExcelWriter对象
        sheet_name: sheet名称
        factor_config: 因子配置
        result: 回测结果
        signal_matrix: 信号矩阵
    """
    rows = []

    # ========== 策略概览部分 ==========
    rows.append({'A': '【策略说明】', 'B': '', 'C': '', 'D': '', 'E': '', 'F': ''})
    rows.append({'A': '策略名称', 'B': factor_config.get('name', ''), 'C': '', 'D': '', 'E': '', 'F': ''})
    rows.append({'A': '策略逻辑', 'B': factor_config.get('logic', ''), 'C': '', 'D': '', 'E': '', 'F': ''})
    rows.append({'A': '调仓频率', 'B': '周频' if factor_config.get('rebalance_type') == 'weekly' else '月频', 'C': '', 'D': '', 'E': '', 'F': ''})
    rows.append({'A': '', 'B': '', 'C': '', 'D': '', 'E': '', 'F': ''})

    # ========== 绩效指标部分 ==========
    rows.append({'A': '【绩效指标】', 'B': '', 'C': '', 'D': '', 'E': '', 'F': ''})
    stats = result['stats']
    rows.append({'A': '总收益', 'B': f"{stats['总收益']:.2%}", 'C': '', 'D': '', 'E': '', 'F': ''})
    rows.append({'A': '基准总收益', 'B': f"{stats['基准总收益']:.2%}", 'C': '', 'D': '', 'E': '', 'F': ''})
    rows.append({'A': '超额总收益', 'B': f"{stats['超额总收益']:.2%}", 'C': '', 'D': '', 'E': '', 'F': ''})
    rows.append({'A': '年化收益', 'B': f"{stats['年化收益']:.2%}", 'C': '', 'D': '', 'E': '', 'F': ''})
    rows.append({'A': '基准年化收益', 'B': f"{stats['基准年化收益']:.2%}", 'C': '', 'D': '', 'E': '', 'F': ''})
    rows.append({'A': '超额年化收益', 'B': f"{stats['超额年化收益']:.2%}", 'C': '', 'D': '', 'E': '', 'F': ''})
    rows.append({'A': '年化波动率', 'B': f"{stats['年化波动率']:.2%}", 'C': '', 'D': '', 'E': '', 'F': ''})
    rows.append({'A': '夏普比率', 'B': f"{stats['夏普比率']:.2f}", 'C': '', 'D': '', 'E': '', 'F': ''})
    rows.append({'A': '信息比率', 'B': f"{stats['信息比率']:.2f}", 'C': '', 'D': '', 'E': '', 'F': ''})
    rows.append({'A': '最大回撤', 'B': f"{stats['最大回撤']:.2%}", 'C': '', 'D': '', 'E': '', 'F': ''})
    rows.append({'A': '超额最大回撤', 'B': f"{stats['超额最大回撤']:.2%}", 'C': '', 'D': '', 'E': '', 'F': ''})
    rows.append({'A': '调仓胜率', 'B': f"{stats['调仓胜率']:.2%}", 'C': '', 'D': '', 'E': '', 'F': ''})
    rows.append({'A': '', 'B': '', 'C': '', 'D': '', 'E': '', 'F': ''})

    # ========== 每年收益统计部分 ==========
    rows.append({'A': '【每年收益统计】', 'B': '', 'C': '', 'D': '', 'E': '', 'F': ''})
    rows.append({'A': '年份', 'B': '多头收益(%)', 'C': '超额收益(%)', 'D': '基准收益(%)', 'E': '', 'F': ''})

    # 计算每年收益
    nav_series = result['nav_series']
    benchmark_nav = result['benchmark_nav']

    # 对齐索引
    aligned_benchmark = benchmark_nav.reindex(nav_series.index)

    # 按年分组计算收益
    nav_df_temp = pd.DataFrame({
        'strategy': nav_series,
        'benchmark': aligned_benchmark
    })
    nav_df_temp.index = pd.to_datetime(nav_df_temp.index)

    years = nav_df_temp.index.year.unique()
    yearly_data = []

    for year in sorted(years):
        year_data = nav_df_temp[nav_df_temp.index.year == year]
        if len(year_data) < 2:
            continue

        # 计算该年收益
        strategy_return = (year_data['strategy'].iloc[-1] / year_data['strategy'].iloc[0] - 1) * 100
        benchmark_return = (year_data['benchmark'].iloc[-1] / year_data['benchmark'].iloc[0] - 1) * 100
        excess_return = strategy_return - benchmark_return

        # 判断是否是最后一年（不完整年份）
        last_date = year_data.index[-1]
        if year == years.max() and last_date.month < 12:
            year_label = f"{year}(截至{last_date.month}月{last_date.day}日)"
        else:
            year_label = str(year)

        yearly_data.append({
            'year': year,
            'year_label': year_label,
            'strategy': strategy_return,
            'excess': excess_return,
            'benchmark': benchmark_return
        })

        rows.append({'A': year_label, 'B': f"{strategy_return:.2f}", 'C': f"{excess_return:.2f}", 'D': f"{benchmark_return:.2f}", 'E': '', 'F': ''})

    # 添加全样本汇总
    if len(nav_series) >= 2:
        total_strategy = (nav_series.iloc[-1] / nav_series.iloc[0] - 1) * 100
        total_benchmark = (aligned_benchmark.iloc[-1] / aligned_benchmark.iloc[0] - 1) * 100
        total_excess = total_strategy - total_benchmark
        rows.append({'A': '全样本', 'B': f"{total_strategy:.2f}", 'C': f"{total_excess:.2f}", 'D': f"{total_benchmark:.2f}", 'E': '', 'F': ''})

    rows.append({'A': '', 'B': '', 'C': '', 'D': '', 'E': '', 'F': ''})

    # ========== 持仓统计部分 ==========
    rows.append({'A': '【持仓统计】', 'B': '', 'C': '', 'D': '', 'E': '', 'F': ''})
    empty_count = result.get('empty_count', 0)
    total_count = result.get('total_count', 1)
    avg_holdings = result.get('avg_holdings', 0)
    rows.append({'A': '空仓月份', 'B': f"{empty_count}/{total_count} ({empty_count/total_count*100:.1f}%)", 'C': '', 'D': '', 'E': '', 'F': ''})
    rows.append({'A': '平均持仓', 'B': f"{avg_holdings:.1f}个行业", 'C': '', 'D': '', 'E': '', 'F': ''})
    rows.append({'A': '', 'B': '', 'C': '', 'D': '', 'E': '', 'F': ''})

    # 创建DataFrame并写入
    overview_df = pd.DataFrame(rows)
    overview_df.columns = ['项目', '内容', '', '', '', '']

    # 写入策略概览
    overview_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)

    current_row = len(rows) + 2

    # ========== 净值序列部分（只保留调仓日） ==========
    nav_series = result['nav_series']
    benchmark_nav = result['benchmark_nav']
    holdings_history = result['holdings_history']

    # 只保留调仓日的净值
    rebalance_dates = sorted(holdings_history.keys())
    nav_series_rebalance = nav_series.reindex(rebalance_dates).dropna()
    benchmark_rebalance = benchmark_nav.reindex(rebalance_dates).dropna()

    # 取交集
    common_dates = nav_series_rebalance.index.intersection(benchmark_rebalance.index)
    nav_series_rebalance = nav_series_rebalance.loc[common_dates]
    benchmark_rebalance = benchmark_rebalance.loc[common_dates]

    # 写入净值序列标题
    ws = writer.sheets[sheet_name]
    ws.cell(row=current_row, column=1, value='【净值序列】')
    current_row += 1

    nav_df = pd.DataFrame({
        '日期': nav_series_rebalance.index.strftime('%Y-%m-%d'),
        '策略净值': nav_series_rebalance.values,
        '基准净值': benchmark_rebalance.values,
    })
    nav_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=current_row)

    # ========== 在H18位置插入净值折线图 ==========
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.axis import ChartLines
    from openpyxl.drawing.line import LineProperties
    from openpyxl.chart.shapes import GraphicalProperties

    # 创建虚线网格线样式
    def create_dashed_gridlines():
        """创建虚线网格线"""
        gridlines = ChartLines()
        gridlines.spPr = GraphicalProperties(ln=LineProperties(prstDash='dash'))
        return gridlines

    chart = LineChart()
    chart.title = "净值序列"
    chart.style = 2  # 使用简洁样式
    chart.y_axis.title = "净值"
    chart.x_axis.title = "日期"
    chart.width = 18  # 宽18cm（与factor_value_backtest一致）
    chart.height = 10  # 高10cm（与factor_value_backtest一致）

    # 设置网格线为虚线
    chart.y_axis.majorGridlines = create_dashed_gridlines()
    chart.x_axis.majorGridlines = None  # 不显示X轴网格线

    # 显示轴刻度标签
    chart.x_axis.tickLblPos = "low"
    chart.y_axis.tickLblPos = "low"
    chart.x_axis.delete = False
    chart.y_axis.delete = False

    # 设置X轴刻度间隔（每隔一定数量显示一个标签）
    if len(nav_df) > 8:
        chart.x_axis.tickLblSkip = max(1, len(nav_df) // 8)
        chart.x_axis.tickMarkSkip = max(1, len(nav_df) // 8)

    # 数据范围：策略净值和基准净值（不含超额净值）
    data_start_row = current_row + 1  # 标题行
    data_end_row = current_row + len(nav_df)

    # 策略净值 (B列) 和 基准净值 (C列)
    data = Reference(ws, min_col=2, min_row=data_start_row, max_col=3, max_row=data_end_row)
    cats = Reference(ws, min_col=1, min_row=data_start_row + 1, max_row=data_end_row)  # 日期作为X轴

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    # 放置在H5位置
    ws.add_chart(chart, "H5")

    current_row += len(nav_df) + 3

    # ========== 历史持仓部分（从最新到最久） ==========
    ws.cell(row=current_row, column=1, value='【历史持仓】')
    current_row += 1

    holdings_data = []
    for date, holdings in sorted(holdings_history.items(), reverse=True):  # reverse=True 从最新到最久
        holdings_data.append({
            '调仓日期': date.strftime('%Y-%m-%d'),
            '持仓数量': len(holdings),
            '持仓行业': ', '.join(holdings) if holdings else '空仓'
        })
    holdings_df = pd.DataFrame(holdings_data)
    holdings_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=current_row)
    current_row += len(holdings_df) + 3

    # ========== 调仓详情部分 ==========
    ws.cell(row=current_row, column=1, value='【调仓详情】')
    current_row += 1

    if 'detail_history' in result:
        detail_data = []
        for detail in result['detail_history']:
            detail_data.append({
                '调仓日期': detail['date'].strftime('%Y-%m-%d'),
                '合成因子Top6': ', '.join(detail.get('top_industries', [])) or '-',
                '泡沫信号行业': ', '.join(detail.get('bubble_industries', [])) or '-',
                '交集持仓': ', '.join(detail.get('intersection', [])) or '空仓'
            })
        detail_df = pd.DataFrame(detail_data)
    else:
        # 泡沫因子没有detail_history，直接用holdings_history
        detail_df = holdings_df.copy()
    detail_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=current_row)
    current_row += len(detail_df) + 3

    # ========== 信号矩阵部分 ==========
    ws.cell(row=current_row, column=1, value='【信号矩阵】')
    current_row += 1

    if signal_matrix is not None:
        signal_export = signal_matrix.copy()
        signal_export.index = signal_export.index.strftime('%Y-%m-%d')
        signal_export.index.name = '日期'
        signal_export.to_excel(writer, sheet_name=sheet_name, startrow=current_row)


def export_signal_backtest_to_excel(results: Dict[str, Dict],
                                     output_file: str = None,
                                     verbose: bool = True) -> str:
    """
    将所有信号因子回测结果导出到Excel，每个因子一个sheet页

    参数:
        results: Dict, {因子名: 回测结果}
        output_file: str, 输出文件路径
        verbose: bool, 是否打印信息

    返回:
        str: 输出文件路径
    """
    if output_file is None:
        # 输出到 factor分析 目录
        output_dir = r"C:\Users\MECHREVO\001_TEMP\Quant\行业动量\factor分析"
        os.makedirs(output_dir, exist_ok=True)
        timestamp = datetime.now().strftime('%H%M%S')
        output_file = os.path.join(output_dir, f"信号因子统一分析_最新_{timestamp}.xlsx")

    if verbose:
        print(f"\n正在导出回测结果到: {output_file}")

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for factor_name, result in results.items():
            factor_config = SIGNAL_FACTOR_CONFIG.get(factor_name, {
                'name': factor_name,
                'logic': '',
                'rebalance_type': 'monthly'
            })

            # 获取信号矩阵
            signal_matrix = result.get('signal_matrix', None)

            # 写入sheet
            write_factor_to_sheet(writer, factor_name, factor_config, result, signal_matrix)

            if verbose:
                print(f"  已写入: {factor_name}")

    if verbose:
        print(f"导出完成！文件: {output_file}")

    return output_file


# ============================================================================
# 主程序入口
# ============================================================================

if __name__ == "__main__":
    from factor_value_backtest import DataContainer
    from factor_signal import compute_positive_bubble_signal
    import factor_value as f

    print("=" * 70)
    print("信号因子回测")
    print("=" * 70)

    start_time = time.time()

    # ========== 1. 加载数据 ==========
    print("\n[Step 1] 加载数据...")
    # 传入 backtest_years=10 以计算 first_holding_date（从2016年开始回测）
    data = DataContainer(load_constituent=True, backtest_years=10)
    prices_df = data.prices_df
    amount_df = data.amount_df

    print(f"  价格数据: {prices_df.shape[0]}天 × {prices_df.shape[1]}个行业")
    print(f"  日期范围: {prices_df.index[0].strftime('%Y-%m-%d')} 至 {prices_df.index[-1].strftime('%Y-%m-%d')}")

    # ========== 2. 获取统一起始日期（从DataContainer.first_holding_date计算）==========
    print("\n[Step 2] 获取统一起始日期...")
    unified_start_date = get_unified_start_date_from_data(data)
    print(f"  统一起始日期: {unified_start_date.strftime('%Y-%m-%d')} (来自DataContainer.first_holding_date)")

    # ========== 3. 运行泡沫因子回测（已注释）==========
    # print("\n[Step 3] 运行泡沫因子回测...")
    # bubble_result = run_bubble_factor_backtest(
    #     prices_df, amount_df, unified_start_date, verbose=True
    # )
    # print(f"  泡沫因子回测完成")
    # for key, value in bubble_result['stats'].items():
    #     if isinstance(value, float):
    #         print(f"    {key}: {value:.2%}" if abs(value) < 100 else f"    {key}: {value:.2f}")
    #     else:
    #         print(f"    {key}: {value}")

    # ========== 4. 运行交集因子回测 ==========
    print("\n[Step 3] 运行交集因子回测...")
    intersection_result = run_intersection_factor_backtest(
        prices_df, amount_df, unified_start_date, top_n=6, verbose=True
    )
    print(f"  交集因子回测完成")
    for key, value in intersection_result['stats'].items():
        if isinstance(value, float):
            print(f"    {key}: {value:.2%}" if abs(value) < 100 else f"    {key}: {value:.2f}")
        else:
            print(f"    {key}: {value}")

    # ========== 5. 导出Excel ==========
    print("\n[Step 4] 导出Excel...")
    results = {
        # 'bubble_factor': bubble_result,
        'intersection_factor': intersection_result
    }
    output_file = export_signal_backtest_to_excel(results, verbose=True)

    # ========== 6. 计算耗时 ==========
    elapsed = time.time() - start_time
    minutes = int(elapsed // 60)
    seconds = int(elapsed % 60)
    print(f"\n总耗时: {minutes}分{seconds}秒")
    print(f"输出文件: {output_file}")
    print("=" * 70)
