"""
因子分析模块
负责批量因子分析、IC/IR 计算、分层回测、统计指标计算
输出到Excel，每个因子一个sheet页，包含因子说明

支持多周期回测：20日、60日、120日、240日
超额收益基准：申万一级行业等权收益
"""
import pandas as pd
import numpy as np
import inspect
import subprocess
import platform
from scipy import stats
from copy import copy
from datetime import datetime
# 导入数据加载模块和因子模块
import data_loader
import factor_value
import os
# 导入openpyxl图表模块
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.shapes import GraphicalProperties


# 输出文件夹
OUTPUT_DIR = "factor分析"
# 默认参数
DEFAULT_REBALANCE_FREQ = 20   # 调仓频率（天）- 仅在非月度调仓时使用
DEFAULT_MONTHLY_REBALANCE = True  # 默认使用月度调仓（每月最后一个交易日）
DEFAULT_REBALANCE_TYPE = 'monthly'  # 默认调仓类型: 'monthly', 'weekly', 'fixed'
N_LAYERS = 5                  # 分层回测层数
LOOKBACK_WINDOWS = [20, 60, 120, 240]  # 回看窗口列表（因子计算的回看窗口）
DEFAULT_BACKTEST_YEARS = 10   # 默认回测年限

# 统一因子配置表
# 每个因子包含：
#   - func: 因子计算函数
#   - base_warmup: 固定预热期（交易日）
#   - window_multiplier: 窗口倍数，预热期 = base_warmup + lookback_window * window_multiplier
#   - lookback_windows: 该因子适用的回看窗口列表，None 表示使用默认的 LOOKBACK_WINDOWS
#   - requires_constituent: 是否需要成分股数据
#
# 按照 factor_value.py 中的定义顺序排列
FACTOR_CONFIG = {
    # ===== 基础动量因子（只需要 lookback_window 天）=====
    'momentum': {
        'func': factor_value.momentum,
        'base_warmup': 0,
        'window_multiplier': 1,
        'lookback_windows': None,
        'requires_constituent': False,
        'description': '基础动量1-5-传统动量因子（区间收益率）'
    },
    'momentum_zscore': {
        'func': factor_value.momentum_zscore,
        'base_warmup': 0,
        'window_multiplier': 1,
        'lookback_windows': None,
        'requires_constituent': False,
        'description': '基础动量2-5-标准化动量因子（横截面Z-score）'
    },
    'momentum_rank_zscore': {
        'func': factor_value.momentum_rank_zscore,
        'base_warmup': 0,
        'window_multiplier': 1,
        'lookback_windows': None,
        'requires_constituent': False,
        'description': '基础动量3-5-Rank标准化动量因子（排名标准化）'
    },
    'momentum_sharpe': {
        'func': factor_value.momentum_sharpe,
        'base_warmup': 0,
        'window_multiplier': 1,
        'lookback_windows': None,
        'requires_constituent': False,
        'description': '基础动量4-5-夏普动量因子（风险调整后的动量）'
    },
    'momentum_calmar_ratio': {
        'func': factor_value.momentum_calmar_ratio,
        'base_warmup': 0,
        'window_multiplier': 1,
        'lookback_windows': None,
        'requires_constituent': False,
        'description': '基础动量5-5-Calmar比率因子（最大回撤调整）'
    },

    # ===== 平稳动量因子 =====
    'momentum_volume_return_corr': {
        'func': factor_value.momentum_volume_return_corr,
        'base_warmup': 0,
        'window_multiplier': 1,
        'lookback_windows': [10, 20, 30, 60, 120, 180, 240],  # 研报最优10日，20日也有一定效果
        'requires_constituent': False,
        'description': '平稳动量1-5-量益相关性动量因子（量价同向）'
    },
    'momentum_turnover_adj': {
        'func': factor_value.momentum_turnover_adj,
        'base_warmup': 0,
        'window_multiplier': 1,
        'lookback_windows': None,
        'requires_constituent': False,
        'description': '平稳动量2-5-换手率惩罚动量因子（量价背离）'
    },
    'momentum_rebound_with_crowding_filter': {
        'func': factor_value.momentum_rebound_with_crowding_filter,
        'base_warmup': 0,
        'window_multiplier': 1,
        'lookback_windows': None,
        'requires_constituent': False,
        'description': '平稳动量3-5-反弹综合拥挤过滤动量因子'
    },
    'momentum_amplitude_cut': {
        'func': factor_value.momentum_amplitude_cut,
        'base_warmup': 0,
        'window_multiplier': 1,
        'lookback_windows': [20, 60, 90, 120, 160, 240],  # 行业最优参数测试
        'requires_constituent': False,
        'description': '平稳动量4-5-振幅切割稳健动量因子（剔高振幅）'
    },
    'momentum_price_volume_icir': {
        'func': factor_value.momentum_price_volume_icir,
        'base_warmup': 240,
        'window_multiplier': 0,
        'lookback_windows': [240], #固定使用10-240日
        'requires_constituent': False,
        'description': '平稳动量5-5-量价清洗ICIR加权动量因子（剔高成交量）'
    },

    # ===== 特质收益动量因子 =====
    'momentum_pure_liquidity_stripped': {
        'func': factor_value.momentum_pure_liquidity_stripped,
        'base_warmup': 240,
        'window_multiplier': 1,
        'lookback_windows': None,
        'requires_constituent': False,
        'description': '特质动量1-2-剥离异动提纯动量因子（特质收益）'
    },
    'momentum_residual': {
        'func': factor_value.momentum_residual,
        'base_warmup': 240,  # 12个月月度数据预热（约240交易日）
        'window_multiplier': 0,  # 研报固定使用12个月，不依赖window参数
        'lookback_windows': [240],  # 固定窗口，研报使用12个月回看
        'requires_constituent': False,
        'requires_barra': True,  # 需要Barra因子数据
        'description': '特质动量2-2-行业残差动量因子Barra兴业'
    },

    # ===== 行业间相关性动量因子 =====
    'momentum_cross_industry_lasso': {
        'func': factor_value.momentum_cross_industry_lasso,
        'base_warmup': 200,
        'window_multiplier': 1,
        'lookback_windows': [60, 120, 240, 480],  # 日数→月数：60→3月, 120→6月, 240→12月, 480→24月
        'requires_constituent': False,
        'description': '行业间动量1-1-Lasso因子'
    },

    # ===== 行业内关系动量因子（需要成分股数据）=====
    'momentum_industry_component': {
        'func': factor_value.momentum_industry_component,
        'base_warmup': 0,  # 无额外固定预热期
        'window_multiplier': 1,  # window已经是交易日数，预热期 = window * 1
        'lookback_windows': [20, 60, 120, 240, 480, 720],  # 交易日数（对应1、3、6、12、24、36个月）
        'requires_constituent': True,
        'description': '行业内动量1-3-行业成分股动量因子 东方'
    },
    'momentum_lead_lag_enhanced': {
        'func': factor_value.momentum_lead_lag_enhanced,
        'base_warmup': 0,
        'window_multiplier': 1,
        'lookback_windows': None,  # 原文固定使用20日窗口（1个月）
        'requires_constituent': True,
        'default_split_ratio': 0.1,  # 最优分割参数（测试结果：0.4表现最好）
        'description': '行业内动量2-3-龙头领先修正动量因子'
    },
    'momentum_pca': {
        'func': factor_value.momentum_pca,
        'base_warmup': 125,  # pca_window(120) + lag(5) = 125
        'window_multiplier': 0,  # 动量窗口固定为10天（双周），不使用外部window
        'lookback_windows': [120],  # 固定窗口，与研报一致
        'requires_constituent': True,
        'rebalance_type': 'weekly',  # 研报使用周度调仓
        'description': '行业内动量3-3-PcaMom集中度分析因子'
    },

    # ===== 多因子合成因子 =====
    # 注意：这两个因子的window参数无实际意义，各成分因子使用各自的最优窗口
    'momentum_synthesis_equal': {
        'func': factor_value.momentum_synthesis_equal,
        'base_warmup': 240,  # 最大成分因子窗口（用于预热期计算）
        'window_multiplier': 0,
        'lookback_windows': [1],  # 占位符，实际不使用window参数
        'requires_constituent': True,
        'requires_barra': True,
        'description': '合成因子1-等权合成动量因子'
    },
    'momentum_synthesis_icir': {
        'func': factor_value.momentum_synthesis_icir,
        'base_warmup': 240,  # 最大成分因子窗口（用于预热期计算）
        'window_multiplier': 0,
        'lookback_windows': [1],  # 占位符，实际不使用window参数
        'requires_constituent': True,
        'requires_barra': True,
        'description': '合成因子2-滚动ICIR加权合成动量因子(12月)'
    },
}

def get_output_path(factor_name=None, duration_str=None, end_date=None):
    """
    获取输出文件路径

    参数:
    factor_name: str 或 list, 因子名称。
        - 如果是单个因子名称（str），文件名为 因子description_截止日期_时间戳_耗时.xlsx
        - 如果是多个因子（list）或 None，文件名为 因子统一分析_截止日期_时间戳_耗时.xlsx
    duration_str: str, 计算耗时字符串，如 "1h30m" 或 "45m"
    end_date: str, 截止日期，格式为 'YYYY-MM-DD'，None表示使用"最新"

    返回:
    str: 完整的输出文件路径，格式为 factor分析/因子description_截止日期_HHMMSS_耗时.xlsx
    """
    # 确保输出文件夹存在
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
        print(f"已创建输出文件夹: {OUTPUT_DIR}")

    # 生成时间戳（只保留小时分秒）
    timestamp = datetime.now().strftime('%H%M%S')

    # 处理截止日期
    if end_date:
        # 将 'YYYY-MM-DD' 格式转换为 'xx年xx月截止'
        parts = end_date.split('-')
        end_date_str = f"{parts[0]}年{parts[1]}月截止"
    else:
        end_date_str = "最新"

    # 确定因子名称（使用description）
    if factor_name is None:
        name_prefix = "因子统一分析"
    elif isinstance(factor_name, str):
        # 从FACTOR_CONFIG获取description
        if factor_name in FACTOR_CONFIG:
            name_prefix = FACTOR_CONFIG[factor_name].get('description', factor_name)
        else:
            name_prefix = factor_name
    elif isinstance(factor_name, list):
        if len(factor_name) == 1:
            # 从FACTOR_CONFIG获取description
            if factor_name[0] in FACTOR_CONFIG:
                name_prefix = FACTOR_CONFIG[factor_name[0]].get('description', factor_name[0])
            else:
                name_prefix = factor_name[0]
        else:
            name_prefix = "因子统一分析"
    else:
        name_prefix = "因子统一分析"

    # 生成文件名：因子description_截止日期_时间戳_耗时.xlsx
    if duration_str:
        filename = f"{name_prefix}_{end_date_str}_{timestamp}_{duration_str}.xlsx"
    else:
        filename = f"{name_prefix}_{end_date_str}_{timestamp}.xlsx"

    return os.path.join(OUTPUT_DIR, filename)


def calculate_backtest_dates(data_start, data_end, end_date=None, backtest_years=10, freq='M'):
    """
    计算回测的起始和结束日期

    逻辑：
    1. 先确定最晚持仓日（最新完整周期末）
    2. 往前推backtest_years年的12月底作为最早持仓日
    3. 如果数据不够，就从数据起始日开始

    参数:
    data_start: pd.Timestamp, 数据交集的起始日期
    data_end: pd.Timestamp, 数据交集的结束日期
    end_date: str 或 None, 用户指定的截止日期，None表示使用数据最新日期
    backtest_years: int, 回测年限，默认12年
    freq: str, 调仓频率，'M'=月频(默认), 'W'=周频

    返回:
    tuple: (first_holding_date, last_holding_date, data_start_needed)
        - first_holding_date: str, 最早持仓日（如 '2015-12-31'）
        - last_holding_date: pd.Timestamp, 最晚持仓日（最新完整周期末）
        - data_start_needed: str, 数据需要的起始日期（考虑预热期后会自动调整）
    """
    import pandas as pd
    import calendar

    # 确定截止日期
    if end_date is None:
        end_dt = data_end
    else:
        end_dt = pd.Timestamp(end_date)
        # 不能超过数据范围
        if end_dt > data_end:
            end_dt = data_end

    if freq == 'W':
        # 周频：最晚持仓日为截止日期之前的最后一个周五
        days_since_friday = (end_dt.weekday() - 4) % 7
        if days_since_friday == 0:
            # 如果正好是周五，往前推一周（确保是完整的一周）
            days_since_friday = 7
        last_holding_date = end_dt - pd.Timedelta(days=days_since_friday)

        # 最早持仓日：基于最晚持仓日的年份往前推backtest_years年
        # 找到该年12月最后一个周五
        # 注意：如果最晚持仓日在1月，应该基于上一年计算
        base_year = last_holding_date.year
        if last_holding_date.month == 1:
            base_year -= 1  # 如果最晚持仓日在1月，基准年份应该是上一年

        first_holding_year = base_year - backtest_years
        # 找到该年12月31日
        dec_31 = pd.Timestamp(year=first_holding_year, month=12, day=31)
        # 找到12月31日之前（含）的最后一个周五
        days_to_friday = (dec_31.weekday() - 4) % 7
        first_holding_date = dec_31 - pd.Timedelta(days=days_to_friday)

    else:
        # 月频：最晚持仓日为截止日期所在月的上一个完整月末
        # 例如：2025-12-19 -> 2025-11-30
        last_holding_year = end_dt.year
        last_holding_month = end_dt.month - 1
        if last_holding_month == 0:
            last_holding_month = 12
            last_holding_year -= 1

        # 获取该月的最后一天
        _, last_day = calendar.monthrange(last_holding_year, last_holding_month)
        last_holding_date = pd.Timestamp(year=last_holding_year, month=last_holding_month, day=last_day)

        # 最早持仓日：往前推backtest_years年的12月31日
        first_holding_year = last_holding_date.year - backtest_years
        first_holding_date = pd.Timestamp(year=first_holding_year, month=12, day=31)

    # 检查数据是否足够
    # 数据起始日期需要比最早持仓日更早（因为需要预热期）
    # 这里先返回最早持仓日，实际数据起始会在加载时根据预热期调整
    if first_holding_date < data_start:
        # 数据不够，从数据起始日开始
        # 找到数据起始后的第一个有效持仓日
        first_year = data_start.year
        if data_start.month > 1 or (data_start.month == 1 and data_start.day > 1):
            # 如果数据起始不是1月1日，需要等到下一年
            first_year += 1

        if freq == 'W':
            # 周频：找到该年12月最后一个周五
            dec_31 = pd.Timestamp(year=first_year, month=12, day=31)
            days_to_friday = (dec_31.weekday() - 4) % 7
            first_holding_date = dec_31 - pd.Timedelta(days=days_to_friday)
        else:
            # 月频：12月31日
            first_holding_date = pd.Timestamp(year=first_year, month=12, day=31)

        # 如果调整后的最早持仓日超过了最晚持仓日，说明数据范围太小
        if first_holding_date >= last_holding_date:
            # 使用数据起始后的第一个周期末作为最早持仓日
            first_holding_date = data_start

    return (first_holding_date.strftime('%Y-%m-%d'),
            last_holding_date,
            first_holding_date.strftime('%Y-%m-%d'))


def close_excel():
    """
    关闭所有Excel进程（仅Windows系统）
    在导出Excel文件之前调用，避免文件被占用
    """
    if platform.system() == 'Windows':
        try:
            # 使用taskkill强制关闭Excel进程
            result = subprocess.run(
                ['taskkill', '/F', '/IM', 'EXCEL.EXE'],
                capture_output=True,
                text=True
            )
            if result.returncode == 0:
                print("已关闭Excel进程")
            # returncode != 0 表示没有Excel进程在运行，不需要提示
        except Exception as e:
            print(f"关闭Excel时出错: {e}")

class DataContainer:
    """
    数据容器类，统一管理所有数据
    一次性加载所有数据，避免重复读取文件
    """
    def __init__(self, file_path=data_loader.DEFAULT_CACHE_FILE, start_date=None, end_date=None,
                 exclude_incomplete_month=True, load_constituent=False, backtest_years=None):
        """
        初始化数据容器，加载所有数据

        参数:
        file_path: str, 数据文件路径
        start_date: str, 数据开始日期 (YYYY-MM-DD)，如果指定了backtest_years则忽略此参数
        end_date: str, 数据结束日期 (YYYY-MM-DD)，None表示使用最新数据
        exclude_incomplete_month: bool, 是否排除最后一个不完整的月份（默认True）
        load_constituent: bool, 是否加载成分股数据（默认False，需要时设为True）
        backtest_years: int, 回测年限，如果指定则自动计算起始日期（从最晚持仓日往前推N年的12月底）
        """
        self.file_path = file_path
        self.exclude_incomplete_month = exclude_incomplete_month
        self.load_constituent = load_constituent
        self.backtest_years = backtest_years

        # 加载行业指数数据
        self.prices_df = data_loader.load_price_df(file_path)
        self.high_df = data_loader.load_high_df(file_path)
        self.low_df = data_loader.load_low_df(file_path)
        self.turnover_df = data_loader.load_turnover_df(file_path)
        self.volume_df = data_loader.load_volume_df(file_path)

        # amount_df 使用 volume_df 作为代理（成交量可近似代表成交额趋势）
        self.amount_df = self.volume_df

        # 成分股相关数据（按需加载）
        self.constituent_df = None
        self.stock_price_df = None
        self.stock_mv_df = None
        self.industry_code_df = None

        # Barra因子数据（用于残差动量因子）
        try:
            self.barra_factor_returns_df = data_loader.load_barra_factor_returns()
        except FileNotFoundError:
            self.barra_factor_returns_df = None

        if load_constituent:
            self._load_constituent_data()

        # 自动计算所有数据源的日期交集
        self._align_date_range()

        # 记录数据交集范围
        self.data_start = self.prices_df.index.min()
        self.data_end = self.prices_df.index.max()

        # 根据 end_date 筛选数据
        if end_date:
            self._filter_by_end_date(end_date)

        # 记录原始数据的最后日期
        self.original_last_date = self.prices_df.index[-1] if len(self.prices_df) > 0 else None
        self.last_complete_month_end = None

        # 排除最后一个不完整的月份（用于月度调仓）
        if exclude_incomplete_month and len(self.prices_df) > 0:
            self._exclude_incomplete_month()

        # 计算回测日期范围（不含Barra的因子）
        self.first_holding_date = None
        self.last_holding_date = None
        if backtest_years is not None and self.last_complete_month_end is not None:
            self._calculate_backtest_dates(backtest_years)

        # 计算Barra因子的回测日期范围（含Barra的因子）
        self.barra_first_holding_date = None
        self.barra_last_holding_date = None
        self.barra_last_complete_month_end = None
        if backtest_years is not None and self.barra_common_end is not None:
            self._calculate_barra_backtest_dates(backtest_years)

        print(f"数据加载完成: {self.prices_df.shape[0]}个交易日, {self.prices_df.shape[1]}个行业")

        # 打印回测日期信息
        if self.first_holding_date and self.last_holding_date:
            print(f"  回测日期范围(不含Barra):")
            print(f"    最早持仓日: {self.first_holding_date}")
            print(f"    最晚持仓日: {self.last_holding_date.strftime('%Y-%m-%d')}")

        if self.barra_first_holding_date and self.barra_last_holding_date:
            print(f"  回测日期范围(含Barra):")
            print(f"    最早持仓日: {self.barra_first_holding_date}")
            print(f"    最晚持仓日: {self.barra_last_holding_date.strftime('%Y-%m-%d')}")

        if self.last_complete_month_end and self.original_last_date:
            if self.last_complete_month_end != self.original_last_date:
                print(f"  注意: 已排除不完整月份，收益计算截止到 {self.last_complete_month_end.date()}")
    def _load_constituent_data(self):
        """
        加载成分股相关数据
        """
        try:
            self.constituent_df = data_loader.load_constituent_df()
            self.stock_price_df = data_loader.load_stock_price_df()
            self.stock_mv_df = data_loader.load_stock_mv_df()
            self.industry_code_df = data_loader.load_industry_code_df()
        except FileNotFoundError as e:
            print(f"  警告: 成分股数据加载失败 - {e}")
    
    def has_constituent_data(self):
        """
        检查是否已加载成分股数据
        """
        return (self.constituent_df is not None and
                self.stock_price_df is not None and
                self.stock_mv_df is not None and
                self.industry_code_df is not None)

    def _align_date_range(self):
        """
        自动计算数据源的日期交集
        分别计算两个日期范围：
        1. 不含Barra的日期范围（用于大多数因子）
        2. 含Barra的日期范围（用于需要Barra的因子，如momentum_residual）
        """
        # 收集基础数据源的日期范围（不含Barra）
        base_date_ranges = []

        # 行业指数数据（必须有）
        base_date_ranges.append((self.prices_df.index.min(), self.prices_df.index.max(), '行业指数'))

        # 成分股数据（如果有）
        if self.stock_price_df is not None:
            base_date_ranges.append((self.stock_price_df.index.min(),
                               self.stock_price_df.index.max(), '个股价格'))
        if self.stock_mv_df is not None:
            base_date_ranges.append((self.stock_mv_df.index.min(),
                               self.stock_mv_df.index.max(), '个股市值'))

        # 计算不含Barra的交集（用于大多数因子）
        common_start_no_barra = max(r[0] for r in base_date_ranges)
        common_end_no_barra = min(r[1] for r in base_date_ranges)

        # 打印不含Barra的数据交集范围
        print(f"  数据交集范围(不含Barra): {common_start_no_barra.date()} 至 {common_end_no_barra.date()}")

        # 计算含Barra的交集（用于需要Barra的因子）
        if self.barra_factor_returns_df is not None:
            barra_date_ranges = base_date_ranges.copy()
            barra_date_ranges.append((self.barra_factor_returns_df.index.min(),
                                     self.barra_factor_returns_df.index.max(), 'Barra因子'))
            common_start_with_barra = max(r[0] for r in barra_date_ranges)
            common_end_with_barra = min(r[1] for r in barra_date_ranges)
            print(f"  数据交集范围(含Barra): {common_start_with_barra.date()} 至 {common_end_with_barra.date()}")

            # 保存Barra日期范围
            self.barra_common_start = common_start_with_barra
            self.barra_common_end = common_end_with_barra
        else:
            self.barra_common_start = None
            self.barra_common_end = None

        # 使用不含Barra的日期范围作为主数据范围
        common_start = common_start_no_barra
        common_end = common_end_no_barra

        # 截断主数据到交集范围（不含Barra）
        self.prices_df = self.prices_df[(self.prices_df.index >= common_start) & (self.prices_df.index <= common_end)]
        self.high_df = self.high_df[(self.high_df.index >= common_start) & (self.high_df.index <= common_end)]
        self.low_df = self.low_df[(self.low_df.index >= common_start) & (self.low_df.index <= common_end)]
        self.turnover_df = self.turnover_df[(self.turnover_df.index >= common_start) & (self.turnover_df.index <= common_end)]
        self.volume_df = self.volume_df[(self.volume_df.index >= common_start) & (self.volume_df.index <= common_end)]
        self.amount_df = self.amount_df[(self.amount_df.index >= common_start) & (self.amount_df.index <= common_end)]

        # Barra数据保持原样，不截断（在使用时根据需要截断）
        # 这样需要Barra的因子可以使用更短的日期范围

        if self.stock_price_df is not None:
            self.stock_price_df = self.stock_price_df[(self.stock_price_df.index >= common_start) & (self.stock_price_df.index <= common_end)]
        if self.stock_mv_df is not None:
            self.stock_mv_df = self.stock_mv_df[(self.stock_mv_df.index >= common_start) & (self.stock_mv_df.index <= common_end)]

    def _exclude_incomplete_month(self):
        """
        排除最后一个不完整的月份
        """
        # 获取最后一个完整月份的月末日期
        self.last_complete_month_end = get_last_complete_month_end(self.prices_df.index)

        if self.last_complete_month_end is None:
            print("警告: 无法确定完整月份，使用全部数据")
            return

        # 如果最后一个完整月末与数据最后日期相同，说明数据是完整的
        if self.last_complete_month_end == self.prices_df.index[-1]:
            return

        # 截断操作在回测时进行，这里只记录最后完整月末日期

    def _filter_by_end_date(self, end_date):
        """
        根据截止日期筛选数据
        """
        self.prices_df = self.prices_df[self.prices_df.index <= end_date]
        self.high_df = self.high_df[self.high_df.index <= end_date]
        self.low_df = self.low_df[self.low_df.index <= end_date]
        self.turnover_df = self.turnover_df[self.turnover_df.index <= end_date]
        self.volume_df = self.volume_df[self.volume_df.index <= end_date]
        self.amount_df = self.amount_df[self.amount_df.index <= end_date]
        if self.stock_price_df is not None:
            self.stock_price_df = self.stock_price_df[self.stock_price_df.index <= end_date]
        if self.stock_mv_df is not None:
            self.stock_mv_df = self.stock_mv_df[self.stock_mv_df.index <= end_date]
        if self.barra_factor_returns_df is not None:
            self.barra_factor_returns_df = self.barra_factor_returns_df[self.barra_factor_returns_df.index <= end_date]

    def _calculate_backtest_dates(self, backtest_years):
        """
        计算回测日期范围

        参数:
        backtest_years: int, 回测年限
        """
        import calendar

        # 最晚持仓日 = 最后一个完整月末
        self.last_holding_date = self.last_complete_month_end

        # 最早持仓日 = 往前推 backtest_years 年的12月31日
        first_holding_year = self.last_holding_date.year - backtest_years
        self.first_holding_date = f"{first_holding_year}-12-31"

        # 检查数据是否足够
        first_holding_ts = pd.Timestamp(self.first_holding_date)
        if first_holding_ts < self.data_start:
            # 数据不够，从数据起始后的第一个12月31日开始
            first_year = self.data_start.year
            if self.data_start.month > 1 or (self.data_start.month == 1 and self.data_start.day > 1):
                first_year += 1
            self.first_holding_date = f"{first_year}-12-31"

    def _calculate_barra_backtest_dates(self, backtest_years):
        """
        计算需要Barra因子的回测日期范围

        Barra因子的最早持仓日与不含Barra的因子保持一致，
        只是最晚持仓日受Barra数据结束日期限制

        参数:
        backtest_years: int, 回测年限
        """
        if self.barra_common_end is None:
            return

        # 获取Barra日期范围内的最后一个完整月末
        barra_prices = self.prices_df[
            (self.prices_df.index >= self.barra_common_start) &
            (self.prices_df.index <= self.barra_common_end)
        ]
        self.barra_last_complete_month_end = get_last_complete_month_end(barra_prices.index)

        if self.barra_last_complete_month_end is None:
            return

        # 最晚持仓日 = Barra范围内的最后一个完整月末
        self.barra_last_holding_date = self.barra_last_complete_month_end

        # 最早持仓日 = 与不含Barra的因子保持一致
        # 使用不含Barra的最早持仓日
        self.barra_first_holding_date = self.first_holding_date

        # 检查数据是否足够
        if self.barra_first_holding_date is not None:
            first_holding_ts = pd.Timestamp(self.barra_first_holding_date)
            if first_holding_ts < self.barra_common_start:
                # 数据不够，从Barra数据起始后的第一个12月31日开始
                first_year = self.barra_common_start.year
                if self.barra_common_start.month > 1 or (self.barra_common_start.month == 1 and self.barra_common_start.day > 1):
                    first_year += 1
                self.barra_first_holding_date = f"{first_year}-12-31"


def get_factor_docstring(factor_name):
    """
    获取因子函数的docstring作为因子说明
    
    参数:
    factor_name: str, 因子名称
    
    返回:
    str: 因子说明文档
    """
    if factor_name not in FACTOR_CONFIG:
        return "未找到因子说明"
    
    func = FACTOR_CONFIG[factor_name]['func']
    docstring = inspect.getdoc(func)
    return docstring if docstring else "无说明文档"


def get_factor_config(factor_name):
    """
    获取因子的完整配置信息
    
    参数:
    factor_name: str, 因子名称
    
    返回:
    dict: 因子配置，包含 func, base_warmup, window_multiplier, windows, requires_constituent, description
    """
    if factor_name not in FACTOR_CONFIG:
        return None
    return FACTOR_CONFIG[factor_name]


def print_factor_config_summary():
    """
    打印因子配置摘要
    """
    print("=" * 80)
    print("因子配置摘要")
    print("=" * 80)
    print(f"{'因子名称':<40} {'需要成分股':<12} {'预热期配置':<20}")
    print("-" * 80)
    for name, config in FACTOR_CONFIG.items():
        requires = "是" if config.get('requires_constituent', False) else "否"
        warmup_info = f"base={config['base_warmup']}, mult={config['window_multiplier']}"
        print(f"{name:<40} {requires:<12} {warmup_info:<20}")
    print("=" * 80)
    print(f"总计: {len(FACTOR_CONFIG)} 个因子")
    constituent_count = sum(1 for c in FACTOR_CONFIG.values() if c.get('requires_constituent', False))
    print(f"需要成分股数据的因子: {constituent_count} 个")


def calculate_factor_warmup_period(factor_name, window):
    """
    计算单个因子在指定窗口下的预热期（交易日数）

    参数:
    factor_name: str, 因子名称
    window: int, 因子计算窗口

    返回:
    int: 预热期（交易日数）
    """
    if factor_name not in FACTOR_CONFIG:
        # 默认只需要 window 天
        return window

    config = FACTOR_CONFIG[factor_name]
    base_warmup = config.get('base_warmup', 0)
    window_multiplier = config.get('window_multiplier', 1)

    return base_warmup + window * window_multiplier


def get_factor_windows(factor_name, default_windows=LOOKBACK_WINDOWS):
    """
    获取因子适用的回看窗口列表

    参数:
    factor_name: str, 因子名称
    default_windows: list, 默认回看窗口列表

    返回:
    list: 该因子适用的回看窗口列表
    """
    if factor_name not in FACTOR_CONFIG:
        return default_windows

    config = FACTOR_CONFIG[factor_name]
    factor_windows = config.get('lookback_windows', None)

    if factor_windows is None:
        return default_windows

    return factor_windows


def calculate_factor_unified_start_date(data: 'DataContainer', factor_name, windows,
                                         rebalance_freq=DEFAULT_REBALANCE_FREQ,
                                         monthly_rebalance=DEFAULT_MONTHLY_REBALANCE):
    """
    计算单个因子在所有窗口下的统一回测起始日期
    
    选择该因子在所有窗口中预热期最长的那个，然后找到该日期之后的第一个月末作为统一起始日期
    
    参数:
    data: DataContainer, 数据容器
    factor_name: str, 因子名称
    windows: list, 窗口列表
    rebalance_freq: int, 调仓频率
    monthly_rebalance: bool, 是否按月调仓
    
    返回:
    pd.Timestamp: 该因子的统一回测起始日期（月末日期）
    """
    all_dates = data.prices_df.index
    
    # 计算该因子在所有窗口下的最大预热期
    max_warmup = 0
    max_warmup_window = None
    
    for window in windows:
        warmup = calculate_factor_warmup_period(factor_name, window)
        if warmup > max_warmup:
            max_warmup = warmup
            max_warmup_window = window
    
    # 确保有足够的数据
    if max_warmup >= len(all_dates):
        raise ValueError(f"因子 {factor_name} 数据不足：需要至少 {max_warmup} 个交易日，但只有 {len(all_dates)} 个")
    
    # 找到预热期结束后的第一个日期
    warmup_end_date = all_dates[max_warmup]
    
    if monthly_rebalance:
        # 找到预热期结束后的第一个月末日期
        monthly_dates = get_monthly_rebalance_dates(all_dates)
        # 找到大于等于 warmup_end_date 的第一个月末
        valid_monthly_dates = monthly_dates[monthly_dates >= warmup_end_date]
        
        if len(valid_monthly_dates) == 0:
            raise ValueError(f"因子 {factor_name} 预热期结束后没有有效的月末日期")
        
        unified_start_date = valid_monthly_dates[0]
    else:
        unified_start_date = warmup_end_date
    
    return unified_start_date, max_warmup, max_warmup_window


def compute_factor(factor_name, data: DataContainer, window, rebalance_freq=DEFAULT_REBALANCE_FREQ, split_ratio=None, train_periods=None):
    """
    计算单个因子值

    参数:
    factor_name: str, 因子名称
    data: DataContainer, 数据容器（包含所有需要的数据）
    window: int, 回溯窗口
    rebalance_freq: int, 调仓频率
    split_ratio: float, 龙头/跟随分割参数（仅用于momentum_lead_lag_enhanced因子）
    train_periods: int, 训练期（月数），仅用于lasso因子

    返回:
    pd.DataFrame: 因子值
    """
    if factor_name not in FACTOR_CONFIG:
        raise ValueError(f"未知因子名称: {factor_name}. 可选: {list(FACTOR_CONFIG.keys())}")

    # 检查是否需要成分股数据
    factor_config = FACTOR_CONFIG[factor_name]
    if factor_config.get('requires_constituent', False) and not data.has_constituent_data():
        raise ValueError(f"因子 '{factor_name}' 需要成分股数据，但 DataContainer 未加载成分股数据。"
                        f"请使用 DataContainer(load_constituent=True) 初始化。")

    factor_func = factor_config['func']
    sig = inspect.signature(factor_func)
    param_names = list(sig.parameters.keys())

    # 计算基准收益率（用于需要的因子）
    benchmark_returns = calculate_benchmark_returns(data.prices_df, rebalance_freq)

    # 构建参数字典 - 包含所有可能需要的参数
    available_params = {
        # 数据参数
        'prices_df': data.prices_df,
        'high_df': data.high_df,
        'low_df': data.low_df,
        'turnover_df': data.turnover_df,
        'volume_df': data.volume_df,
        'amount_df': data.amount_df,  # 使用volume_df作为代理
        # 成分股数据参数
        'constituent_df': data.constituent_df,
        'stock_price_df': data.stock_price_df,
        'stock_mv_df': data.stock_mv_df,
        'industry_code_df': data.industry_code_df,
        'industry_code_file': data_loader.DEFAULT_INDUSTRY_CODE_FILE,
        # 窗口参数（所有因子统一使用 window）
        'window': window,
        'rebalance_freq': rebalance_freq,
        # 其他参数（仅保留无默认值的必需参数）
        'benchmark_returns': benchmark_returns,
        # 残差动量因子参数（月频版本使用价格数据）
        'industry_prices_df': data.prices_df,
        'barra_factor_returns_df': data.barra_factor_returns_df,
    }

    # 如果指定了split_ratio，添加到参数字典；否则从配置中读取default_split_ratio
    if split_ratio is not None:
        available_params['split_ratio'] = split_ratio
    elif 'default_split_ratio' in factor_config:
        available_params['split_ratio'] = factor_config['default_split_ratio']

    # 根据函数签名自动选择参数
    call_kwargs = {}
    for param_name in param_names:
        if param_name in available_params:
            call_kwargs[param_name] = available_params[param_name]
        else:
            param = sig.parameters[param_name]
            if param.default is inspect.Parameter.empty:
                raise ValueError(f"因子 '{factor_name}' 需要参数 '{param_name}'，但未定义。")

    return factor_func(**call_kwargs)


def calculate_benchmark_returns(prices_df, rebalance_freq):
    """
    计算基准收益率（申万一级行业等权）
    
    参数:
    prices_df: pd.DataFrame, 价格数据
    rebalance_freq: int, 调仓频率
    
    返回:
    pd.Series: 基准收益率序列（每期）
    """
    # 计算每个行业的期收益率
    period_returns = prices_df.pct_change(rebalance_freq)
    # 等权平均作为基准
    benchmark_returns = period_returns.mean(axis=1)
    return benchmark_returns


def get_monthly_rebalance_dates(all_trade_dates):
    """
    根据给定的所有交易日期，生成每月最后一个交易日作为调仓日期。
    
    参数:
    all_trade_dates: pd.DatetimeIndex, 所有的交易日期序列
    
    返回:
    pd.DatetimeIndex: 每月最后一个交易日序列
    """
    # 确保日期是DatetimeIndex类型
    if not isinstance(all_trade_dates, pd.DatetimeIndex):
        all_trade_dates = pd.to_datetime(all_trade_dates)

    # 按照年月分组，取每个月的最后一个日期
    # 由于all_trade_dates已经是交易日序列，直接取每个月最后一个日期就是该月的最后一个交易日
    monthly_last_dates = all_trade_dates.to_series().groupby([all_trade_dates.year, all_trade_dates.month]).apply(lambda x: x.iloc[-1])
    return pd.DatetimeIndex(monthly_last_dates)


def get_weekly_rebalance_dates(all_trade_dates):
    """
    根据给定的所有交易日期，生成每周最后一个交易日作为调仓日期。
    
    参数:
    all_trade_dates: pd.DatetimeIndex, 所有的交易日期序列
    
    返回:
    pd.DatetimeIndex: 每周最后一个交易日序列
    """
    # 确保日期是DatetimeIndex类型
    if not isinstance(all_trade_dates, pd.DatetimeIndex):
        all_trade_dates = pd.to_datetime(all_trade_dates)

    # 按照年周分组，取每周的最后一个日期
    # 使用 isocalendar() 获取年份和周数
    dates_series = all_trade_dates.to_series()
    year_week = dates_series.apply(lambda x: (x.isocalendar()[0], x.isocalendar()[1]))
    weekly_last_dates = dates_series.groupby(year_week).apply(lambda x: x.iloc[-1])
    return pd.DatetimeIndex(weekly_last_dates)


def get_rebalance_dates_by_type(all_trade_dates, rebalance_type='monthly', rebalance_freq=20):
    """
    根据调仓类型获取调仓日期
    
    参数:
    all_trade_dates: pd.DatetimeIndex, 所有的交易日期序列
    rebalance_type: str, 调仓类型: 'monthly', 'weekly', 'fixed'
    rebalance_freq: int, 固定调仓频率（仅在 rebalance_type='fixed' 时使用）
    
    返回:
    pd.DatetimeIndex: 调仓日期序列
    """
    if rebalance_type == 'monthly':
        return get_monthly_rebalance_dates(all_trade_dates)
    elif rebalance_type == 'weekly':
        return get_weekly_rebalance_dates(all_trade_dates)
    elif rebalance_type == 'fixed':
        rebalance_indices = list(range(0, len(all_trade_dates), rebalance_freq))
        return all_trade_dates[rebalance_indices]
    else:
        raise ValueError(f"未知的调仓类型: {rebalance_type}. 可选: 'monthly', 'weekly', 'fixed'")


def get_last_complete_month_end(all_trade_dates):
    """
    获取最后一个完整月份的月末日期。

    判断逻辑：
    1. 如果数据最后日期不是该月最后几个交易日（通常月末最后一个交易日在25号之后），
       则认为该月不完整
    2. 保守起见，如果数据最后日期的日期号小于25，认为该月不完整

    参数:
    all_trade_dates: pd.DatetimeIndex, 所有的交易日期序列

    返回:
    pd.Timestamp: 最后一个完整月份的月末日期
    """
    if not isinstance(all_trade_dates, pd.DatetimeIndex):
        all_trade_dates = pd.to_datetime(all_trade_dates)

    if len(all_trade_dates) == 0:
        return None

    # 获取所有月末日期
    monthly_ends = get_monthly_rebalance_dates(all_trade_dates)

    # 数据的最后一个日期
    last_date = all_trade_dates[-1]

    # 判断最后一个月是否完整：
    # 1. 如果最后日期的日期号 < 25，认为该月不完整（月末交易日通常在25号之后）
    # 2. 如果最后日期的日期号 >= 28，认为该月完整（已经是月末）
    # 3. 如果在25-27之间，需要进一步判断
    if last_date.day < 25:
        # 该月明显不完整，返回上一个月末
        complete_month_ends = monthly_ends[monthly_ends < last_date]
        if len(complete_month_ends) > 0:
            return complete_month_ends[-1]
        else:
            return None
    elif last_date.day >= 28:
        # 该月基本完整（28号及以后）
        return last_date if last_date in monthly_ends.values else monthly_ends[monthly_ends <= last_date][-1]
    else:
        # 25-27号之间，保守处理，认为不完整
        complete_month_ends = monthly_ends[monthly_ends < last_date]
        if len(complete_month_ends) > 0:
            return complete_month_ends[-1]
        else:
            return None


def is_month_complete(all_trade_dates, check_date):
    """
    检查指定月份是否完整（即该月的最后一个交易日是否在数据中）
    
    参数:
    all_trade_dates: pd.DatetimeIndex, 所有的交易日期序列
    check_date: pd.Timestamp, 要检查的日期
    
    返回:
    bool: 该月是否完整
    """
    if not isinstance(all_trade_dates, pd.DatetimeIndex):
        all_trade_dates = pd.to_datetime(all_trade_dates)
    
    # 获取该月的所有交易日
    year, month = check_date.year, check_date.month
    month_dates = all_trade_dates[(all_trade_dates.year == year) & (all_trade_dates.month == month)]
    
    if len(month_dates) == 0:
        return False
    
    # 获取该月在数据中的最后一个交易日
    last_trade_day_in_data = month_dates[-1]
    
    # 检查这是否是该月的最后一个交易日
    # 方法：检查下一个交易日是否在下个月
    last_date_idx = all_trade_dates.get_loc(last_trade_day_in_data)
    
    # 如果是数据的最后一天，需要判断是否是月末
    if last_date_idx == len(all_trade_dates) - 1:
        # 检查日期是否接近月末（最后5个自然日内）
        import calendar
        _, last_day = calendar.monthrange(year, month)
        return last_trade_day_in_data.day >= last_day - 5
    else:
        # 检查下一个交易日是否在下个月
        next_trade_day = all_trade_dates[last_date_idx + 1]
        return next_trade_day.month != month or next_trade_day.year != year


def calculate_monthly_forward_returns(prices_df, all_trade_dates):
    """
    计算月度未来收益率（从当前月末到下一个月末的收益率）
    
    参数:
    prices_df: pd.DataFrame, 价格数据
    all_trade_dates: pd.DatetimeIndex, 所有交易日期
    
    返回:
    pd.DataFrame: 月度未来收益率 (index=日期, columns=行业)
    """
    # 获取每月最后一个交易日
    monthly_dates = get_monthly_rebalance_dates(all_trade_dates)
    
    # 计算每个月末到下一个月末的收益率
    forward_returns = pd.DataFrame(index=prices_df.index, columns=prices_df.columns, dtype=float)
    
    for i, date in enumerate(monthly_dates[:-1]):
        next_date = monthly_dates[i + 1]
        if date in prices_df.index and next_date in prices_df.index:
            # 计算从当前月末到下一个月末的收益率
            ret = (prices_df.loc[next_date] / prices_df.loc[date]) - 1
            forward_returns.loc[date] = ret
    
    return forward_returns


def calculate_weekly_forward_returns(prices_df, all_trade_dates):
    """
    计算周度未来收益率（从当前周末到下一个周末的收益率）
    
    参数:
    prices_df: pd.DataFrame, 价格数据
    all_trade_dates: pd.DatetimeIndex, 所有交易日期
    
    返回:
    pd.DataFrame: 周度未来收益率 (index=日期, columns=行业)
    """
    # 获取每周最后一个交易日
    weekly_dates = get_weekly_rebalance_dates(all_trade_dates)
    
    # 计算每个周末到下一个周末的收益率
    forward_returns = pd.DataFrame(index=prices_df.index, columns=prices_df.columns, dtype=float)
    
    for i, date in enumerate(weekly_dates[:-1]):
        next_date = weekly_dates[i + 1]
        if date in prices_df.index and next_date in prices_df.index:
            # 计算从当前周末到下一个周末的收益率
            ret = (prices_df.loc[next_date] / prices_df.loc[date]) - 1
            forward_returns.loc[date] = ret
    
    return forward_returns


def calculate_forward_returns_by_type(prices_df, all_trade_dates, rebalance_type='monthly', rebalance_freq=20):
    """
    根据调仓类型计算未来收益率
    
    参数:
    prices_df: pd.DataFrame, 价格数据
    all_trade_dates: pd.DatetimeIndex, 所有交易日期
    rebalance_type: str, 调仓类型: 'monthly', 'weekly', 'fixed'
    rebalance_freq: int, 固定调仓频率（仅在 rebalance_type='fixed' 时使用）
    
    返回:
    pd.DataFrame: 未来收益率 (index=日期, columns=行业)
    """
    if rebalance_type == 'monthly':
        return calculate_monthly_forward_returns(prices_df, all_trade_dates)
    elif rebalance_type == 'weekly':
        return calculate_weekly_forward_returns(prices_df, all_trade_dates)
    elif rebalance_type == 'fixed':
        # 固定频率调仓
        forward_returns_df = prices_df.pct_change(rebalance_freq).shift(-rebalance_freq)
        return forward_returns_df
    else:
        raise ValueError(f"未知的调仓类型: {rebalance_type}")


def calc_pearson_ic(factor_series, return_series):
    """
    计算Pearson IC (Pearson相关系数)

    参数:
        factor_series: pd.Series, 因子值序列
        return_series: pd.Series, 收益率序列

    返回:
        float, Pearson IC值
    """
    valid_mask = factor_series.notna() & return_series.notna()
    if valid_mask.sum() < 3:
        return np.nan

    factor_valid = factor_series[valid_mask]
    return_valid = return_series[valid_mask]

    ic, _ = stats.pearsonr(factor_valid, return_valid)
    return ic


def calc_rank_ic(factor_series, return_series):
    """
    计算Rank IC (Spearman相关系数)

    参数:
        factor_series: pd.Series, 因子值序列
        return_series: pd.Series, 收益率序列

    返回:
        float, Rank IC值
    """
    valid_mask = factor_series.notna() & return_series.notna()
    if valid_mask.sum() < 3:
        return np.nan

    factor_valid = factor_series[valid_mask]
    return_valid = return_series[valid_mask]

    ic, _ = stats.spearmanr(factor_valid, return_valid)
    return ic


def calculate_ic_ir(factor_df, forward_returns_df, rebalance_freq=DEFAULT_REBALANCE_FREQ, monthly_rebalance=DEFAULT_MONTHLY_REBALANCE, unified_start_date=None, rebalance_type=None):
    """
    计算因子的 Pearson IC 和 Rank IC 及其 IR

    参数:
        factor_df: pd.DataFrame, 因子值 (index=日期, columns=行业)
        forward_returns_df: pd.DataFrame, 未来收益率 (index=日期, columns=行业)
        rebalance_freq: int, 调仓频率（交易日），默认20。如果 monthly_rebalance 为 True，此参数无效。
        monthly_rebalance: bool, 是否按月调仓，默认为 True (每月最后一个交易日调仓)
        unified_start_date: pd.Timestamp, 统一的回测起始日期（如果指定，只计算该日期之后的IC）
        rebalance_type: str, 调仓类型: 'monthly', 'weekly', 'fixed'（优先级高于 monthly_rebalance）

    返回:
        dict: 包含 Pearson IC 和 Rank IC 的所有指标
    """
    all_dates = factor_df.index

    # 优先使用 rebalance_type 参数
    if rebalance_type is not None:
        rebalance_dates = get_rebalance_dates_by_type(all_dates, rebalance_type, rebalance_freq)
    elif monthly_rebalance:
        # 按月调仓
        rebalance_dates = get_monthly_rebalance_dates(all_dates)
    else:
        # 按固定频率调仓
        rebalance_indices = list(range(0, len(all_dates), rebalance_freq))
        rebalance_dates = all_dates[rebalance_indices]

    # 如果指定了统一起始日期，过滤调仓日期
    if unified_start_date is not None:
        rebalance_dates = rebalance_dates[rebalance_dates >= unified_start_date]

    # 计算每个调仓日的 Pearson IC 和 Rank IC
    pearson_ic_list = []
    rank_ic_list = []
    ic_dates = []

    for date in rebalance_dates:
        if date not in factor_df.index or date not in forward_returns_df.index:
            continue

        pearson_ic = calc_pearson_ic(factor_df.loc[date], forward_returns_df.loc[date])
        rank_ic = calc_rank_ic(factor_df.loc[date], forward_returns_df.loc[date])

        if not np.isnan(pearson_ic) and not np.isnan(rank_ic):
            pearson_ic_list.append(pearson_ic)
            rank_ic_list.append(rank_ic)
            ic_dates.append(date)

    # 构建IC时间序列
    pearson_ic_series = pd.Series(pearson_ic_list, index=ic_dates)
    rank_ic_series = pd.Series(rank_ic_list, index=ic_dates)

    # 计算IC累积序列
    pearson_ic_cumsum = pearson_ic_series.cumsum()
    rank_ic_cumsum = rank_ic_series.cumsum()

    # 计算 Pearson IC 统计指标
    pearson_ic_mean = pearson_ic_series.mean() if len(pearson_ic_series) > 0 else np.nan
    pearson_ic_std = pearson_ic_series.std() if len(pearson_ic_series) > 0 else np.nan
    pearson_icir = pearson_ic_mean / pearson_ic_std if pearson_ic_std > 0 else np.nan

    # 计算 Rank IC 统计指标
    rank_ic_mean = rank_ic_series.mean() if len(rank_ic_series) > 0 else np.nan
    rank_ic_std = rank_ic_series.std() if len(rank_ic_series) > 0 else np.nan
    rank_icir = rank_ic_mean / rank_ic_std if rank_ic_std > 0 else np.nan

    # IC胜率（使用 Rank IC）
    ic_win_rate = (rank_ic_series > 0).sum() / len(rank_ic_series) if len(rank_ic_series) > 0 else np.nan

    return {
        # Pearson IC 相关
        'pearson_ic_series': pearson_ic_series,
        'pearson_ic_cumsum': pearson_ic_cumsum,
        'pearson_ic_mean': pearson_ic_mean,
        'pearson_icir': pearson_icir,
        # Rank IC 相关
        'rank_ic_series': rank_ic_series,
        'rank_ic_cumsum': rank_ic_cumsum,
        'rank_ic_mean': rank_ic_mean,
        'rank_icir': rank_icir,
        # 通用指标
        'ic_win_rate': ic_win_rate,
    }


def get_latest_month_holdings(factor_df, prices_df, window, n_layers=N_LAYERS, monthly_rebalance=DEFAULT_MONTHLY_REBALANCE, last_complete_month_end=None):
    """
    获取最后一个完整月末的分层持仓（不包括不完整月份）
    用于补充回测中可能遗漏的最后一个调仓日持仓

    参数:
    factor_df: pd.DataFrame, 因子值 (index=日期, columns=行业)
    prices_df: pd.DataFrame, 价格数据
    window: int, 回溯窗口
    n_layers: int, 分层数
    monthly_rebalance: bool, 是否按月调仓
    last_complete_month_end: pd.Timestamp, 最后一个完整月末日期

    返回:
    dict: {layer_idx: {date: [行业列表]}} 最后一个完整月末的持仓
    """
    all_dates = factor_df.index

    if not monthly_rebalance:
        return {}

    # 获取所有月末日期
    all_month_ends = get_monthly_rebalance_dates(all_dates)

    if len(all_month_ends) == 0:
        return {}

    # 如果指定了最后一个完整月末，只获取该日期的持仓
    # 否则获取最后一个月末的持仓
    if last_complete_month_end is not None:
        target_date = last_complete_month_end
    else:
        target_date = all_month_ends[-1]

    # 确保该日期在因子数据中存在
    if target_date not in factor_df.index:
        return {}

    # 获取该日期的因子值
    fac = factor_df.loc[target_date]
    valid_mask = ~fac.isna()
    fac = fac[valid_mask]

    if len(fac) < n_layers * 2:
        return {}

    # 按因子值排序并分层
    sorted_assets = fac.sort_values(ascending=True)
    n_assets = len(sorted_assets)
    layer_size = n_assets // n_layers

    latest_holdings = {i: {} for i in range(n_layers)}

    for layer in range(n_layers):
        start_idx = layer * layer_size
        end_idx = start_idx + layer_size if layer < n_layers - 1 else n_assets
        layer_assets = sorted_assets.index[start_idx:end_idx].tolist()
        latest_holdings[layer][target_date] = layer_assets

    return latest_holdings


def stratified_backtest(factor_df, prices_df, window, rebalance_freq=DEFAULT_REBALANCE_FREQ, n_layers=N_LAYERS, monthly_rebalance=DEFAULT_MONTHLY_REBALANCE, last_complete_month_end=None, unified_start_date=None, rebalance_type=None):
    """
    分层回测：根据因子值将资产分成n层，计算每层的收益率累计净值
    同时返回每期各层选中的行业
    
    参数:
    factor_df: pd.DataFrame, 因子值 (index=日期, columns=行业)
    prices_df: pd.DataFrame, 价格数据
    window: int, 回溯窗口（因子计算的回溯窗口）
    rebalance_freq: int, 调仓频率（交易日），默认20。如果 monthly_rebalance 为 True，此参数无效。
    n_layers: int, 分层数
    monthly_rebalance: bool, 是否按月调仓，默认为 True (每月最后一个交易日调仓)
    last_complete_month_end: pd.Timestamp, 最后一个完整月份的月末日期（用于排除不完整月份的收益计算）
    unified_start_date: pd.Timestamp, 统一的回测起始日期（如果指定，从该日期开始回测）
    rebalance_type: str, 调仓类型: 'monthly', 'weekly', 'fixed'（优先级高于 monthly_rebalance）
    
    返回:
    tuple: (nav_df, layer_returns, layer_holdings_history)
    """
    daily_returns = prices_df.pct_change()
    
    all_dates_factor_df = factor_df.index

    # 优先使用 rebalance_type 参数
    if rebalance_type is not None:
        rebalance_dates_raw = get_rebalance_dates_by_type(all_dates_factor_df, rebalance_type, rebalance_freq)
        
        # 如果指定了统一起始日期，从该日期开始
        if unified_start_date is not None:
            rebalance_dates_raw = rebalance_dates_raw[rebalance_dates_raw >= unified_start_date]
        
        # 如果指定了最后一个完整月末且是月度调仓，排除之后的调仓日期
        if last_complete_month_end is not None and rebalance_type == 'monthly':
            rebalance_dates_raw = rebalance_dates_raw[rebalance_dates_raw <= last_complete_month_end]
    elif monthly_rebalance:
        # 按月调仓，获取所有月末日期
        rebalance_dates_raw = get_monthly_rebalance_dates(all_dates_factor_df)
        
        # 如果指定了统一起始日期，从该日期开始
        if unified_start_date is not None:
            rebalance_dates_raw = rebalance_dates_raw[rebalance_dates_raw >= unified_start_date]
        
        # 如果指定了最后一个完整月末，排除之后的调仓日期
        if last_complete_month_end is not None:
            rebalance_dates_raw = rebalance_dates_raw[rebalance_dates_raw <= last_complete_month_end]
    else:
        # 按固定频率调仓
        # 如果指定了统一起始日期，从该日期开始
        if unified_start_date is not None:
            start_idx = all_dates_factor_df.get_loc(unified_start_date) if unified_start_date in all_dates_factor_df else 0
            effective_start_dates = all_dates_factor_df[start_idx:]
        else:
            # 默认从 window 位置开始
            start_date_idx = min(window, len(all_dates_factor_df) - 1)
            effective_start_dates = all_dates_factor_df[start_date_idx:]
        
        rebalance_indices = list(range(0, len(effective_start_dates), rebalance_freq))
        rebalance_dates_raw = effective_start_dates[rebalance_indices]
    
    # 确保调仓日期在 prices_df 中存在
    valid_dates = [d for d in rebalance_dates_raw if d in prices_df.index]

    if not valid_dates:
        nav_df = pd.DataFrame()
        return nav_df, pd.DataFrame(), {}

    layer_nav = {i: [1.0] for i in range(n_layers)}
    nav_dates = [valid_dates[0]]  # 第一个调仓日，净值=1.0
    layer_holdings_history = {i: {} for i in range(n_layers)}  # 记录每期每层持仓

    for i, date in enumerate(valid_dates[:-1]):
        next_date = valid_dates[i + 1]

        fac = factor_df.loc[date]
        valid_mask = ~fac.isna()
        fac = fac[valid_mask]

        if len(fac) < n_layers * 2:
            for layer in range(n_layers):
                layer_nav[layer].append(layer_nav[layer][-1])
            nav_dates.append(next_date)
            continue

        # 按因子值排序并分层（升序：G1=因子值最小，G5=因子值最大）
        sorted_assets = fac.sort_values(ascending=True)
        n_assets = len(sorted_assets)
        layer_size = n_assets // n_layers

        holding_period = daily_returns.loc[date:next_date].iloc[1:]
        for layer in range(n_layers):
            start_idx = layer * layer_size
            end_idx = start_idx + layer_size if layer < n_layers - 1 else n_assets
            layer_assets = sorted_assets.index[start_idx:end_idx].tolist()

            # 记录持仓（在调仓日记录）
            layer_holdings_history[layer][date] = layer_assets

            if layer_assets and len(holding_period) > 0:
                layer_ret = holding_period[layer_assets].mean(axis=1)
                cumulative_ret = (1 + layer_ret).prod() - 1
                layer_nav[layer].append(layer_nav[layer][-1] * (1 + cumulative_ret))
            else:
                layer_nav[layer].append(layer_nav[layer][-1])

        nav_dates.append(next_date)
    
    nav_df = pd.DataFrame(layer_nav, index=nav_dates)
    nav_df.columns = [f'G{i+1}' for i in range(n_layers)]
    layer_returns = nav_df.pct_change().dropna()
    
    return nav_df, layer_returns, layer_holdings_history


def calculate_excess_metrics(nav_df, benchmark_nav, rebalance_freq=DEFAULT_REBALANCE_FREQ, monthly_rebalance=DEFAULT_MONTHLY_REBALANCE, rebalance_type=None):
    """
    计算超额收益指标
    
    参数:
    nav_df: pd.DataFrame, 各层净值
    benchmark_nav: pd.Series, 基准净值
    rebalance_freq: int, 调仓频率（仅在非月度调仓时使用）
    monthly_rebalance: bool, 是否按月调仓
    rebalance_type: str, 调仓类型: 'monthly', 'weekly', 'fixed'
    
    返回:
    dict: 各层的超额收益指标
    """
    results = {}

    start_date = nav_df.index[0]
    end_date = nav_df.index[-1]

    # 根据调仓类型计算年数和每年周期数
    if rebalance_type == 'weekly':
        periods_per_year = 52  # 周度调仓，每年约52周
        years = len(nav_df) / 52  # 用周数计算年数
    elif rebalance_type == 'monthly' or monthly_rebalance:
        periods_per_year = 12  # 月度调仓
        years = len(nav_df) / 12  # 用月数计算年数
    else:
        periods_per_year = 252 / rebalance_freq
        years = len(nav_df) / periods_per_year  # 用周期数计算年数
    
    # 计算基准收益率序列
    benchmark_returns = benchmark_nav.pct_change().dropna()
    
    for col in nav_df.columns:
        nav = nav_df[col]
        
        # 超额净值 = 策略净值 / 基准净值
        excess_nav = nav / benchmark_nav
        excess_returns = excess_nav.pct_change().dropna()
        
        # 超额累计收益率
        excess_total_return = (excess_nav.iloc[-1] / excess_nav.iloc[0] - 1) * 100
        
        # 超额年化收益率
        excess_total_ratio = excess_nav.iloc[-1] / excess_nav.iloc[0]
        excess_annual_return = (excess_total_ratio ** (1 / years) - 1) * 100 if years > 0 else 0
        
        # 绝对收益指标
        total_return = (nav.iloc[-1] / nav.iloc[0] - 1) * 100
        total_ratio = nav.iloc[-1] / nav.iloc[0]
        annual_return = (total_ratio ** (1 / years) - 1) * 100 if years > 0 else 0
        
        returns = nav.pct_change().dropna()
        volatility = returns.std() * np.sqrt(periods_per_year) * 100
        
        # 夏普比率
        sharpe = (returns.mean() / returns.std() * np.sqrt(periods_per_year)) if returns.std() > 0 else 0
        
        # 最大回撤
        cummax = nav.cummax()
        drawdown = (nav - cummax) / cummax
        max_dd = drawdown.min() * 100
        
        # 多头胜率：策略收益 > 基准收益的期数占比
        aligned_returns = returns.align(benchmark_returns, join='inner')
        strategy_ret = aligned_returns[0]
        bench_ret = aligned_returns[1]
        win_count = (strategy_ret > bench_ret).sum()
        total_periods = len(strategy_ret)
        win_rate = (win_count / total_periods * 100) if total_periods > 0 else 0
        
        results[col] = {
            '累计收益率(%)': total_return,
            '年化收益率(%)': annual_return,
            '超额累计收益率(%)': excess_total_return,
            '超额年化收益率(%)': excess_annual_return,
            '年化波动率(%)': volatility,
            '夏普比率': sharpe,
            '最大回撤(%)': max_dd,
            '多头胜率(%)': win_rate,
        }
    
    return results


def calculate_yearly_returns(nav_df, benchmark_nav, start_year=None, last_complete_month_end=None):
    """
    计算每年的收益统计（仅针对G5）

    参数:
    nav_df: pd.DataFrame, 各层净值（包含G5列）
    benchmark_nav: pd.Series, 基准净值
    start_year: int, 起始年份（默认None，自动从数据起始年份开始）
    last_complete_month_end: pd.Timestamp, 最后一个完整月份的月末日期（用于标注不完整年）

    返回:
    pd.DataFrame: 每年的多头收益、超额收益、基准收益
    """
    if 'G5' not in nav_df.columns:
        return pd.DataFrame()

    g5_nav = nav_df['G5']

    # 获取所有年份
    years = sorted(set(nav_df.index.year))

    # 如果未指定起始年份，使用数据的第一个完整年份（第二年开始）
    if start_year is None:
        # 默认从数据的第二年开始（第一年可能不完整）
        if len(years) > 1:
            start_year = years[1]
        else:
            start_year = years[0] if years else 2012

    years = [y for y in years if y >= start_year]

    # 判断最后一年是否完整（12月是否有数据）
    last_year = years[-1] if years else None
    last_year_incomplete = False
    if last_year and last_complete_month_end is not None:
        # 如果最后完整月末不是12月，则最后一年不完整
        if last_complete_month_end.month != 12:
            last_year_incomplete = True

    yearly_data = []

    for i, year in enumerate(years):
        # 获取该年的数据
        year_mask = nav_df.index.year == year
        year_dates = nav_df.index[year_mask]

        if len(year_dates) == 0:
            continue

        # 年度收益的起始日期：
        # - 第一年：使用上一年的最后一个日期（即初始持仓日）
        # - 其他年：使用上一年的最后一个日期
        if i == 0:
            # 第一年：找到该年之前的最后一个日期（初始持仓日）
            prev_year_mask = nav_df.index.year < year
            if prev_year_mask.any():
                start_date = nav_df.index[prev_year_mask][-1]
            else:
                # 如果没有上一年数据，使用该年第一个日期
                start_date = year_dates[0]
        else:
            # 其他年：使用上一年的最后一个日期
            prev_year = years[i - 1]
            prev_year_mask = nav_df.index.year == prev_year
            prev_year_dates = nav_df.index[prev_year_mask]
            if len(prev_year_dates) > 0:
                start_date = prev_year_dates[-1]
            else:
                start_date = year_dates[0]

        end_date = year_dates[-1]

        # G5多头收益
        g5_return = (g5_nav.loc[end_date] / g5_nav.loc[start_date] - 1) * 100

        # 基准收益
        bench_return = (benchmark_nav.loc[end_date] / benchmark_nav.loc[start_date] - 1) * 100

        # 超额收益
        excess_return = g5_return - bench_return

        # 年份标注（最后一年如果不完整，加上标注）
        year_label = year
        if year == last_year and last_year_incomplete:
            year_label = f"{year}(截至{last_complete_month_end.month}月{last_complete_month_end.day}日)"

        yearly_data.append({
            '年份': year_label,
            'G5多头收益(%)': round(g5_return, 2),
            '超额收益(%)': round(excess_return, 2),
            '基准收益(%)': round(bench_return, 2),
        })
    
    # 添加全样本统计
    if len(nav_df) >= 2 and len(years) > 0:
        # 全样本起始日期：初始持仓日（第一年之前的最后一个日期，即净值=1.0的日期）
        first_year = years[0]
        prev_year_mask = nav_df.index.year < first_year

        if prev_year_mask.any():
            # 有上一年数据，使用上一年最后一个日期（初始持仓日）
            first_valid_date_for_full_sample = nav_df.index[prev_year_mask][-1]
        else:
            # 没有上一年数据，使用第一年的第一个日期
            first_year_mask = nav_df.index.year == first_year
            first_year_dates = nav_df.index[first_year_mask]
            if len(first_year_dates) > 0:
                first_valid_date_for_full_sample = first_year_dates[0]
            else:
                first_valid_date_for_full_sample = nav_df.index[0]

        # 确保起始日期在g5_nav和benchmark_nav中存在
        if first_valid_date_for_full_sample in g5_nav.index and first_valid_date_for_full_sample in benchmark_nav.index:
            total_g5_return = (g5_nav.iloc[-1] / g5_nav.loc[first_valid_date_for_full_sample] - 1) * 100
            total_bench_return = (benchmark_nav.iloc[-1] / benchmark_nav.loc[first_valid_date_for_full_sample] - 1) * 100
            total_excess_return = total_g5_return - total_bench_return

            yearly_data.append({
                '年份': '全样本',
                'G5多头收益(%)': round(total_g5_return, 2),
                '超额收益(%)': round(total_excess_return, 2),
                '基准收益(%)': round(total_bench_return, 2),
            })
        else:
            print(f"警告: 无法在全样本计算中找到起始日期 {first_valid_date_for_full_sample} 的对应数据。")
    
    return pd.DataFrame(yearly_data)


def analyze_single_factor_window(factor_name, data: DataContainer, window, rebalance_freq=DEFAULT_REBALANCE_FREQ, monthly_rebalance=DEFAULT_MONTHLY_REBALANCE, unified_start_date=None, split_ratio=None):
    """
    分析单个因子在单个窗口下的表现

    参数:
    factor_name: str, 因子名称
    data: DataContainer, 数据容器
    window: int, 因子计算的回溯窗口
    rebalance_freq: int, 调仓频率（仅在非月度调仓时使用）
    monthly_rebalance: bool, 是否按月调仓，默认为 True
    unified_start_date: pd.Timestamp, 统一的回测起始日期（如果指定，所有因子从该日期开始回测）
    split_ratio: float, 龙头/跟随分割参数（仅用于momentum_lead_lag_enhanced因子）

    返回:
    dict: 包含IC/IR、分层指标、持仓等信息
    """
    # 检查因子配置中是否有特定的 rebalance_type
    factor_rebalance_type = None
    requires_barra = False
    if factor_name in FACTOR_CONFIG:
        factor_rebalance_type = FACTOR_CONFIG[factor_name].get('rebalance_type', None)
        requires_barra = FACTOR_CONFIG[factor_name].get('requires_barra', False)

    # 如果因子配置了特定的调仓类型，使用该类型
    if factor_rebalance_type is not None:
        print(f"    使用因子配置的调仓类型: {factor_rebalance_type}")

    # 根据因子是否需要Barra数据，选择使用的价格数据范围
    if requires_barra and hasattr(data, 'barra_common_start') and data.barra_common_start is not None:
        # 需要Barra的因子，使用含Barra的日期范围
        prices_df_for_analysis = data.prices_df[
            (data.prices_df.index >= data.barra_common_start) &
            (data.prices_df.index <= data.barra_common_end)
        ]
        print(f"    使用含Barra的日期范围: {data.barra_common_start.date()} 至 {data.barra_common_end.date()}")
    else:
        # 不需要Barra的因子，使用完整的日期范围
        prices_df_for_analysis = data.prices_df

    # 计算因子值（传入split_ratio参数）
    factor_df = compute_factor(factor_name, data, window, rebalance_freq, split_ratio=split_ratio)

    # 如果需要Barra，截断因子数据到Barra日期范围
    if requires_barra and hasattr(data, 'barra_common_start') and data.barra_common_start is not None:
        factor_df = factor_df[
            (factor_df.index >= data.barra_common_start) &
            (factor_df.index <= data.barra_common_end)
        ]

    # 获取最后一个完整月末日期（用于排除不完整月份的收益计算）
    last_complete_month_end = None
    if (factor_rebalance_type == 'monthly' or (factor_rebalance_type is None and monthly_rebalance)) and hasattr(data, 'last_complete_month_end'):
        last_complete_month_end = data.last_complete_month_end
        # 如果需要Barra，使用Barra日期范围内的最后完整月末
        if requires_barra and hasattr(data, 'barra_common_end') and data.barra_common_end is not None:
            if last_complete_month_end is not None and last_complete_month_end > data.barra_common_end:
                # 重新计算Barra范围内的最后完整月末
                last_complete_month_end = get_last_complete_month_end(prices_df_for_analysis.index)

    # 计算未来收益率（根据调仓类型）
    if factor_rebalance_type is not None:
        forward_returns_df = calculate_forward_returns_by_type(
            prices_df_for_analysis, prices_df_for_analysis.index, factor_rebalance_type, rebalance_freq
        )
    elif monthly_rebalance:
        # 月度调仓：计算从当前月末到下一个月末的收益率
        forward_returns_df = calculate_monthly_forward_returns(prices_df_for_analysis, prices_df_for_analysis.index)
    else:
        # 固定频率调仓：使用固定周期的未来收益率
        forward_returns_df = prices_df_for_analysis.pct_change(rebalance_freq).shift(-rebalance_freq)
    
    # 计算IC/IR（包含 Pearson IC 和 Rank IC 累积序列）
    # 注意：IC计算需要完整的未来收益率，不完整月份的IC会自动被排除（因为forward_returns为NaN）
    ic_results = calculate_ic_ir(
        factor_df, forward_returns_df, rebalance_freq=rebalance_freq, monthly_rebalance=monthly_rebalance,
        unified_start_date=unified_start_date, rebalance_type=factor_rebalance_type
    )

    # 分层回测（传入最后一个完整月末日期，排除不完整月份的收益计算）
    # 使用对应的价格数据（需要Barra的因子使用截断后的价格数据）
    nav_df, layer_returns, layer_holdings = stratified_backtest(
        factor_df, prices_df_for_analysis, window, rebalance_freq=rebalance_freq, monthly_rebalance=monthly_rebalance,
        last_complete_month_end=last_complete_month_end, unified_start_date=unified_start_date,
        rebalance_type=factor_rebalance_type
    )

    # 获取最后一个完整月末的持仓（补充回测中可能遗漏的最后一个调仓日持仓）
    if monthly_rebalance:
        latest_holdings = get_latest_month_holdings(factor_df, prices_df_for_analysis, window, N_LAYERS, monthly_rebalance, last_complete_month_end)
        # 合并最新持仓到 layer_holdings（如果该日期不在回测持仓中）
        for layer_idx, holdings in latest_holdings.items():
            for date, assets in holdings.items():
                if date not in layer_holdings.get(layer_idx, {}):
                    if layer_idx not in layer_holdings:
                        layer_holdings[layer_idx] = {}
                    layer_holdings[layer_idx][date] = assets

    # 计算基准净值（等权行业指数）
    # 对齐到调仓日期，使用对应的价格数据
    benchmark_nav = pd.Series(index=nav_df.index, dtype=float)
    benchmark_nav.iloc[0] = 1.0
    for i in range(1, len(nav_df.index)):
        prev_date = nav_df.index[i-1]
        curr_date = nav_df.index[i]
        # 计算期间基准收益（所有行业等权平均）
        if prev_date in prices_df_for_analysis.index and curr_date in prices_df_for_analysis.index:
            period_ret = (prices_df_for_analysis.loc[curr_date] / prices_df_for_analysis.loc[prev_date] - 1).mean()
        else:
            period_ret = 0
        benchmark_nav.iloc[i] = benchmark_nav.iloc[i-1] * (1 + period_ret)
    
    # 计算超额指标
    excess_metrics = calculate_excess_metrics(
        nav_df,
        benchmark_nav,
        rebalance_freq,
        monthly_rebalance,
        rebalance_type=factor_rebalance_type,
    )

    # 计算每年收益统计（仅针对G5）
    # 自动从数据起始年份开始，传入最后完整月末用于标注不完整年
    yearly_returns = calculate_yearly_returns(nav_df, benchmark_nav, start_year=None,
                                              last_complete_month_end=last_complete_month_end)
    
    return {
        # Pearson IC 相关
        'ic_mean': ic_results['pearson_ic_mean'],
        'icir': ic_results['pearson_icir'],
        'ic_series': ic_results['pearson_ic_series'],
        'ic_cumsum': ic_results['pearson_ic_cumsum'],
        # Rank IC 相关
        'rank_ic_mean': ic_results['rank_ic_mean'],
        'rank_icir': ic_results['rank_icir'],
        'rank_ic_series': ic_results['rank_ic_series'],
        'rank_ic_cumsum': ic_results['rank_ic_cumsum'],
        # 通用指标
        'ic_win_rate': ic_results['ic_win_rate'],
        # 回测结果
        'nav_df': nav_df,
        'benchmark_nav': benchmark_nav,
        'excess_metrics': excess_metrics,
        'layer_holdings': layer_holdings,
        'yearly_returns': yearly_returns,
    }


def analyze_all_factors(data: DataContainer, windows=LOOKBACK_WINDOWS, rebalance_freq=DEFAULT_REBALANCE_FREQ, monthly_rebalance=DEFAULT_MONTHLY_REBALANCE, use_unified_start_date=True):
    """
    分析所有因子在所有窗口下的表现

    参数:
    data: DataContainer, 数据容器
    windows: list, 窗口列表（因子计算的回溯窗口，作为默认值）
    rebalance_freq: int, 调仓频率（仅在非月度调仗时使用）
    monthly_rebalance: bool, 是否按月调仓，默认为 True (每月最后一个交易日调仓)
    use_unified_start_date: bool, 是否使用统一的回测起始日期，默认为 True
        - True: 每个因子内部不同窗口使用该因子的最大预热期对应的统一起始日期
        - False: 每个窗口使用各自的预热期

    返回:
    dict: {factor_name: {window: analysis_result}}
    """
    import time

    all_results = {}
    factor_timing = {}  # 记录每个因子的耗时

    # 打印预热期分析
    if use_unified_start_date:
        print("\n" + "=" * 60)
        print("各因子预热期分析（每个因子内部不同窗口使用统一起始日期）")
        print("=" * 60)

    total_start_time = time.time()
    factor_list = list(FACTOR_CONFIG.keys())

    for i, factor_name in enumerate(factor_list):
        factor_start_time = time.time()

        # 调用单因子分析函数
        factor_results = analyze_single_factor(
            factor_name, data, windows, rebalance_freq, monthly_rebalance, use_unified_start_date
        )
        all_results.update(factor_results)

        # 计算该因子耗时
        factor_elapsed = time.time() - factor_start_time
        factor_minutes = int(factor_elapsed // 60)
        factor_seconds = int(factor_elapsed % 60)
        factor_timing[factor_name] = factor_minutes

        # 打印进度和耗时
        print(f"  ✓ {factor_name} 完成 ({i+1}/{len(factor_list)}) - 耗时: {factor_minutes}分{factor_seconds}秒")

    # 打印总耗时
    total_elapsed = time.time() - total_start_time
    total_hours = int(total_elapsed // 3600)
    total_minutes = int((total_elapsed % 3600) // 60)
    print("\n" + "=" * 60)
    print("因子分析耗时汇总")
    print("=" * 60)
    for factor_name, minutes in factor_timing.items():
        print(f"  {factor_name}: {minutes}分钟")
    print(f"\n总耗时: {total_hours}小时{total_minutes}分钟" if total_hours > 0 else f"\n总耗时: {total_minutes}分钟")

    # 将耗时信息存储到结果中，供后续导出Excel使用
    all_results['_factor_timing'] = factor_timing

    return all_results


def analyze_single_factor(factor_name, data: DataContainer, windows=LOOKBACK_WINDOWS, rebalance_freq=DEFAULT_REBALANCE_FREQ, monthly_rebalance=DEFAULT_MONTHLY_REBALANCE, use_unified_start_date=True):
    """
    分析单个因子在所有窗口下的表现

    参数:
    factor_name: str, 因子名称（必须在 FACTOR_CONFIG 中定义）
    data: DataContainer, 数据容器
    windows: list, 窗口列表（因子计算的回溯窗口）
    rebalance_freq: int, 调仓频率（仅在非月度调仓时使用）
    monthly_rebalance: bool, 是否按月调仓，默认为 True
    use_unified_start_date: bool, 是否使用统一的回测起始日期，默认为 True

    返回:
    dict: {factor_name: {window: analysis_result}} 或 {factor_name: {(window, split_ratio): analysis_result}}
    """
    if factor_name not in FACTOR_CONFIG:
        raise ValueError(f"因子 '{factor_name}' 不存在，可用因子: {list(FACTOR_CONFIG.keys())}")

    # 获取该因子适用的窗口列表
    factor_windows = get_factor_windows(factor_name, windows)

    # 获取该因子的split_ratios配置（如果有）
    factor_config = FACTOR_CONFIG[factor_name]
    split_ratios = factor_config.get('split_ratios', None)

    print(f"\n正在分析因子: {factor_name}")
    print(f"  适用窗口: {factor_windows}")
    if split_ratios:
        print(f"  分割参数: {split_ratios}")
    factor_results = {}

    # 获取 DataContainer 中的 first_holding_date（如果有）
    first_holding_date = getattr(data, 'first_holding_date', None)

    # 计算统一起始日期
    factor_unified_start_date = None
    if use_unified_start_date:
        try:
            warmup_start_date, max_warmup, max_warmup_window = calculate_factor_unified_start_date(
                data, factor_name, factor_windows, rebalance_freq, monthly_rebalance
            )
            print(f"  预热期: {max_warmup}天 (来自{max_warmup_window}日窗口)")
            print(f"  预热期结束日期: {warmup_start_date.strftime('%Y-%m-%d')}")

            # 如果 DataContainer 指定了最早持仓日期，取两者中较晚的
            if first_holding_date is not None:
                first_holding_ts = pd.Timestamp(first_holding_date)

                # 将 first_holding_date 转换为实际的月末交易日（小于等于该日期的最近月末）
                if monthly_rebalance:
                    all_dates = data.prices_df.index
                    monthly_dates = get_monthly_rebalance_dates(all_dates)
                    # 找到小于等于 first_holding_ts 的最近月末交易日
                    valid_monthly_dates = monthly_dates[monthly_dates <= first_holding_ts]
                    if len(valid_monthly_dates) > 0:
                        first_holding_ts = valid_monthly_dates[-1]

                if first_holding_ts > warmup_start_date:
                    factor_unified_start_date = first_holding_ts
                    print(f"  使用指定的最早持仓日: {factor_unified_start_date.strftime('%Y-%m-%d')}")
                else:
                    factor_unified_start_date = warmup_start_date
                    print(f"  使用预热期结束日期: {factor_unified_start_date.strftime('%Y-%m-%d')}（早于指定日期）")
            else:
                factor_unified_start_date = warmup_start_date
                print(f"  统一起始日期: {factor_unified_start_date.strftime('%Y-%m-%d')}")
        except Exception as e:
            print(f"  计算统一起始日期失败: {e}")
            factor_unified_start_date = None

    # 如果有split_ratios配置，遍历所有(window, split_ratio)组合
    if split_ratios:
        for window in factor_windows:
            for split_ratio in split_ratios:
                key = (window, split_ratio)
                print(f"  窗口: {window}日, 分割参数: {split_ratio}...")
                try:
                    result = analyze_single_factor_window(
                        factor_name, data, window, rebalance_freq, monthly_rebalance=monthly_rebalance,
                        unified_start_date=factor_unified_start_date, split_ratio=split_ratio
                    )
                    factor_results[key] = result
                except Exception as e:
                    print(f"    错误: {e}")
                    factor_results[key] = None
    else:
        # 没有split_ratios配置，按原来的方式遍历窗口
        # 如果有default_split_ratio配置，使用该值
        default_split_ratio = factor_config.get('default_split_ratio', None)
        for window in factor_windows:
            print(f"  窗口: {window}日...")
            try:
                result = analyze_single_factor_window(
                    factor_name, data, window, rebalance_freq, monthly_rebalance=monthly_rebalance,
                    unified_start_date=factor_unified_start_date, split_ratio=default_split_ratio
                )
                factor_results[window] = result
            except Exception as e:
                print(f"    错误: {e}")
                factor_results[window] = None

    return {factor_name: factor_results}


def find_best_windows(factor_results, top_n=1):
    """
    找到IC和ICIR最大的窗口（取原始值最大，不取绝对值）

    动量因子是正向因子：因子值越大，预期未来收益越高
    因此IC和ICIR应该为正，取原始值最大的窗口

    参数:
    factor_results: dict, {window: analysis_result} 单个因子的所有窗口结果
    top_n: int, 返回前N个最优窗口，默认1（只取最优的1个）

    返回:
    dict: {
        'best_ic_windows': [(window, ic_mean), ...],  # IC最大的窗口列表
        'best_icir_windows': [(window, icir), ...],   # ICIR最大的窗口列表
        'best_combined_windows': [window, ...],       # IC和ICIR综合最优的窗口（取并集）
    }
    """
    ic_values = {}
    icir_values = {}

    for window, result in factor_results.items():
        if result is None:
            continue
        ic_mean = result.get('ic_mean', np.nan)
        icir = result.get('icir', np.nan)

        # 直接使用原始值比较（动量因子是正向因子，IC/ICIR应为正）
        if not np.isnan(ic_mean):
            ic_values[window] = ic_mean
        if not np.isnan(icir):
            icir_values[window] = icir

    # 按原始值排序，找到最大的窗口
    sorted_ic = sorted(ic_values.items(), key=lambda x: x[1], reverse=True)
    sorted_icir = sorted(icir_values.items(), key=lambda x: x[1], reverse=True)

    # 取前top_n个
    best_ic_windows = sorted_ic[:top_n]
    best_icir_windows = sorted_icir[:top_n]

    # 综合最优窗口（IC和ICIR的并集）
    best_windows_set = set()
    for window, _ in best_ic_windows:
        best_windows_set.add(window)
    for window, _ in best_icir_windows:
        best_windows_set.add(window)

    return {
        'best_ic_windows': best_ic_windows,
        'best_icir_windows': best_icir_windows,
        'best_combined_windows': sorted(list(best_windows_set)),
    }


def create_best_windows_summary(all_results, top_n=1):
    """
    创建所有因子的最优窗口汇总表

    参数:
    all_results: dict, {factor_name: {window: analysis_result}}
    top_n: int, 每个指标取前N个最优窗口，默认1

    返回:
    pd.DataFrame: 最优窗口汇总表
    """
    summary_data = []

    for factor_name, factor_results in all_results.items():
        if not factor_results:
            continue

        best_info = find_best_windows(factor_results, top_n)

        # 格式化最优窗口信息
        ic_windows_str = ', '.join([f"{w}日(IC={v:.4f})" for w, v in best_info['best_ic_windows']])
        icir_windows_str = ', '.join([f"{w}日(ICIR={v:.4f})" for w, v in best_info['best_icir_windows']])
        combined_windows_str = ', '.join([f"{w}日" for w in best_info['best_combined_windows']])

        summary_data.append({
            '因子名称': factor_name,
            '最优IC窗口': ic_windows_str,
            '最优ICIR窗口': icir_windows_str,
            '综合最优窗口': combined_windows_str,
        })

    return pd.DataFrame(summary_data)


def create_factor_summary_df(factor_name, factor_results, windows=LOOKBACK_WINDOWS):
    """
    创建单个因子的汇总DataFrame

    格式：
    - 列名为窗口周期（20, 60, 120, 240）
    - 行名为指标名称
    - G5放在最前面，然后是G4, G3, G2, G1
    - IC和ICIR保留4位小数，其他指标保留2位小数

    返回:
    pd.DataFrame: 汇总表格（行为指标，列为窗口）
    """
    # 构建数据字典：{窗口: {指标: 值}}
    data_dict = {}

    for window in windows:
        result = factor_results.get(window)
        if result is None:
            continue

        window_data = {}
        # Pearson IC 和 ICIR 保留4位小数
        window_data['IC均值'] = round(result['ic_mean'], 4) if not np.isnan(result['ic_mean']) else np.nan
        window_data['ICIR'] = round(result['icir'], 4) if not np.isnan(result['icir']) else np.nan
        # Rank IC 和 RankICIR 保留4位小数
        window_data['RankIC均值'] = round(result['rank_ic_mean'], 4) if not np.isnan(result['rank_ic_mean']) else np.nan
        window_data['RankICIR'] = round(result['rank_icir'], 4) if not np.isnan(result['rank_icir']) else np.nan
        # IC胜率
        window_data['IC胜率'] = round(result['ic_win_rate'], 4) if not np.isnan(result['ic_win_rate']) else np.nan

        # 按G5, G4, G3, G2, G1顺序添加各层指标
        layer_order = ['G5', 'G4', 'G3', 'G2', 'G1']
        for layer_name in layer_order:
            if layer_name in result['excess_metrics']:
                metrics = result['excess_metrics'][layer_name]
                for metric_name, value in metrics.items():
                    row_name = f'{layer_name}_{metric_name}'
                    window_data[row_name] = round(value, 2)

        data_dict[window] = window_data

    # 转换为DataFrame并转置（行为指标，列为窗口）
    df = pd.DataFrame(data_dict)

    # 确保列按窗口顺序排列
    df = df.reindex(columns=windows)

    return df


def clean_industry_name(name):
    """
    清理行业名称，删除"（申万）"后缀
    
    参数:
    name: str, 行业名称
    
    返回:
    str: 清理后的行业名称
    """
    if isinstance(name, str):
        return name.replace('（申万）', '').replace('(申万)', '')
    return name


def format_date_index(df):
    """
    格式化DataFrame的日期索引，只保留日期部分（不含时间）
    
    参数:
    df: pd.DataFrame, 带有日期索引的DataFrame
    
    返回:
    pd.DataFrame: 格式化后的DataFrame
    """
    if isinstance(df.index, pd.DatetimeIndex):
        df.index = df.index.strftime('%Y-%m-%d')
    return df


def create_g5_holdings_df(factor_results, windows=LOOKBACK_WINDOWS):
    """
    创建G5持仓记录DataFrame（按列输出每个窗口的G5持仓）
    
    格式：
    - 列名为窗口周期（20, 60, 120, 240）
    - 行为日期（降序，最新日期在最前面）
    - 值为持仓行业（逗号分隔，不含"（申万）"后缀）
    
    参数:
    factor_results: dict, 因子分析结果 {window: result}
    windows: list, 窗口列表
    
    返回:
    pd.DataFrame: G5持仓记录（行为日期，列为窗口）
    """
    # 收集所有窗口的G5持仓
    holdings_dict = {}
    all_dates = set()
    
    for window in windows:
        result = factor_results.get(window)
        if result is None or result.get('layer_holdings') is None:
            continue
        
        # G5 是 layer_idx=4
        g5_holdings = result['layer_holdings'].get(4, {})
        
        # 转换为 {日期: 持仓行业字符串}，同时清理行业名称
        window_holdings = {}
        for date, industries in g5_holdings.items():
            # 清理行业名称，删除"（申万）"后缀
            cleaned_industries = [clean_industry_name(ind) for ind in industries] if industries else []
            window_holdings[date] = ', '.join(cleaned_industries) if cleaned_industries else ''
            all_dates.add(date)
        
        holdings_dict[window] = window_holdings
    
    if not all_dates:
        return pd.DataFrame()
    
    # 按日期降序排列（最新日期在最前面）
    sorted_dates = sorted(all_dates, reverse=True)
    
    # 构建DataFrame
    data = {}
    for window in windows:
        if window in holdings_dict:
            data[window] = [holdings_dict[window].get(date, '') for date in sorted_dates]
        else:
            data[window] = [''] * len(sorted_dates)
    
    # 格式化日期索引（只保留日期部分）
    formatted_dates = [d.strftime('%Y-%m-%d') if hasattr(d, 'strftime') else str(d) for d in sorted_dates]
    
    df = pd.DataFrame(data, index=formatted_dates)
    df.index.name = '日期'
    
    return df


def create_ic_cumsum_df(factor_results, windows=LOOKBACK_WINDOWS):
    """
    创建IC累积序列DataFrame（包含 Pearson IC 和 Rank IC）

    参数:
    factor_results: dict, 因子分析结果 {window: result}
    windows: list, 窗口列表

    返回:
    tuple: (pearson_ic_cumsum_df, rank_ic_cumsum_df) 两个IC累积序列DataFrame
    """
    pearson_ic_cumsum_dict = {}
    rank_ic_cumsum_dict = {}
    all_dates = set()

    for window in windows:
        result = factor_results.get(window)
        if result is None:
            continue

        # Pearson IC 累积序列
        if result.get('ic_cumsum') is not None:
            pearson_ic_cumsum_dict[window] = result['ic_cumsum']
            all_dates.update(result['ic_cumsum'].index)

        # Rank IC 累积序列
        if result.get('rank_ic_cumsum') is not None:
            rank_ic_cumsum_dict[window] = result['rank_ic_cumsum']
            all_dates.update(result['rank_ic_cumsum'].index)

    if not all_dates:
        return pd.DataFrame(), pd.DataFrame()

    # 按日期升序排列
    sorted_dates = sorted(all_dates)

    # 构建 Pearson IC DataFrame
    pearson_data = {}
    for window in windows:
        if window in pearson_ic_cumsum_dict:
            ic_series = pearson_ic_cumsum_dict[window]
            pearson_data[f'{window}日'] = [ic_series.loc[date] if date in ic_series.index else np.nan for date in sorted_dates]
        else:
            pearson_data[f'{window}日'] = [np.nan] * len(sorted_dates)

    # 构建 Rank IC DataFrame
    rank_data = {}
    for window in windows:
        if window in rank_ic_cumsum_dict:
            ic_series = rank_ic_cumsum_dict[window]
            rank_data[f'{window}日'] = [ic_series.loc[date] if date in ic_series.index else np.nan for date in sorted_dates]
        else:
            rank_data[f'{window}日'] = [np.nan] * len(sorted_dates)

    # 格式化日期索引（只保留日期部分）
    formatted_dates = [d.strftime('%Y-%m-%d') if hasattr(d, 'strftime') else str(d) for d in sorted_dates]

    pearson_df = pd.DataFrame(pearson_data, index=formatted_dates)
    pearson_df.index.name = '日期'
    pearson_df = pearson_df.round(4)
    pearson_df = pearson_df.dropna()

    rank_df = pd.DataFrame(rank_data, index=formatted_dates)
    rank_df.index.name = '日期'
    rank_df = rank_df.round(4)
    rank_df = rank_df.dropna()

    return pearson_df, rank_df


def create_layer_nav_df(factor_results, windows=LOOKBACK_WINDOWS):
    """
    创建分层累积净值DataFrame（每个窗口的G1-G5净值）
    
    参数:
    factor_results: dict, 因子分析结果 {window: result}
    windows: list, 窗口列表
    
    返回:
    dict: {window: nav_df} 每个窗口的分层净值DataFrame
    """
    layer_nav_dict = {}
    
    for window in windows:
        result = factor_results.get(window)
        if result is None or result.get('nav_df') is None:
            continue
        
        nav_df = result['nav_df'].copy()
        # 添加基准净值列
        if result.get('benchmark_nav') is not None:
            nav_df['基准'] = result['benchmark_nav']
        
        # 保留四位小数
        nav_df = nav_df.round(4)
        
        # 格式化日期索引（只保留日期部分）
        nav_df = format_date_index(nav_df)
        
        layer_nav_dict[window] = nav_df
    
    return layer_nav_dict


def create_g5_yearly_returns_df(factor_results, windows=LOOKBACK_WINDOWS):
    """
    创建G5每年收益统计DataFrame（仅针对指定窗口，如G5对应的窗口）
    
    参数:
    factor_results: dict, 因子分析结果 {window: result}
    windows: list, 窗口列表
    
    返回:
    dict: {window: yearly_returns_df}
    """
    yearly_dict = {}
    
    for window in windows:
        result = factor_results.get(window)
        if result is None or result.get('yearly_returns') is None:
            continue
        
        yearly_df = result['yearly_returns']
        if not yearly_df.empty:
            yearly_dict[window] = yearly_df
    
    return yearly_dict


def get_data_date_range(factor_results, windows=LOOKBACK_WINDOWS):
    """
    获取数据日期范围
    
    参数:
    factor_results: dict, 因子分析结果 {window: result}
    windows: list, 窗口列表
    
    返回:
    tuple: (start_date_str, end_date_str) 日期范围字符串
    """
    all_dates = []
    
    for window in windows:
        result = factor_results.get(window)
        if result is None or result.get('nav_df') is None:
            continue
        
        nav_df = result['nav_df']
        if len(nav_df) > 0:
            all_dates.extend(nav_df.index.tolist())
    
    if not all_dates:
        return None, None
    
    start_date = min(all_dates)
    end_date = max(all_dates)
    
    # 格式化日期
    start_str = start_date.strftime('%Y-%m-%d') if hasattr(start_date, 'strftime') else str(start_date)
    end_str = end_date.strftime('%Y-%m-%d') if hasattr(end_date, 'strftime') else str(end_date)
    
    return start_str, end_str


def create_best_factors_summary_df(all_results):
    """
    创建所有因子最优窗口的汇总DataFrame（横向排列）

    参数:
    all_results: dict, 所有因子的分析结果

    返回:
    pd.DataFrame: 汇总表格（行为指标，列为因子名称+窗口）
    """
    all_data = {}

    for factor_name, factor_results in all_results.items():
        # 获取最优窗口（只使用IC最优）
        best_info = find_best_windows(factor_results, top_n=1)
        best_ic_windows = [w for w, _ in best_info['best_ic_windows']]

        # 获取因子描述作为列名前缀
        if factor_name in FACTOR_CONFIG:
            factor_desc = FACTOR_CONFIG[factor_name].get('description', factor_name)
        else:
            factor_desc = factor_name

        # 只使用IC最优窗口
        for window in best_ic_windows:
            result = factor_results.get(window)
            if result is None:
                continue

            # 列名：因子描述 + 窗口
            col_name = f"{factor_desc}_{window}日"

            col_data = {}
            # Pearson IC 和 ICIR
            col_data['IC均值'] = round(result['ic_mean'], 4) if not np.isnan(result['ic_mean']) else np.nan
            col_data['ICIR'] = round(result['icir'], 4) if not np.isnan(result['icir']) else np.nan
            # Rank IC 和 RankICIR
            col_data['RankIC均值'] = round(result['rank_ic_mean'], 4) if not np.isnan(result['rank_ic_mean']) else np.nan
            col_data['RankICIR'] = round(result['rank_icir'], 4) if not np.isnan(result['rank_icir']) else np.nan
            # IC胜率
            col_data['IC胜率'] = round(result['ic_win_rate'], 4) if not np.isnan(result['ic_win_rate']) else np.nan

            # 只输出G5的指标
            if 'G5' in result['excess_metrics']:
                metrics = result['excess_metrics']['G5']
                for metric_name, value in metrics.items():
                    col_data[f'G5_{metric_name}'] = round(value, 2)

            all_data[col_name] = col_data

    if not all_data:
        return pd.DataFrame()

    df = pd.DataFrame(all_data)
    return df


# ============================================================
# 多因子相关性分析
# ============================================================

def collect_best_factor_data(all_results):
    """
    从分析结果中收集各因子最优窗口的IC序列和G5净值数据

    参数:
        all_results: dict, {factor_name: {window: analysis_result}}

    返回:
        ic_df: pd.DataFrame, 各因子IC序列 (index=日期, columns=因子名)
        nav_df: pd.DataFrame, 各因子G5净值 (index=日期, columns=因子名)
        factor_info: dict, 各因子信息 {factor_name: {'window': w, 'ic_mean': ic, 'icir': icir}}
    """
    ic_dict = {}
    nav_dict = {}
    factor_info = {}

    for factor_name, factor_results in all_results.items():
        # 跳过内部键（如 _factor_timing）
        if factor_name.startswith('_'):
            continue

        # 找到该因子的最优窗口（按IC均值）
        best_info = find_best_windows(factor_results, top_n=1)
        if not best_info['best_ic_windows']:
            continue

        best_window, best_ic = best_info['best_ic_windows'][0]
        result = factor_results.get(best_window)

        if result is None:
            continue

        # 获取因子描述作为简短名称
        config = FACTOR_CONFIG.get(factor_name, {})
        short_name = config.get('description', factor_name)
        # 提取简短名称（如 "基础动量1-传统动量因子" -> "基础动量1"）
        if '-' in short_name:
            short_name = short_name.split('-')[0]

        # 提取IC序列
        ic_series = result.get('ic_series')
        if ic_series is not None and len(ic_series) > 0:
            ic_dict[short_name] = ic_series

        # 提取G5净值
        nav = result.get('nav_df')
        if nav is not None and 'G5' in nav.columns:
            nav_dict[short_name] = nav['G5']

        # 记录因子信息
        factor_info[short_name] = {
            'factor_name': factor_name,
            'description': config.get('description', factor_name),  # 完整描述用于显示
            'window': best_window,
            'ic_mean': result.get('ic_mean', np.nan),
            'icir': result.get('icir', np.nan),
        }

    # 合并为DataFrame并对齐日期
    ic_df = pd.DataFrame(ic_dict)
    nav_df = pd.DataFrame(nav_dict)

    return ic_df, nav_df, factor_info


def calculate_return_correlation(nav_df):
    """
    计算收益率相关性矩阵

    注意：使用收益率而非净值！（避免伪回归）

    返回:
        corr_matrix: pd.DataFrame, 相关性矩阵
        returns_df: pd.DataFrame, 收益率序列
    """
    # 计算收益率（月度调仓，实际是月收益率）
    returns_df = nav_df.pct_change().dropna()

    if returns_df.empty or len(returns_df) < 3:
        return pd.DataFrame(), returns_df

    # 计算Pearson相关系数矩阵
    corr_matrix = returns_df.corr()

    return corr_matrix, returns_df


def calculate_ic_correlation(ic_df):
    """
    计算IC序列相关性矩阵

    返回:
        corr_matrix: pd.DataFrame, 相关性矩阵
        ic_aligned: pd.DataFrame, 对齐后的IC序列
    """
    # 对齐日期，去除NaN
    ic_aligned = ic_df.dropna()

    if ic_aligned.empty or len(ic_aligned) < 3:
        return pd.DataFrame(), ic_aligned

    # 计算Pearson相关系数矩阵
    corr_matrix = ic_aligned.corr()

    return corr_matrix, ic_aligned


def interpret_correlation_pair(r, corr_type='return'):
    """
    解读单对因子的相关性

    参数:
        r: float, 相关系数
        corr_type: str, 'return' 或 'ic'

    返回:
        tuple: (级别, 建议)
    """
    if r > 0.6:
        level = "高相关"
        if corr_type == 'return':
            advice = "策略没分散，风险集中 → 建议合成为一个因子"
        else:
            advice = "同类Alpha来源 → 建议合成，不要分开用"
    elif r > 0.2:
        level = "中等相关"
        advice = "有一定相关性，需关注"
    elif r > -0.2:
        level = "低相关"
        if corr_type == 'return':
            advice = "实现了分散化，曲线更平滑"
        else:
            advice = "各有各的赚法，独立保留"
    elif r > -0.4:
        level = "负相关"
        advice = "互补性较强"
    else:
        level = "强负相关"
        if corr_type == 'return':
            advice = "完美对冲组合（前提是都有正收益）"
        else:
            advice = "互补性极强，重仓组合可提升夏普"

    return level, advice


def create_correlation_interpretation_df(corr_matrix, corr_type='return', factor_info=None):
    """
    创建相关性解读表格（按建议分类为三列）

    参数:
        corr_matrix: 相关性矩阵
        corr_type: 'return' 或 'ic'
        factor_info: dict, 因子信息（用于获取全名）

    返回:
        pd.DataFrame: 三列格式（高相关/中等相关/低相关），每个单元格为"【相关系数】因子1 vs 因子2"
    """
    if corr_matrix.empty:
        return pd.DataFrame()

    columns = corr_matrix.columns.tolist()

    # 按相关级别分类
    high_corr = []      # > 0.6
    medium_corr = []    # 0.2 ~ 0.6
    low_corr = []       # < 0.2 (包括负相关)

    for i in range(len(columns)):
        for j in range(i+1, len(columns)):
            col1, col2 = columns[i], columns[j]
            r = corr_matrix.loc[col1, col2]

            # 获取因子全名（使用description）
            if factor_info:
                full_name1 = factor_info.get(col1, {}).get('description', col1)
                full_name2 = factor_info.get(col2, {}).get('description', col2)
            else:
                full_name1, full_name2 = col1, col2

            # 格式化单元格内容
            cell_content = f"【{r:.3f}】{full_name1} vs {full_name2}"

            # 按相关系数分类
            if r > 0.6:
                high_corr.append((abs(r), cell_content))
            elif r > 0.2:
                medium_corr.append((abs(r), cell_content))
            else:
                low_corr.append((abs(r), cell_content))

    # 按相关系数绝对值降序排列
    high_corr.sort(key=lambda x: x[0], reverse=True)
    medium_corr.sort(key=lambda x: x[0], reverse=True)
    low_corr.sort(key=lambda x: x[0], reverse=True)

    # 提取内容
    high_list = [x[1] for x in high_corr]
    medium_list = [x[1] for x in medium_corr]
    low_list = [x[1] for x in low_corr]

    # 对齐长度
    max_len = max(len(high_list), len(medium_list), len(low_list), 1)
    high_list.extend([''] * (max_len - len(high_list)))
    medium_list.extend([''] * (max_len - len(medium_list)))
    low_list.extend([''] * (max_len - len(low_list)))

    # 创建DataFrame
    df = pd.DataFrame({
        '高相关(>0.6) → 建议合成': high_list,
        '中等相关(0.2~0.6) → 需关注': medium_list,
        '低相关(<0.2) → 独立保留': low_list,
    })

    return df


def create_factor_correlation_summary_df(corr_matrix, factor_info=None):
    """
    创建从因子角度出发的相关性汇总表

    每个因子作为一行，列出与其他因子的关系分类

    参数:
        corr_matrix: 相关性矩阵
        factor_info: dict, 因子信息（用于获取全名）

    返回:
        pd.DataFrame: 因子为行，三列（建议合并/保持关注/独立保留）
    """
    if corr_matrix.empty:
        return pd.DataFrame()

    columns = corr_matrix.columns.tolist()

    # 为每个因子收集相关性信息
    factor_summary = {}

    for factor in columns:
        high_corr = []      # > 0.6 建议合并
        medium_corr = []    # 0.2 ~ 0.6 保持关注
        low_corr = []       # < 0.2 独立保留

        for other_factor in columns:
            if factor == other_factor:
                continue

            r = corr_matrix.loc[factor, other_factor]

            # 获取另一个因子的全名（使用description）
            if factor_info:
                other_full_name = factor_info.get(other_factor, {}).get('description', other_factor)
            else:
                other_full_name = other_factor

            # 格式化内容：【相关系数】因子全名
            cell_content = f"【{r:.3f}】{other_full_name}"

            # 按相关系数分类
            if r > 0.6:
                high_corr.append((r, cell_content))
            elif r > 0.2:
                medium_corr.append((r, cell_content))
            else:
                low_corr.append((abs(r), cell_content))

        # 按相关系数排序
        high_corr.sort(key=lambda x: x[0], reverse=True)
        medium_corr.sort(key=lambda x: x[0], reverse=True)
        low_corr.sort(key=lambda x: x[0], reverse=True)

        # 合并为字符串（换行分隔）
        high_str = '\n'.join([x[1] for x in high_corr]) if high_corr else ''
        medium_str = '\n'.join([x[1] for x in medium_corr]) if medium_corr else ''
        low_str = '\n'.join([x[1] for x in low_corr]) if low_corr else ''

        # 获取当前因子的全名（使用description）
        if factor_info:
            factor_full_name = factor_info.get(factor, {}).get('description', factor)
        else:
            factor_full_name = factor

        factor_summary[factor_full_name] = {
            '建议合并(>0.6)': high_str,
            '保持关注(0.2~0.6)': medium_str,
            '独立保留(<0.2)': low_str,
        }

    # 创建DataFrame
    df = pd.DataFrame(factor_summary).T
    df.index.name = '因子'

    return df


def generate_factor_synthesis_recommendation(ic_corr, factor_info, corr_threshold=0.6):
    """
    生成因子合成与保留建议

    使用层次聚类识别高相关因子组，并基于ICIR给出合成建议

    参数:
        ic_corr: pd.DataFrame, IC相关性矩阵
        factor_info: dict, 因子信息（包含ICIR等）
        corr_threshold: float, 高相关阈值，默认0.6

    返回:
        dict: 包含聚类结果和建议
    """
    from scipy.cluster.hierarchy import linkage, fcluster
    from scipy.spatial.distance import squareform

    if ic_corr.empty or len(ic_corr) < 2:
        return None

    factors = ic_corr.columns.tolist()
    n_factors = len(factors)

    # 将相关性转换为距离矩阵 (1 - corr)
    # 相关性越高，距离越近
    dist_matrix = 1 - ic_corr.values
    np.fill_diagonal(dist_matrix, 0)  # 对角线设为0

    # 确保距离矩阵对称且非负
    dist_matrix = (dist_matrix + dist_matrix.T) / 2
    dist_matrix = np.maximum(dist_matrix, 0)

    # 转换为压缩距离矩阵
    condensed_dist = squareform(dist_matrix)

    # 层次聚类
    linkage_matrix = linkage(condensed_dist, method='average')

    # 根据阈值切割聚类树 (距离阈值 = 1 - corr_threshold)
    dist_threshold = 1 - corr_threshold
    clusters = fcluster(linkage_matrix, t=dist_threshold, criterion='distance')

    # 整理聚类结果
    cluster_groups = {}
    for factor, cluster_id in zip(factors, clusters):
        if cluster_id not in cluster_groups:
            cluster_groups[cluster_id] = []
        cluster_groups[cluster_id].append(factor)

    # 为每个聚类组生成建议
    recommendations = []
    independent_factors = []  # 独立因子（单独成组）
    synthesis_groups = []     # 需要合成的组

    for cluster_id, group_factors in cluster_groups.items():
        if len(group_factors) == 1:
            # 单独成组，独立保留
            factor = group_factors[0]
            info = factor_info.get(factor, {})
            independent_factors.append({
                'factor': factor,
                'description': info.get('description', factor),
                'icir': info.get('icir', 0),
                'reason': '与其他因子相关性低，Alpha来源独立'
            })
        else:
            # 多因子成组，需要合成
            group_info = []
            for factor in group_factors:
                info = factor_info.get(factor, {})
                group_info.append({
                    'factor': factor,
                    'description': info.get('description', factor),
                    'icir': info.get('icir', 0),
                })

            # 按ICIR排序，选出最优因子
            group_info.sort(key=lambda x: x['icir'], reverse=True)
            best_factor = group_info[0]
            other_factors = group_info[1:]

            # 计算每个因子与组内其他因子的平均相关性
            factor_avg_corrs = {}
            for f1 in group_factors:
                corrs_with_others = []
                for f2 in group_factors:
                    if f1 != f2:
                        corrs_with_others.append(ic_corr.loc[f1, f2])
                factor_avg_corrs[f1] = np.mean(corrs_with_others) if corrs_with_others else 0

            # 计算组内所有因子对的平均相关性
            all_pair_corrs = []
            for i, f1 in enumerate(group_factors):
                for f2 in group_factors[i+1:]:
                    all_pair_corrs.append(ic_corr.loc[f1, f2])
            group_avg_corr = np.mean(all_pair_corrs) if all_pair_corrs else 0

            synthesis_groups.append({
                'group_id': cluster_id,
                'factors': group_factors,
                'factor_details': group_info,
                'best_factor': best_factor,
                'other_factors': other_factors,
                'factor_avg_corrs': factor_avg_corrs,  # 每个因子与组内其他因子的平均相关性
                'avg_correlation': group_avg_corr,  # 组内平均相关性
                'recommendation': f"建议保留 {best_factor['factor']}（ICIR最高: {best_factor['icir']:.3f}），"
                                  f"或将组内{len(group_factors)}个因子等权/ICIR加权合成"
            })

    # 生成汇总建议DataFrame
    summary_rows = []

    # 1. 需要合成的组
    for group in synthesis_groups:
        for i, detail in enumerate(group['factor_details']):
            factor_name = detail['factor']
            row = {
                '分组': f"合成组{group['group_id']}",
                '因子简称': factor_name,
                '因子全名': detail['description'],
                'ICIR': detail['icir'],
                '组内排名': i + 1,
                '与组内其他因子平均相关性': group['factor_avg_corrs'].get(factor_name, 0),
                '建议': '★ 推荐保留' if i == 0 else '可合成或剔除',
            }
            summary_rows.append(row)

    # 2. 独立因子
    for factor_data in independent_factors:
        row = {
            '分组': '独立因子',
            '因子简称': factor_data['factor'],
            '因子全名': factor_data['description'],
            'ICIR': factor_data['icir'],
            '组内排名': 1,
            '与组内其他因子平均相关性': 0,
            '建议': '★ 独立保留',
        }
        summary_rows.append(row)

    summary_df = pd.DataFrame(summary_rows)

    # 生成最终建议文本
    final_recommendations = []
    final_recommendations.append("=" * 60)
    final_recommendations.append("【因子合成与保留建议】")
    final_recommendations.append("=" * 60)
    final_recommendations.append(f"\n分析基于IC序列相关性，阈值: {corr_threshold}")
    final_recommendations.append(f"共 {n_factors} 个因子，分为 {len(cluster_groups)} 组\n")

    # 合成组建议
    if synthesis_groups:
        final_recommendations.append("-" * 40)
        final_recommendations.append("【需要合成/精简的因子组】")
        final_recommendations.append("-" * 40)
        for group in synthesis_groups:
            final_recommendations.append(f"\n◆ 合成组{group['group_id']}（{len(group['factors'])}个因子，平均相关性: {group['avg_correlation']:.3f}）")
            for i, detail in enumerate(group['factor_details']):
                marker = "★" if i == 0 else "  "
                final_recommendations.append(f"  {marker} {detail['factor']}: ICIR={detail['icir']:.3f}")
            final_recommendations.append(f"  → {group['recommendation']}")

    # 独立因子
    if independent_factors:
        final_recommendations.append("\n" + "-" * 40)
        final_recommendations.append("【独立保留的因子】（与其他因子相关性低）")
        final_recommendations.append("-" * 40)
        for factor_data in sorted(independent_factors, key=lambda x: x['icir'], reverse=True):
            final_recommendations.append(f"  ★ {factor_data['factor']}: ICIR={factor_data['icir']:.3f}")

    # 最终精简建议
    final_recommendations.append("\n" + "=" * 60)
    final_recommendations.append("【最终精简建议】")
    final_recommendations.append("=" * 60)

    recommended_factors = []
    for group in synthesis_groups:
        recommended_factors.append(group['best_factor']['factor'])
    for factor_data in independent_factors:
        recommended_factors.append(factor_data['factor'])

    final_recommendations.append(f"\n从 {n_factors} 个因子精简为 {len(recommended_factors)} 个:")
    for f in recommended_factors:
        info = factor_info.get(f, {})
        final_recommendations.append(f"  • {f} ({info.get('description', f)})")

    return {
        'cluster_groups': cluster_groups,
        'synthesis_groups': synthesis_groups,
        'independent_factors': independent_factors,
        'summary_df': summary_df,
        'recommended_factors': recommended_factors,
        'recommendation_text': '\n'.join(final_recommendations),
    }


def analyze_multi_factor_correlation(all_results):
    """
    执行多因子相关性分析

    参数:
        all_results: dict, 所有因子的分析结果

    返回:
        dict: 包含所有相关性分析结果
    """
    print("\n" + "=" * 60)
    print("多因子相关性分析")
    print("=" * 60)

    # 收集各因子最优窗口的数据
    ic_df, nav_df, factor_info = collect_best_factor_data(all_results)

    if ic_df.empty or nav_df.empty:
        print("警告: 没有足够的因子数据进行相关性分析")
        return None

    print(f"\n成功收集 {len(ic_df.columns)} 个因子的数据")
    print(f"IC序列长度: {len(ic_df)}, 净值序列长度: {len(nav_df)}")

    # 计算收益率相关性
    print("\n计算收益率相关性...")
    return_corr, returns_df = calculate_return_correlation(nav_df)
    return_interp = create_correlation_interpretation_df(return_corr, 'return', factor_info)

    # 计算IC序列相关性
    print("计算IC序列相关性...")
    ic_corr, ic_aligned = calculate_ic_correlation(ic_df)
    ic_interp = create_correlation_interpretation_df(ic_corr, 'ic', factor_info)

    # 打印摘要
    if not return_corr.empty:
        print("\n【收益率相关性矩阵】（判断策略分散化效果）")
        print(return_corr.round(3).to_string())

    if not ic_corr.empty:
        print("\n【IC序列相关性矩阵】（判断Alpha来源独立性）")
        print(ic_corr.round(3).to_string())

    # 生成因子角度的汇总表（基于IC相关性）
    factor_summary_df = create_factor_correlation_summary_df(ic_corr, factor_info)

    # 生成因子合成与保留建议
    print("\n生成因子合成与保留建议...")
    synthesis_recommendation = generate_factor_synthesis_recommendation(ic_corr, factor_info)

    if synthesis_recommendation:
        print(synthesis_recommendation['recommendation_text'])

    return {
        'factor_info': factor_info,
        'ic_df': ic_df,
        'nav_df': nav_df,
        'returns_df': returns_df,
        'return_corr': return_corr,
        'return_interpretation': return_interp,
        'ic_corr': ic_corr,
        'ic_interpretation': ic_interp,
        'factor_summary': factor_summary_df,
        'synthesis_recommendation': synthesis_recommendation,
    }


def _export_correlation_sheet(writer, correlation_results):
    """
    导出多因子相关性分析到Excel sheet

    参数:
        writer: pd.ExcelWriter, Excel写入器
        correlation_results: dict, 相关性分析结果
    """
    sheet_name = '多因子相关性分析'
    start_row = 0

    # 1. 写入标题
    title_df = pd.DataFrame({'【多因子相关性分析】': [
        '分析维度：',
        '1. 收益率相关性 - 判断策略分散化效果（使用G5多头组月收益率）',
        '2. IC序列相关性 - 判断Alpha来源独立性',
        '',
        '解读标准：',
        '- 高相关(>0.6): 策略没分散/同类Alpha → 建议合成',
        '- 中等相关(0.2~0.6): 有一定相关性 → 需关注',
        '- 低相关(<0.2): 分散化良好/独立Alpha → 独立保留'
    ]})
    title_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
    start_row += len(title_df) + 2

    # 2. 写入因子信息汇总（只保留基本信息，不含解读行）
    factor_info = correlation_results.get('factor_info', {})
    if factor_info:
        info_rows = []
        for name, info in factor_info.items():
            info_rows.append({
                '因子简称': name,
                '因子全名': info.get('description', ''),
                '最优窗口': f"{info.get('window', '')}日",
                'IC均值': round(info.get('ic_mean', 0), 4),
                'ICIR': round(info.get('icir', 0), 4),
            })
        info_df = pd.DataFrame(info_rows)

        header_df = pd.DataFrame({'【因子信息汇总】': ['']})
        header_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
        start_row += 1
        info_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
        start_row += len(info_df) + 3

    # 3. 写入收益率相关性矩阵
    return_corr = correlation_results.get('return_corr')
    if return_corr is not None and not return_corr.empty:
        header_df = pd.DataFrame({'【收益率相关性矩阵】（判断策略分散化效果）': ['']})
        header_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
        start_row += 1
        return_corr.round(4).to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=True)
        start_row += len(return_corr) + 3

    # 4. 写入收益率相关性解读（三列格式）
    return_interp = correlation_results.get('return_interpretation')
    if return_interp is not None and not return_interp.empty:
        header_df = pd.DataFrame({'【收益率相关性解读】': ['']})
        header_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
        start_row += 1
        return_interp.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
        start_row += len(return_interp) + 3

    # 5. 写入IC序列相关性矩阵
    ic_corr = correlation_results.get('ic_corr')
    if ic_corr is not None and not ic_corr.empty:
        header_df = pd.DataFrame({'【IC序列相关性矩阵】（判断Alpha来源独立性）': ['']})
        header_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
        start_row += 1
        ic_corr.round(4).to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=True)
        start_row += len(ic_corr) + 3

    # 6. 写入IC序列相关性解读（三列格式）
    ic_interp = correlation_results.get('ic_interpretation')
    if ic_interp is not None and not ic_interp.empty:
        header_df = pd.DataFrame({'【IC序列相关性解读】': ['']})
        header_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
        start_row += 1
        ic_interp.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
        start_row += len(ic_interp) + 3

    # 7. 写入因子角度汇总表（基于IC相关性）
    factor_summary = correlation_results.get('factor_summary')
    if factor_summary is not None and not factor_summary.empty:
        header_df = pd.DataFrame({'【因子角度汇总表】（基于IC序列相关性，每行为一个因子）': ['']})
        header_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
        start_row += 1
        factor_summary.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=True)
        start_row += len(factor_summary) + 3

    # 8. 写入因子合成与保留建议
    synthesis_rec = correlation_results.get('synthesis_recommendation')
    if synthesis_rec is not None:
        # 8.1 写入汇总表
        summary_df = synthesis_rec.get('summary_df')
        if summary_df is not None and not summary_df.empty:
            header_df = pd.DataFrame({'【因子合成与保留建议】（基于层次聚类，相关性阈值0.6）': ['']})
            header_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
            start_row += 1
            summary_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
            start_row += len(summary_df) + 3

        # 8.2 写入推荐保留的因子列表
        recommended = synthesis_rec.get('recommended_factors', [])
        if recommended:
            rec_rows = []
            factor_info = correlation_results.get('factor_info', {})
            for f in recommended:
                info = factor_info.get(f, {})
                rec_rows.append({
                    '推荐保留因子': f,
                    '因子全名': info.get('description', f),
                    'ICIR': round(info.get('icir', 0), 4),
                })
            rec_df = pd.DataFrame(rec_rows)

            n_original = len(factor_info)
            n_recommended = len(recommended)
            header_df = pd.DataFrame({f'【最终精简建议】从{n_original}个因子精简为{n_recommended}个': ['']})
            header_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
            start_row += 1
            rec_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
            start_row += len(rec_df) + 3


def export_to_excel(all_results, output_file='factors_analysis_report.xlsx', windows=LOOKBACK_WINDOWS, correlation_results=None):
    """
    将所有因子分析结果导出到Excel
    每个因子一个sheet页

    参数:
    all_results: dict, 所有因子的分析结果
    output_file: str, 输出文件名
    windows: list, 窗口列表
    correlation_results: dict, 多因子相关性分析结果（可选）
    """
    # 提取耗时信息（如果有）
    factor_timing = all_results.pop('_factor_timing', {})

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 创建所有因子最优窗口汇总（作为第一个sheet）
        if len(all_results) >= 1:
            print("正在导出因子最优窗口汇总...")
            best_factors_df = create_best_factors_summary_df(all_results)
            if not best_factors_df.empty:
                best_factors_df.to_excel(writer, sheet_name='因子最优窗口汇总', index=True)

        # 导出多因子相关性分析结果（作为第二个sheet）
        if correlation_results is not None:
            print("正在导出多因子相关性分析...")
            _export_correlation_sheet(writer, correlation_results)

        for factor_name, factor_results in all_results.items():
            print(f"正在导出因子: {factor_name}")

            # 获取该因子实际使用的窗口列表
            factor_windows = list(factor_results.keys())

            # 获取因子说明
            docstring = get_factor_docstring(factor_name)

            # 获取数据日期范围
            start_date, end_date = get_data_date_range(factor_results, factor_windows)

            # 获取该因子的最优窗口信息（只取最优的1个）
            best_info = find_best_windows(factor_results, top_n=1)

            # 创建汇总表
            summary_df = create_factor_summary_df(factor_name, factor_results, factor_windows)

            # 创建G5持仓记录（按列输出每个窗口）
            g5_holdings_df = create_g5_holdings_df(factor_results, factor_windows)

            # 创建IC累积序列（Pearson IC 和 Rank IC）
            pearson_ic_cumsum_df, rank_ic_cumsum_df = create_ic_cumsum_df(factor_results, factor_windows)

            # 创建分层累积净值
            layer_nav_dict = create_layer_nav_df(factor_results, factor_windows)

            # 创建G5每年收益统计
            g5_yearly_dict = create_g5_yearly_returns_df(factor_results, factor_windows)

            # 写入sheet（使用description作为sheet名，最长31字符）
            if factor_name in FACTOR_CONFIG:
                sheet_name = FACTOR_CONFIG[factor_name].get('description', factor_name)[:31]
            else:
                sheet_name = factor_name[:31]

            # 写入因子说明标题
            start_row = 0
            title_df = pd.DataFrame({f'【因子说明】': [docstring]})
            title_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
            start_row += 3

            # 写入数据日期范围
            if start_date and end_date:
                date_range_df = pd.DataFrame({f'【数据日期范围】': [f'{start_date} 至 {end_date}']})
                date_range_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
                start_row += 3
            else:
                start_row += 1

            # 写入最优窗口信息（只显示IC最优）
            best_ic_str = ', '.join([f"{w}日(IC={v:.4f})" for w, v in best_info['best_ic_windows']])
            best_windows_info = pd.DataFrame({
                '【最优窗口】': [f'IC最优窗口: {best_ic_str}'],
            })
            best_windows_info.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
            start_row += 3

            # 写入汇总表（行为指标，列为窗口，需要写入index）
            if not summary_df.empty:
                # 写入汇总表标题
                header_df = pd.DataFrame({f'【因子汇总指标】': ['']})
                header_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
                start_row += 1
                summary_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=True)
                start_row += len(summary_df) + 3

            # 只对IC最优窗口输出详细的G5每年收益统计
            best_ic_windows = [w for w, _ in best_info['best_ic_windows']]

            for window in best_ic_windows:
                if window in g5_yearly_dict:
                    yearly_df = g5_yearly_dict[window]
                    # 写入标题行
                    header_df = pd.DataFrame({f'【G5每年收益统计 - {window}日窗口】': ['']})
                    header_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
                    start_row += 1
                    yearly_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
                    start_row += len(yearly_df) + 3

            # 只对IC最优窗口输出IC累积序列（合并Pearson IC和Rank IC到一个表）
            ic_chart_info = None  # 记录IC累积序列的位置信息，用于生成图表
            if best_ic_windows:
                best_ic_cols = [f'{w}日' for w in best_ic_windows]

                # 合并IC和RankIC到一个DataFrame
                combined_ic_data = {}
                has_data = False

                for w in best_ic_windows:
                    col_name = f'{w}日'
                    # 添加Pearson IC列
                    if not pearson_ic_cumsum_df.empty and col_name in pearson_ic_cumsum_df.columns:
                        combined_ic_data[f'IC_{w}日'] = pearson_ic_cumsum_df[col_name]
                        has_data = True
                    # 添加Rank IC列
                    if not rank_ic_cumsum_df.empty and col_name in rank_ic_cumsum_df.columns:
                        combined_ic_data[f'RankIC_{w}日'] = rank_ic_cumsum_df[col_name]
                        has_data = True

                if has_data:
                    combined_ic_df = pd.DataFrame(combined_ic_data)
                    # 按列名排序：IC_20日, RankIC_20日, IC_60日, RankIC_60日, ...
                    sorted_cols = []
                    for w in best_ic_windows:
                        if f'IC_{w}日' in combined_ic_df.columns:
                            sorted_cols.append(f'IC_{w}日')
                        if f'RankIC_{w}日' in combined_ic_df.columns:
                            sorted_cols.append(f'RankIC_{w}日')
                    combined_ic_df = combined_ic_df[sorted_cols]

                    header_df = pd.DataFrame({'【IC/RankIC累积序列 - IC最优窗口】': ['']})
                    header_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
                    start_row += 1
                    ic_data_start_row = start_row  # 记录数据起始行
                    combined_ic_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=True)
                    ic_chart_info = {
                        'start_row': ic_data_start_row,
                        'end_row': start_row + len(combined_ic_df),
                        'num_cols': len(sorted_cols) + 1,  # +1 for index column
                        'label': 'IC'
                    }
                    start_row += len(combined_ic_df) + 3

            # 只对IC最优窗口输出分层累积净值
            nav_chart_info = None  # 记录分层净值的位置信息，用于生成图表
            for window in best_ic_windows:
                if window in layer_nav_dict:
                    nav_df = layer_nav_dict[window]
                    # 写入标题行
                    header_df = pd.DataFrame({f'【分层累积净值 - {window}日窗口】': ['']})
                    header_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
                    start_row += 1
                    nav_data_start_row = start_row  # 记录数据起始行
                    nav_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=True)
                    # 只记录第一个窗口的图表信息
                    if nav_chart_info is None:
                        nav_chart_info = {
                            'start_row': nav_data_start_row,
                            'end_row': start_row + len(nav_df),
                            'num_cols': len(nav_df.columns) + 1,  # +1 for index column
                            'window': window
                        }
                    start_row += len(nav_df) + 3

            # 只对IC最优窗口输出G5持仓记录
            if not g5_holdings_df.empty and best_ic_windows:
                # G5持仓记录的列名是整数（20, 60, 120, 240）
                best_cols = [w for w in best_ic_windows if w in g5_holdings_df.columns]
                if best_cols:
                    g5_holdings_best_df = g5_holdings_df[best_cols]
                    # 写入标题行
                    header_df = pd.DataFrame({f'【G5持仓行业 - IC最优窗口】': ['']})
                    header_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
                    start_row += 1
                    g5_holdings_best_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=True)
                    start_row += len(g5_holdings_best_df) + 3

            # 生成图表（在所有数据写入后）
            # 获取工作表对象
            ws = writer.sheets[sheet_name]

            # 导入图表样式相关模块
            from openpyxl.chart.axis import ChartLines
            from openpyxl.drawing.line import LineProperties

            # 创建虚线网格线样式
            def create_dashed_gridlines():
                """创建虚线网格线"""
                gridlines = ChartLines()
                gridlines.spPr = GraphicalProperties(ln=LineProperties(prstDash='dash'))
                return gridlines

            # 1. 生成IC累积曲线图（同时显示IC和RankIC）
            if ic_chart_info is not None:
                from openpyxl.chart.legend import Legend
                from openpyxl.chart.layout import Layout, ManualLayout
                from openpyxl.drawing.line import LineProperties as DrawingLineProperties

                ic_chart = LineChart()
                ic_chart.title = "IC/RankIC累积曲线"
                ic_chart.style = 2  # 使用简洁样式，避免分段变色
                ic_chart.y_axis.title = "累积值"
                ic_chart.x_axis.title = "日期"
                ic_chart.width = 10
                ic_chart.height = 6

                # 设置网格线为虚线，减少网格线数量
                ic_chart.y_axis.majorGridlines = create_dashed_gridlines()
                ic_chart.x_axis.majorGridlines = None  # 不显示X轴网格线

                # 减少Y轴网格线数量
                ic_chart.y_axis.majorUnit = None  # 让Excel自动计算合适的间隔

                # 显示轴刻度标签
                ic_chart.x_axis.tickLblPos = "low"
                ic_chart.y_axis.tickLblPos = "low"
                ic_chart.x_axis.delete = False
                ic_chart.y_axis.delete = False

                # 设置X轴标签间隔（每隔一定数量显示一个标签）
                data_rows = ic_chart_info['end_row'] - ic_chart_info['start_row']
                ic_chart.x_axis.tickLblSkip = max(1, data_rows // 8)
                ic_chart.x_axis.tickMarkSkip = max(1, data_rows // 8)

                # 数据范围（从第2列开始，跳过日期列）
                data_ref = Reference(ws,
                                    min_col=2,
                                    min_row=ic_chart_info['start_row'] + 1,  # +1 因为Excel行号从1开始
                                    max_col=ic_chart_info['num_cols'],
                                    max_row=ic_chart_info['end_row'] + 1)
                # 日期作为X轴
                cats_ref = Reference(ws,
                                    min_col=1,
                                    min_row=ic_chart_info['start_row'] + 2,  # 跳过标题行
                                    max_row=ic_chart_info['end_row'] + 1)

                ic_chart.add_data(data_ref, titles_from_data=True)
                ic_chart.set_categories(cats_ref)

                # 设置线条颜色：IC黑色(000000)，RankIC蓝色(0000FF)
                for i, series in enumerate(ic_chart.series):
                    if i == 0:  # IC - 黑色
                        series.graphicalProperties.line.solidFill = "000000"
                    elif i == 1:  # RankIC - 蓝色
                        series.graphicalProperties.line.solidFill = "0000FF"

                # 设置图例在左上角（使用手动布局）
                ic_chart.legend = Legend()
                ic_chart.legend.layout = Layout(
                    manualLayout=ManualLayout(
                        xMode='edge',
                        yMode='edge',
                        x=0.02,  # 距左边2%
                        y=0.02,  # 距顶部2%
                    )
                )

                # 将图表添加到工作表（放在H5）
                ws.add_chart(ic_chart, "H5")

            # 2. 生成分层净值曲线图
            if nav_chart_info is not None:
                from openpyxl.chart.legend import Legend
                from openpyxl.chart.layout import Layout, ManualLayout

                nav_chart = LineChart()
                nav_chart.title = f"分层累积净值曲线 - {nav_chart_info['window']}日窗口"
                nav_chart.style = 2  # 使用简洁样式，避免分段变色
                nav_chart.y_axis.title = "净值"
                nav_chart.x_axis.title = "日期"
                nav_chart.width = 18
                nav_chart.height = 10

                # 设置网格线为虚线，减少网格线数量
                nav_chart.y_axis.majorGridlines = create_dashed_gridlines()
                nav_chart.x_axis.majorGridlines = None  # 不显示X轴网格线

                # 显示轴刻度标签
                nav_chart.x_axis.tickLblPos = "low"
                nav_chart.y_axis.tickLblPos = "low"
                nav_chart.x_axis.delete = False
                nav_chart.y_axis.delete = False

                # 设置X轴标签间隔
                data_rows = nav_chart_info['end_row'] - nav_chart_info['start_row']
                nav_chart.x_axis.tickLblSkip = max(1, data_rows // 8)
                nav_chart.x_axis.tickMarkSkip = max(1, data_rows // 8)

                # 数据范围（从第2列开始，跳过日期列）
                # 数据列顺序是 G1, G2, G3, G4, G5, 基准
                # 需要按 G5, G4, G3, G2, G1, 基准 的顺序添加
                num_cols = nav_chart_info['num_cols']

                # 日期作为X轴
                cats_ref = Reference(ws,
                                    min_col=1,
                                    min_row=nav_chart_info['start_row'] + 2,
                                    max_row=nav_chart_info['end_row'] + 1)

                # 按 G5, G4, G3, G2, G1, 基准 顺序添加数据系列
                # 原始列顺序: 2=G1, 3=G2, 4=G3, 5=G4, 6=G5, 7=基准
                # 目标顺序: G5(6), G4(5), G3(4), G2(3), G1(2), 基准(7)
                col_order = [6, 5, 4, 3, 2, 7]  # G5, G4, G3, G2, G1, 基准

                for col_idx in col_order:
                    if col_idx <= num_cols:
                        data_ref = Reference(ws,
                                            min_col=col_idx,
                                            min_row=nav_chart_info['start_row'] + 1,
                                            max_row=nav_chart_info['end_row'] + 1)
                        nav_chart.add_data(data_ref, titles_from_data=True)

                nav_chart.set_categories(cats_ref)

                # 设置图例位置到左上角
                nav_chart.legend = Legend()
                nav_chart.legend.position = 'l'  # 左侧
                nav_chart.legend.layout = Layout(
                    manualLayout=ManualLayout(
                        xMode='edge',
                        yMode='edge',
                        x=0.08,  # 左边距（相对于绘图区）
                        y=0.08,  # 上边距（更靠上）
                        h=0.25,  # 高度
                        w=0.12   # 宽度
                    )
                )

                # 将图表添加到工作表（放在H18）
                ws.add_chart(nav_chart, "H18")

        # 在最后添加计算耗时sheet
        if factor_timing:
            print("正在导出计算耗时统计...")
            timing_data = []
            total_minutes = 0
            for factor_name, minutes in factor_timing.items():
                # 获取因子描述
                if factor_name in FACTOR_CONFIG:
                    factor_desc = FACTOR_CONFIG[factor_name].get('description', factor_name)
                else:
                    factor_desc = factor_name
                timing_data.append({
                    '因子名称': factor_desc,
                    '计算耗时(分钟)': minutes
                })
                total_minutes += minutes
            # 添加总计行
            timing_data.append({
                '因子名称': '【总计】',
                '计算耗时(分钟)': total_minutes
            })
            timing_df = pd.DataFrame(timing_data)
            timing_df.to_excel(writer, sheet_name='计算耗时统计', index=False)

    print(f"\n分析报告已导出到: {output_file}")


def list_factors():
    """列出所有可用因子"""
    return list(FACTOR_CONFIG.keys())


def format_excel_report(file_path: str):
    """
    调整Excel报告文件格式

    功能：
    1. 将每个sheet页的A列设置为左对齐
    2. 将每个sheet页的A列宽度设置为11
    3. 隐藏每个sheet页的22行到55行（调整后的行号）
    4. 动态查找ICIR行，比较各窗口的ICIR值，将最大值所在列加粗
    5. 在"因子最优窗口汇总"sheet中为因子名添加超链接到对应sheet
    6. 在每个因子sheet的A2位置添加返回"因子最优窗口汇总"的链接

    参数:
        file_path: str, Excel文件路径
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font

    # 加载工作簿
    wb = load_workbook(file_path)

    # 获取所有sheet名称，用于建立因子名到sheet名的映射
    all_sheet_names = wb.sheetnames

    # 处理"因子最优窗口汇总"sheet的超链接
    summary_sheet_name = '因子最优窗口汇总'
    if summary_sheet_name in wb.sheetnames:
        ws_summary = wb[summary_sheet_name]

        # 设置A列左对齐
        for row in range(1, ws_summary.max_row + 1):
            cell = ws_summary.cell(row=row, column=1)
            cell.alignment = Alignment(horizontal='left')

        # 设置第1行自动换行、顶对齐、左对齐
        for col in range(1, ws_summary.max_column + 1):
            cell = ws_summary.cell(row=1, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        # 遍历第一行（标题行），从B列开始（A列是指标名）
        for col in range(2, ws_summary.max_column + 1):
            cell = ws_summary.cell(row=1, column=col)
            col_name = cell.value

            if col_name:
                # 列名格式: "因子描述_窗口日(标签)"，例如 "行业间动量1-1-Lasso因子_20日(IC最优 ICIR最优)"
                # 提取因子描述部分（最后一个下划线前的内容）
                # 注意：因子描述本身可能包含下划线，所以用rsplit从右边分割
                parts = col_name.rsplit('_', 1)
                if len(parts) >= 1:
                    factor_desc = parts[0]  # 因子描述部分
                    # sheet名最长31字符
                    target_sheet_name = factor_desc[:31]

                    if target_sheet_name in all_sheet_names:
                        # 添加超链接到对应的sheet
                        cell.hyperlink = f"#'{target_sheet_name}'!A1"
                        # 设置超链接样式（蓝色下划线，自动换行）
                        cell.font = Font(color="0000FF", underline="single")

    # 遍历所有sheet页
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 跳过汇总sheet页
        if sheet_name == '最优窗口汇总' or sheet_name == summary_sheet_name:
            continue

        # 0. 在B1位置添加返回"因子最优窗口汇总"的链接
        if summary_sheet_name in wb.sheetnames:
            return_cell = ws.cell(row=1, column=2)
            return_cell.value = "← 返回汇总"
            return_cell.hyperlink = f"#'{summary_sheet_name}'!A1"
            return_cell.font = Font(color="0000FF", underline="single")

        # 1. 设置A列宽度为11
        ws.column_dimensions['A'].width = 11

        # 2. 设置A列左对齐
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=1)  # A列是第1列
            cell.alignment = Alignment(horizontal='left')

        # 3. 动态查找ICIR行
        icir_row = None
        for row in range(1, min(ws.max_row + 1, 30)):  # 只在前30行查找
            cell_value = ws.cell(row=row, column=1).value
            if cell_value == 'ICIR':
                icir_row = row
                break

        # 4. 隐藏25行到56行
        for row in range(25, 57):  # 25到56行（包含56）
            if row <= ws.max_row:
                ws.row_dimensions[row].hidden = True

        # 5. 比较ICIR行的B-E列值，将原始值最大的列加粗（动量因子是正向因子）
        if icir_row:
            values = {}
            for col in range(2, 6):  # B到E列
                cell_value = ws.cell(row=icir_row, column=col).value
                if cell_value is not None:
                    try:
                        values[col] = float(cell_value)  # 使用原始值比较（不取绝对值）
                    except (ValueError, TypeError):
                        values[col] = float('-inf')
                else:
                    values[col] = float('-inf')

            if values and max(values.values()) > float('-inf'):
                max_col = max(values, key=values.get)

                # 将该列从汇总表标题行到隐藏区域结束的行加粗
                summary_header_row = icir_row - 2  # 汇总表标题行
                bold_end_row = 56  # 加粗到第56行（与隐藏区域一致）
                for row in range(summary_header_row, bold_end_row + 1):
                    if row <= ws.max_row:
                        cell = ws.cell(row=row, column=max_col)
                        # 保留原有字体属性，只修改bold
                        if cell.font:
                            new_font = Font(
                                name=cell.font.name,
                                size=cell.font.size,
                                bold=True,
                                italic=cell.font.italic,
                                color=cell.font.color
                            )
                        else:
                            new_font = Font(bold=True)
                        cell.font = new_font

    # 保存文件
    wb.save(file_path)


if __name__ == "__main__":
    import time
    start_time = time.time()

    print("=" * 60)
    print("因子批量分析")
    print("=" * 60)

    # 配置参数
    MONTHLY_REBALANCE = True  # 使用月度调仓（每月最后一个交易日）
    REBALANCE_FREQ = 20  # 调仓频率（天）- 仅在 MONTHLY_REBALANCE=False 时使用
    WINDOWS_TO_TEST = [20, 60, 120, 240]  # 因子计算的回溯窗口
    LOAD_CONSTITUENT = True  # 是否加载成分股数据（用于计算需要成分股的因子）

    # 固定日期范围（设为None使用全部数据，与单因子测试.py保持一致）
    end_date = None    # 最新日期

    # 加载所有数据（一次性加载）
    data = DataContainer(data_loader.DEFAULT_CACHE_FILE, end_date=end_date,
                         backtest_years=DEFAULT_BACKTEST_YEARS,
                         load_constituent=LOAD_CONSTITUENT)

    # 显示可用因子
    print(f"\n可用因子列表: {list_factors()}")

    # 显示调仓方式
    if MONTHLY_REBALANCE:
        print("\n调仓方式: 每月最后一个交易日调仓")
    else:
        print(f"\n调仓方式: 每 {REBALANCE_FREQ} 个交易日调仓")

    # 分析所有因子
    print("\n开始分析所有因子...")
    all_results = analyze_all_factors(
        data,
        windows=WINDOWS_TO_TEST,
        rebalance_freq=REBALANCE_FREQ,
        monthly_rebalance=MONTHLY_REBALANCE
    )

    # 多因子相关性分析（至少需要2个因子）
    correlation_results = None
    if len(all_results) >= 2:
        correlation_results = analyze_multi_factor_correlation(all_results)

    # 计算耗时
    elapsed_time = time.time() - start_time
    hours = int(elapsed_time // 3600)
    minutes = int((elapsed_time % 3600) // 60)
    duration_str = f"{hours}h{minutes:02d}m" if hours > 0 else f"{minutes}m"

    # 获取输出路径（使用 get_output_path）
    output_file = get_output_path(None, duration_str, end_date)

    # 导出到Excel（包含相关性分析结果）
    print("\n正在导出到Excel...")
    export_to_excel(all_results, output_file, WINDOWS_TO_TEST, correlation_results=correlation_results)

    # 格式化Excel报告
    print("\n正在格式化Excel报告...")
    format_excel_report(output_file)

    print(f"\n分析完成！报告已保存至: {output_file}")
    print(f"总耗时: {hours}小时{minutes}分钟" if hours > 0 else f"总耗时: {minutes}分钟")

