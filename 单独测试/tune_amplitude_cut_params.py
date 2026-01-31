"""
振幅切割动量因子参数调优脚本

测试不同的λ（selection_ratio）和N（window）组合，找出最优参数
"""

import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pandas as pd
import numpy as np
from datetime import datetime
from itertools import product
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from tqdm import tqdm

from data_loader import load_price_df, load_high_df, load_low_df
from factors_analysis import (
    stratified_backtest,
    calculate_ic_ir,
    calculate_monthly_forward_returns,
    calculate_excess_metrics,
    get_last_complete_month_end
)
import factor_


def tune_amplitude_cut_parameters(
    test_periods=None,
    window_range=None,
    lambda_range=None,
    output_dir='factor分析/参数调优'
):
    """
    参数调优主函数

    参数:
        test_periods: list of tuple, 测试时间段列表，例如 [('2010-01-01', '2020-12-31'), ('2015-01-01', '2025-01-31')]
        window_range: list of int, 窗口N的测试范围，例如 [120, 140, 160, 180, 200]
        lambda_range: list of float, λ的测试范围，例如 [0.50, 0.60, 0.70, 0.80]
        output_dir: str, 输出目录
    """

    # 创建输出目录
    os.makedirs(output_dir, exist_ok=True)

    # 加载数据
    print("=" * 80)
    print("加载数据...")
    print("=" * 80)

    prices_df = load_price_df()
    high_df = load_high_df()
    low_df = load_low_df()

    # 默认参数：自动计算测试时间段
    if test_periods is None:
        # 获取最后一个完整月末
        last_complete = get_last_complete_month_end(prices_df.index)
        if last_complete is None:
            last_complete = prices_df.index[-1]
        # 往前推10年的12月31日
        start_year = last_complete.year - 10
        start_date = f"{start_year}-12-31"
        end_date = last_complete.strftime('%Y-%m-%d')
        test_periods = [(start_date, end_date)]
        print(f"  自动计算测试时间段: {start_date} 至 {end_date}")

    if window_range is None:
        # 研报最优160，测试±40范围
        window_range = [120, 140, 160, 180, 200]

    if lambda_range is None:
        # 研报最优0.70，测试0.50-0.80范围
        lambda_range = [0.50, 0.60, 0.70, 0.80]

    # 存储所有结果
    all_results = []

    # 遍历所有参数组合
    total_combinations = len(test_periods) * len(window_range) * len(lambda_range)

    print(f"\n总共需要测试 {total_combinations} 个参数组合")
    print(f"测试时间段: {len(test_periods)} 个")
    print(f"窗口N范围: {window_range}")
    print(f"λ范围: {lambda_range}")
    print("=" * 80)

    for period_idx, (start_date, end_date) in enumerate(test_periods):
        period_name = f"{start_date[:4]}-{end_date[:4]}"

        print(f"\n{'=' * 80}")
        print(f"测试时间段 [{period_idx+1}/{len(test_periods)}]: {period_name}")
        print(f"{'=' * 80}")

        # 直接过滤价格数据（不使用 DataContainer 避免 Barra 数据限制）
        prices_filtered = prices_df.loc[
            (prices_df.index >= start_date) &
            (prices_df.index <= end_date)
        ].copy()

        # 获取最后一个完整月末
        last_complete_month = get_last_complete_month_end(prices_filtered.index)
        print(f"  数据范围: {prices_filtered.index[0].date()} - {prices_filtered.index[-1].date()}")
        print(f"  最后完整月末: {last_complete_month.date() if last_complete_month else 'N/A'}")

        # 生成参数组合
        param_combinations = [(w, l) for w in window_range for l in lambda_range]

        # 使用 tqdm 进度条（单行刷新）
        pbar = tqdm(param_combinations, desc=f"调优 {period_name}",
                    ncols=80, leave=True, dynamic_ncols=False,
                    bar_format='{desc}: {percentage:3.0f}%|{bar}| {n}/{total} [{elapsed}<{remaining}]')

        for window, selection_ratio in pbar:
            pbar.set_description(f"调优 N={window},λ={selection_ratio:.1f}")

            try:
                # 计算因子
                factor_df = factor_.momentum_amplitude_cut(
                    high_df=high_df,
                    low_df=low_df,
                    prices_df=prices_df,
                    window=window,
                    selection_ratio=selection_ratio
                )

                # 过滤日期范围
                factor_df = factor_df.loc[
                    (factor_df.index >= start_date) &
                    (factor_df.index <= end_date)
                ]

                # 计算IC/IR
                forward_returns_df = calculate_monthly_forward_returns(
                    prices_filtered, prices_filtered.index
                )
                ic_series, ic_cumsum, ic_mean, ic_std, icir, ic_win_rate, ic_abs_mean = calculate_ic_ir(
                    factor_df, forward_returns_df, monthly_rebalance=True
                )

                # 分层回测
                nav_df, layer_returns, layer_holdings = stratified_backtest(
                    factor_df, prices_filtered, window,
                    n_layers=5, monthly_rebalance=True, rebalance_type='monthly'
                )

                # 计算基准净值（等权行业指数）
                benchmark_nav = pd.Series(index=nav_df.index, dtype=float)
                benchmark_nav.iloc[0] = 1.0
                for i in range(1, len(nav_df.index)):
                    prev_date = nav_df.index[i-1]
                    curr_date = nav_df.index[i]
                    period_ret = (prices_filtered.loc[curr_date] / prices_filtered.loc[prev_date] - 1).mean()
                    benchmark_nav.iloc[i] = benchmark_nav.iloc[i-1] * (1 + period_ret)

                # 计算超额指标
                excess_metrics = calculate_excess_metrics(
                    nav_df, benchmark_nav, monthly_rebalance=True, rebalance_type='monthly'
                )

                # 提取关键指标
                g5_metrics = excess_metrics.get('G5', {})
                g5_annual_return = g5_metrics.get('年化收益率(%)', np.nan)
                g5_excess_return = g5_metrics.get('超额年化收益率(%)', np.nan)
                g5_sharpe = g5_metrics.get('夏普比率', np.nan)
                g5_max_drawdown = g5_metrics.get('最大回撤(%)', np.nan)

                # 多空组合指标
                ls_metrics = excess_metrics.get('多空组合', {})
                ls_annual_return = ls_metrics.get('年化收益率(%)', np.nan)
                ls_sharpe = ls_metrics.get('夏普比率', np.nan)
                ls_max_drawdown = ls_metrics.get('最大回撤(%)', np.nan)

                # 存储结果
                result = {
                    '测试时间段': period_name,
                    '窗口N': window,
                    'λ': selection_ratio,
                    'IC均值': ic_mean,
                    'ICIR': icir,
                    'G5年化收益(%)': g5_annual_return,
                    'G5超额收益(%)': g5_excess_return,
                    'G5 Sharpe': g5_sharpe,
                    'G5最大回撤(%)': g5_max_drawdown,
                    '多空年化收益(%)': ls_annual_return,
                    '多空Sharpe': ls_sharpe,
                    '多空最大回撤(%)': ls_max_drawdown
                }

                all_results.append(result)

            except Exception as e:
                tqdm.write(f"  ❌ N={window}, λ={selection_ratio:.2f} 计算失败: {str(e)}")
                continue

    # 转换为DataFrame
    results_df = pd.DataFrame(all_results)

    # 生成报告
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = os.path.join(output_dir, f'amplitude_cut_tuning_{timestamp}.xlsx')

    print(f"\n{'=' * 80}")
    print("生成Excel报告...")
    print(f"{'=' * 80}")

    generate_tuning_report(results_df, output_file, test_periods, window_range, lambda_range)

    print(f"\n✅ 参数调优完成！")
    print(f"📊 结果已保存至: {output_file}")

    return results_df


def generate_tuning_report(results_df, output_file, test_periods, window_range, lambda_range):
    """
    生成格式化的Excel调优报告
    """

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:

        # Sheet 1: 完整结果表
        results_df.to_excel(writer, sheet_name='完整结果', index=False)

        # Sheet 2-N: 每个时间段的热力图
        for period_idx, (start_date, end_date) in enumerate(test_periods):
            period_name = f"{start_date[:4]}-{end_date[:4]}"
            period_data = results_df[results_df['测试时间段'] == period_name].copy()

            if len(period_data) == 0:
                continue

            # 为每个关键指标创建热力图
            metrics = [
                ('ICIR', 'ICIR热力图'),
                ('G5超额收益(%)', 'G5超额收益热力图'),
                ('多空年化收益(%)', '多空收益热力图')
            ]

            for metric_col, sheet_suffix in metrics:
                sheet_name = f"{period_name}_{sheet_suffix}"[:31]  # Excel sheet名称限制31字符

                # 创建透视表（λ为行，N为列）
                pivot_table = period_data.pivot_table(
                    values=metric_col,
                    index='λ',
                    columns='窗口N',
                    aggfunc='first'
                )

                # 写入Excel
                pivot_table.to_excel(writer, sheet_name=sheet_name)

        # Sheet: 最优参数汇总
        summary_data = []

        for period_idx, (start_date, end_date) in enumerate(test_periods):
            period_name = f"{start_date[:4]}-{end_date[:4]}"
            period_data = results_df[results_df['测试时间段'] == period_name].copy()

            if len(period_data) == 0:
                continue

            # 找出各指标的最优参数（处理 NaN 值）
            # ICIR
            icir_valid = period_data['ICIR'].dropna()
            if len(icir_valid) > 0:
                best_icir = period_data.loc[icir_valid.idxmax()]
            else:
                best_icir = period_data.iloc[0]

            # G5超额收益
            g5_valid = period_data['G5超额收益(%)'].dropna()
            if len(g5_valid) > 0:
                best_g5_excess = period_data.loc[g5_valid.idxmax()]
            else:
                best_g5_excess = period_data.iloc[0]

            # 多空收益
            ls_valid = period_data['多空年化收益(%)'].dropna()
            if len(ls_valid) > 0:
                best_ls_return = period_data.loc[ls_valid.idxmax()]
            else:
                best_ls_return = period_data.iloc[0]

            summary_data.append({
                '测试时间段': period_name,
                '优化目标': 'ICIR最大',
                '最优N': int(best_icir['窗口N']),
                '最优λ': best_icir['λ'],
                'ICIR': best_icir['ICIR'],
                'G5超额收益(%)': best_icir['G5超额收益(%)'],
                '多空年化收益(%)': best_icir['多空年化收益(%)']
            })

            summary_data.append({
                '测试时间段': period_name,
                '优化目标': 'G5超额收益最大',
                '最优N': int(best_g5_excess['窗口N']),
                '最优λ': best_g5_excess['λ'],
                'ICIR': best_g5_excess['ICIR'],
                'G5超额收益(%)': best_g5_excess['G5超额收益(%)'],
                '多空年化收益(%)': best_g5_excess['多空年化收益(%)']
            })

            summary_data.append({
                '测试时间段': period_name,
                '优化目标': '多空收益最大',
                '最优N': int(best_ls_return['窗口N']),
                '最优λ': best_ls_return['λ'],
                'ICIR': best_ls_return['ICIR'],
                'G5超额收益(%)': best_ls_return['G5超额收益(%)'],
                '多空年化收益(%)': best_ls_return['多空年化收益(%)']
            })

        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='最优参数汇总', index=False)

    # 格式化Excel
    format_tuning_excel(output_file)


def format_tuning_excel(file_path):
    """
    格式化Excel报告
    """
    wb = openpyxl.load_workbook(file_path)

    # 定义样式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 格式化每个sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 设置表头样式
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # 设置数据区域样式
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width

        # 冻结首行
        ws.freeze_panes = 'A2'

    wb.save(file_path)


if __name__ == '__main__':

    # 配置参数范围
    TEST_PERIODS = None  # 自动计算：最新数据往前推10年

    # 窗口N范围：研报最优160，测试120-200
    WINDOW_RANGE = [120, 140, 160, 180, 200]

    # λ范围：研报最优0.70，测试0.50-0.80
    LAMBDA_RANGE = [0.50, 0.60, 0.70, 0.80]

    # 运行参数调优
    results = tune_amplitude_cut_parameters(
        test_periods=TEST_PERIODS,
        window_range=WINDOW_RANGE,
        lambda_range=LAMBDA_RANGE
    )

    # 打印最优参数汇总
    print("\n" + "=" * 80)
    print("最优参数汇总")
    print("=" * 80)

    # 从结果中获取时间段
    if len(results) > 0:
        for period_name in results['测试时间段'].unique():
            period_data = results[results['测试时间段'] == period_name]

            if len(period_data) == 0:
                continue

            print(f"\n时间段: {period_name}")
            print("-" * 80)

            # ICIR最大（处理 NaN）
            icir_valid = period_data['ICIR'].dropna()
            if len(icir_valid) > 0:
                best_icir = period_data.loc[icir_valid.idxmax()]
                print(f"ICIR最大: N={int(best_icir['窗口N'])}, λ={best_icir['λ']:.2f} "
                      f"→ ICIR={best_icir['ICIR']:.2f}, G5超额={best_icir['G5超额收益(%)']:.2f}%")

            # G5超额收益最大
            g5_valid = period_data['G5超额收益(%)'].dropna()
            if len(g5_valid) > 0:
                best_g5 = period_data.loc[g5_valid.idxmax()]
                print(f"G5超额最大: N={int(best_g5['窗口N'])}, λ={best_g5['λ']:.2f} "
                      f"→ ICIR={best_g5['ICIR']:.2f}, G5超额={best_g5['G5超额收益(%)']:.2f}%")

            # 多空收益最大
            ls_valid = period_data['多空年化收益(%)'].dropna()
            if len(ls_valid) > 0:
                best_ls = period_data.loc[ls_valid.idxmax()]
                print(f"多空收益最大: N={int(best_ls['窗口N'])}, λ={best_ls['λ']:.2f} "
                      f"→ ICIR={best_ls['ICIR']:.2f}, 多空收益={best_ls['多空年化收益(%)']:.2f}%")
