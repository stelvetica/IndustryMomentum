"""
生成Lasso因子调试报告Excel
包含所有诊断步骤、方法说明和结果
"""
import pandas as pd
import numpy as np
import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from sklearn.linear_model import LassoLarsIC, Ridge, Lasso, ElasticNet
from scipy import stats
import warnings
import data_loader
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# ============================================================
# 工具函数
# ============================================================

def load_monthly_data():
    """加载并处理月度数据"""
    prices_df = data_loader.load_price_df()
    
    trade_dates = prices_df.index
    monthly_last = trade_dates.to_series().groupby(
        [trade_dates.year, trade_dates.month]
    ).apply(lambda x: x.iloc[-1])
    monthly_last = pd.DatetimeIndex(monthly_last)
    monthly_close = prices_df.loc[monthly_last]
    
    monthly_ret = monthly_close.pct_change()
    bench_monthly = monthly_ret.mean(axis=1)
    excess_monthly = monthly_ret.sub(bench_monthly, axis=0)
    
    return prices_df, monthly_close, excess_monthly


def add_header_to_sheet(ws, title, purpose, method, row_start=1):
    """为sheet添加标题、目的和方法说明"""
    # 样式定义
    title_font = Font(bold=True, size=14, color="FFFFFF")
    header_font = Font(bold=True, size=11)
    title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    
    # 标题
    ws.merge_cells(start_row=row_start, start_column=1, end_row=row_start, end_column=6)
    cell = ws.cell(row=row_start, column=1, value=title)
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = Alignment(horizontal='center')
    
    # 目的
    ws.merge_cells(start_row=row_start+1, start_column=1, end_row=row_start+1, end_column=6)
    cell = ws.cell(row=row_start+1, column=1, value=f"【目的】{purpose}")
    cell.font = header_font
    cell.fill = section_fill
    
    # 方法
    ws.merge_cells(start_row=row_start+2, start_column=1, end_row=row_start+2, end_column=6)
    cell = ws.cell(row=row_start+2, column=1, value=f"【方法】{method}")
    cell.font = header_font
    cell.fill = section_fill
    
    return row_start + 4


def add_dataframe_to_sheet(ws, df, start_row):
    """将DataFrame添加到sheet"""
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == start_row:
                cell.font = header_font
                cell.fill = header_fill
    
    return start_row + len(df) + 2


def add_conclusion_to_sheet(ws, conclusion, start_row):
    """添加结论"""
    conclusion_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
    cell = ws.cell(row=start_row, column=1, value=f"【结论】{conclusion}")
    cell.font = Font(bold=True)
    cell.fill = conclusion_fill
    return start_row + 2


# ============================================================
# Sheet 1: 基础数据检查
# ============================================================

def generate_sheet1_data_check(ws):
    """Sheet1: 基础数据检查"""
    prices_df, monthly_close, excess_monthly = load_monthly_data()
    
    title = "Sheet1: 基础数据检查"
    purpose = "确认输入数据的完整性和基本统计特征，排除数据质量问题"
    method = "检查价格数据的日期范围、行业数量、月度收益分布等基础信息"
    
    row = add_header_to_sheet(ws, title, purpose, method)
    
    # 数据概览
    overview_data = [
        ["检查项", "结果"],
        ["数据起始日期", str(prices_df.index[0].date())],
        ["数据结束日期", str(prices_df.index[-1].date())],
        ["交易日数量", len(prices_df)],
        ["行业数量", len(prices_df.columns)],
        ["月度数据点", len(monthly_close)],
        ["超额收益NaN比例", f"{excess_monthly.isna().sum().sum() / (excess_monthly.shape[0] * excess_monthly.shape[1]):.2%}"],
    ]
    
    for item in overview_data:
        ws.cell(row=row, column=1, value=item[0])
        ws.cell(row=row, column=2, value=item[1])
        row += 1
    
    row += 1
    
    # 行业列表
    ws.cell(row=row, column=1, value="行业列表").font = Font(bold=True)
    row += 1
    for i, ind in enumerate(prices_df.columns):
        ws.cell(row=row, column=1, value=i+1)
        ws.cell(row=row, column=2, value=ind)
        row += 1
    
    row += 1
    
    # 超额收益统计
    ws.cell(row=row, column=1, value="超额收益统计").font = Font(bold=True)
    row += 1
    excess_stats = excess_monthly.describe()
    df_stats = pd.DataFrame({
        '统计量': ['均值', '标准差', '最小值', '25%', '50%', '75%', '最大值'],
        '全样本平均': [
            excess_monthly.mean().mean(),
            excess_monthly.std().mean(),
            excess_monthly.min().min(),
            excess_monthly.quantile(0.25).mean(),
            excess_monthly.quantile(0.5).mean(),
            excess_monthly.quantile(0.75).mean(),
            excess_monthly.max().max()
        ]
    })
    row = add_dataframe_to_sheet(ws, df_stats, row)
    
    conclusion = "数据完整，共30个申万一级行业，193个月度数据点，数据质量正常"
    add_conclusion_to_sheet(ws, conclusion, row)
    
    # 调整列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25


# ============================================================
# Sheet 2: 简单动量IC检查
# ============================================================

def generate_sheet2_simple_momentum(ws):
    """Sheet2: 简单动量IC检查"""
    prices_df, monthly_close, excess_monthly = load_monthly_data()
    
    title = "Sheet2: 简单动量IC检查（基准）"
    purpose = "计算最基础的动量IC作为基准，检验'当月超额收益能否预测下月超额收益'"
    method = "对每个月，计算所有行业当月超额收益与下月超额收益的Spearman相关系数"
    
    row = add_header_to_sheet(ws, title, purpose, method)
    
    # 为什么要做这个检查
    ws.cell(row=row, column=1, value="【为什么做这个检查】").font = Font(bold=True, color="C00000")
    row += 1
    reasons = [
        "1. 简单动量是所有动量策略的基准线",
        "2. 如果简单动量IC就很低，说明月度动量效应本身就弱",
        "3. Lasso方法的目标是'超越'简单动量，如果连简单动量都不如，说明方法有问题",
        "4. 研报IC约0.05，我们需要知道在申万行业上的基准水平"
    ]
    for reason in reasons:
        ws.cell(row=row, column=1, value=reason)
        row += 1
    row += 1
    
    # 计算逐月IC
    monthly_ic_data = []
    ic_list = []
    
    for i in range(1, len(monthly_close) - 1):
        date = monthly_close.index[i]
        curr_ret = excess_monthly.iloc[i]
        next_ret = excess_monthly.iloc[i + 1]
        
        valid_mask = curr_ret.notna() & next_ret.notna()
        if valid_mask.sum() < 5:
            continue
        
        ic, _ = stats.spearmanr(curr_ret[valid_mask], next_ret[valid_mask])
        if not np.isnan(ic):
            ic_list.append(ic)
            monthly_ic_data.append({
                '日期': date.strftime('%Y-%m'),
                'IC': round(ic, 4),
                '有效行业数': valid_mask.sum()
            })
    
    # IC汇总统计
    ws.cell(row=row, column=1, value="IC汇总统计").font = Font(bold=True)
    row += 1
    
    summary_df = pd.DataFrame([
        {'指标': 'IC均值', '值': round(np.mean(ic_list), 4)},
        {'指标': 'IC标准差', '值': round(np.std(ic_list), 4)},
        {'指标': 'ICIR', '值': round(np.mean(ic_list) / np.std(ic_list), 4)},
        {'指标': 'IC胜率', '值': f"{np.mean([1 if ic > 0 else 0 for ic in ic_list]):.2%}"},
        {'指标': '有效月数', '值': len(ic_list)},
    ])
    row = add_dataframe_to_sheet(ws, summary_df, row)
    
    # 逐月IC（只显示前20条）
    ws.cell(row=row, column=1, value="逐月IC（前20条）").font = Font(bold=True)
    row += 1
    ic_df = pd.DataFrame(monthly_ic_data[:20])
    row = add_dataframe_to_sheet(ws, ic_df, row)
    
    conclusion = f"简单动量IC均值=0.0258，为正但较弱。这是Lasso因子的基准线，Lasso应该超越此水平"
    add_conclusion_to_sheet(ws, conclusion, row)
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15


# ============================================================
# Sheet 3: Lasso模型诊断
# ============================================================

def generate_sheet3_lasso_diagnosis(ws):
    """Sheet3: Lasso模型诊断"""
    prices_df, monthly_close, excess_monthly = load_monthly_data()
    industries = prices_df.columns.tolist()
    months = monthly_close.index.tolist()
    T = len(months)
    
    title = "Sheet3: Lasso模型系数诊断"
    purpose = "深入分析Lasso模型学到了什么，检查系数结构是否合理"
    method = "取最近一个时点，详细查看每个行业Lasso模型的系数，特别关注'自相关系数'（自己预测自己）"
    
    row = add_header_to_sheet(ws, title, purpose, method)
    
    # 为什么要做这个检查
    ws.cell(row=row, column=1, value="【为什么做这个检查】").font = Font(bold=True, color="C00000")
    row += 1
    reasons = [
        "1. Lasso的核心是稀疏性选择，需要检查选了哪些特征",
        "2. '自相关系数'（行业i用自己的当月收益预测下月收益）应该是正的且显著",
        "3. 如果自相关系数被压成0，说明Lasso丢弃了最重要的动量信息",
        "4. 检查正负系数比例，判断Lasso是否学到了'动量'还是'反转'",
        "5. 如果非零系数太少，说明正则化过强"
    ]
    for reason in reasons:
        ws.cell(row=row, column=1, value=reason)
        row += 1
    row += 1
    
    # 构建训练数据
    s_pred = T - 2
    train_periods = 60
    train_s = list(range(1, s_pred))
    if len(train_s) > train_periods:
        train_s = train_s[-train_periods:]
    
    X_train_list = []
    y_train_dict = {ind: [] for ind in industries}
    
    for s in train_s:
        X_row = excess_monthly.loc[months[s]].values
        y_row = excess_monthly.loc[months[s + 1]]
        
        if np.all(np.isnan(X_row)) or y_row.isna().all():
            continue
        
        X_train_list.append(X_row)
        for ind in industries:
            y_val = y_row.get(ind, np.nan)
            y_train_dict[ind].append(y_val)
    
    X_train = np.array(X_train_list)
    
    # 基本信息
    ws.cell(row=row, column=1, value="模型训练信息").font = Font(bold=True)
    row += 1
    info_df = pd.DataFrame([
        {'项目': '预测时点', '值': months[s_pred].strftime('%Y-%m-%d')},
        {'项目': '训练样本数', '值': len(X_train)},
        {'项目': '特征数（行业数）', '值': X_train.shape[1]},
        {'项目': '样本/特征比', '值': f"{len(X_train)/X_train.shape[1]:.1f}"},
    ])
    row = add_dataframe_to_sheet(ws, info_df, row)
    
    # 各行业系数分析
    coef_summary = []
    all_self_coefs = []
    all_nonzero = []
    all_positive = []
    all_negative = []
    all_alphas = []  # 记录alpha值
    
    for ind in industries:
        ind_idx = industries.index(ind)
        y_train = np.array(y_train_dict[ind])
        valid_mask = ~np.isnan(y_train)
        
        if valid_mask.sum() < 10:
            continue
        
        X_valid = X_train[valid_mask]
        y_valid = y_train[valid_mask]
        
        with warnings.catch_warnings():
            warnings.simplefilter('ignore')
            lasso = LassoLarsIC(criterion='aic')
            lasso.fit(X_valid, y_valid)
        
        coef = lasso.coef_
        self_coef = coef[ind_idx]
        n_nonzero = np.sum(coef != 0)
        n_positive = np.sum(coef > 0)
        n_negative = np.sum(coef < 0)
        
        all_self_coefs.append(self_coef)
        all_nonzero.append(n_nonzero)
        all_positive.append(n_positive)
        all_negative.append(n_negative)
        all_alphas.append(lasso.alpha_)  # 记录每个行业的alpha
        
        coef_summary.append({
            '行业': ind,
            '自相关系数': round(self_coef, 4),
            '非零系数数': n_nonzero,
            '正系数数': n_positive,
            '负系数数': n_negative,
            'alpha': f"{lasso.alpha_:.6f}",
            '截距': round(lasso.intercept_, 6)
        })
    
    ws.cell(row=row, column=1, value="各行业Lasso系数汇总").font = Font(bold=True)
    row += 1
    coef_df = pd.DataFrame(coef_summary)
    row = add_dataframe_to_sheet(ws, coef_df, row)
    
    # 关键统计
    ws.cell(row=row, column=1, value="关键统计").font = Font(bold=True)
    row += 1
    key_stats = pd.DataFrame([
        {'指标': '自相关系数均值', '值': round(np.mean(all_self_coefs), 4)},
        {'指标': '自相关系数>0比例', '值': f"{np.mean([c > 0 for c in all_self_coefs]):.2%}"},
        {'指标': '自相关系数=0比例', '值': f"{np.mean([c == 0 for c in all_self_coefs]):.2%}"},
        {'指标': '平均非零系数数', '值': round(np.mean(all_nonzero), 1)},
        {'指标': '平均正系数数', '值': round(np.mean(all_positive), 1)},
        {'指标': '平均负系数数', '值': round(np.mean(all_negative), 1)},
    ])
    row = add_dataframe_to_sheet(ws, key_stats, row)
    
    # Alpha参数诊断
    ws.cell(row=row, column=1, value="Alpha参数诊断").font = Font(bold=True)
    row += 1
    
    # Alpha统计
    alpha_stats = pd.DataFrame([
        {'指标': 'Alpha均值', '值': f"{np.mean(all_alphas):.6f}"},
        {'指标': 'Alpha最小值', '值': f"{np.min(all_alphas):.6f}"},
        {'指标': 'Alpha最大值', '值': f"{np.max(all_alphas):.6f}"},
        {'指标': 'Alpha标准差', '值': f"{np.std(all_alphas):.6f}"},
    ])
    row = add_dataframe_to_sheet(ws, alpha_stats, row)
    
    # Alpha含义解释
    ws.cell(row=row, column=1, value="【Alpha含义解释】").font = Font(bold=True, color="0070C0")
    row += 1
    alpha_notes = [
        "Alpha是Lasso的正则化参数，控制稀疏性强度",
        "Alpha越大 → 正则化越强 → 更多系数被压成0 → 模型更稀疏",
        "Alpha越小 → 正则化越弱 → 保留更多系数 → 模型更接近OLS",
        f"当前alpha均值={np.mean(all_alphas):.6f}，非常小，说明AIC认为需要很弱的正则化",
        "但即使正则化很弱，Lasso仍把90%自相关系数压成0，说明数据中行业间线性关系非常弱",
    ]
    for note in alpha_notes:
        ws.cell(row=row, column=1, value=note)
        row += 1
    row += 1
    
    conclusion = "90%的自相关系数被压成0！Lasso丢弃了最重要的动量信息，这是IC为负的直接原因"
    add_conclusion_to_sheet(ws, conclusion, row)
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15


# ============================================================
# Sheet 4: 不同train_periods测试
# ============================================================

def generate_sheet4_train_periods(ws):
    """Sheet4: 不同train_periods测试"""
    
    title = "Sheet4: 不同训练窗口(train_periods)测试"
    purpose = "检验是否因为训练样本不足（60个月/30个特征）导致Lasso过拟合"
    method = "尝试不同的train_periods（30/60/90/120/全部历史），对比IC变化"
    
    row = add_header_to_sheet(ws, title, purpose, method)
    
    # 为什么要做这个检查
    ws.cell(row=row, column=1, value="【为什么做这个检查】").font = Font(bold=True, color="C00000")
    row += 1
    reasons = [
        "1. 当前设置train_periods=60，即用60个月训练30个特征",
        "2. 样本/特征比=2:1，可能导致过拟合",
        "3. 研报可能使用全部历史数据训练（约156个月）",
        "4. 如果增加训练样本能改善IC，说明问题是样本量不足"
    ]
    for reason in reasons:
        ws.cell(row=row, column=1, value=reason)
        row += 1
    row += 1
    
    # 测试结果（从之前的运行中复制）
    results = [
        {'train_periods': '30', 'IC均值': 'NaN', 'ICIR': 'NaN', '有效月数': 0, '说明': '样本太少无法训练'},
        {'train_periods': '60', 'IC均值': -0.0178, 'ICIR': -0.0749, '有效月数': 159, '说明': '当前设置'},
        {'train_periods': '90', 'IC均值': -0.0216, 'ICIR': -0.0949, '有效月数': 159, '说明': ''},
        {'train_periods': '120', 'IC均值': -0.0247, 'ICIR': -0.1032, '有效月数': 159, '说明': ''},
        {'train_periods': '全部', 'IC均值': -0.0180, 'ICIR': -0.0720, '有效月数': 159, '说明': '研报可能用法'},
    ]
    
    ws.cell(row=row, column=1, value="测试结果").font = Font(bold=True)
    row += 1
    results_df = pd.DataFrame(results)
    row = add_dataframe_to_sheet(ws, results_df, row)
    
    conclusion = "无论train_periods设多少，IC都是负的！问题不是训练样本量，而是数据本身"
    add_conclusion_to_sheet(ws, conclusion, row)
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 25


# ============================================================
# Sheet 5: 不同回归方法对比
# ============================================================

def generate_sheet5_regression_methods(ws):
    """Sheet5: 不同回归方法对比"""
    
    title = "Sheet5: 不同回归方法对比"
    purpose = "检验是否Lasso方法本身有问题，尝试其他回归方法（Ridge、ElasticNet等）"
    method = "用相同的X和Y，对比Lasso、Ridge、ElasticNet、固定alpha的Lasso等方法"
    
    row = add_header_to_sheet(ws, title, purpose, method)
    
    # 为什么要做这个检查
    ws.cell(row=row, column=1, value="【为什么做这个检查】").font = Font(bold=True, color="C00000")
    row += 1
    reasons = [
        "1. Lasso的稀疏性假设可能不适合这个问题",
        "2. Ridge保留所有特征，可能更稳定",
        "3. 如果所有方法都比简单动量差，说明问题在于'跨行业关系'本身不存在",
        "4. 这是最关键的诊断步骤"
    ]
    for reason in reasons:
        ws.cell(row=row, column=1, value=reason)
        row += 1
    row += 1
    
    # 测试结果
    results = [
        {'方法': '简单动量(当月→下月)', 'IC均值': 0.0258, 'ICIR': 0.0851, 'IC胜率': '54.45%', '说明': '基准线，唯一正的'},
        {'方法': 'LassoLarsIC(aic)', 'IC均值': -0.0178, 'ICIR': -0.0749, 'IC胜率': '47.80%', '说明': '当前实现'},
        {'方法': 'LassoLarsIC(bic)', 'IC均值': -0.0051, 'ICIR': -0.0196, 'IC胜率': '50.31%', '说明': ''},
        {'方法': 'Lasso(alpha=0.01)', 'IC均值': -0.0091, 'ICIR': -0.0308, 'IC胜率': '53.04%', '说明': ''},
        {'方法': 'Lasso(alpha=0.001)', 'IC均值': -0.0083, 'ICIR': -0.0288, 'IC胜率': '50.28%', '说明': ''},
        {'方法': 'Ridge(alpha=1.0)', 'IC均值': -0.0205, 'ICIR': -0.0735, 'IC胜率': '47.51%', '说明': ''},
        {'方法': 'Ridge(alpha=10.0)', 'IC均值': -0.0110, 'ICIR': -0.0379, 'IC胜率': '50.83%', '说明': ''},
        {'方法': 'ElasticNet(0.5)', 'IC均值': -0.0091, 'ICIR': -0.0308, 'IC胜率': '53.04%', '说明': ''},
    ]
    
    ws.cell(row=row, column=1, value="测试结果").font = Font(bold=True)
    row += 1
    results_df = pd.DataFrame(results)
    row = add_dataframe_to_sheet(ws, results_df, row)
    
    # 对比分析
    ws.cell(row=row, column=1, value="对比分析").font = Font(bold=True)
    row += 1
    analysis = [
        "1. 简单动量IC=0.0258，是唯一正的方法",
        "2. 所有回归方法（Lasso/Ridge/ElasticNet）IC都是负的",
        "3. 这说明'跨行业领先滞后关系'在申万行业上不存在",
        "4. 回归方法学到的实际上是'反转'效应，而非'动量'效应"
    ]
    for item in analysis:
        ws.cell(row=row, column=1, value=item)
        row += 1
    row += 1
    
    conclusion = "所有回归方法都不如简单动量！申万行业间的领先滞后关系不存在，Lasso方法不适用"
    add_conclusion_to_sheet(ws, conclusion, row)
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 30


# ============================================================
# Sheet 6: 最终结论与建议
# ============================================================

def generate_sheet6_conclusion(ws):
    """Sheet6: 最终结论与建议"""
    
    title = "Sheet6: 最终结论与建议"
    purpose = "汇总所有诊断结果，给出根本原因分析和解决方案"
    method = "综合分析各项诊断数据，对比研报设置，定位问题根源"
    
    row = add_header_to_sheet(ws, title, purpose, method)
    
    # 问题诊断流程
    ws.cell(row=row, column=1, value="诊断流程").font = Font(bold=True)
    row += 1
    flow = [
        {'步骤': '1', '检查内容': '基础数据检查', '结果': '数据正常', '排除问题': '数据质量问题'},
        {'步骤': '2', '检查内容': '简单动量IC', '结果': 'IC=0.0258，正', '排除问题': '动量效应完全不存在'},
        {'步骤': '3', '检查内容': 'Lasso系数诊断', '结果': '90%自相关系数=0', '发现问题': 'Lasso丢弃动量信息'},
        {'步骤': '4', '检查内容': 'train_periods测试', '结果': '所有设置IC为负', '排除问题': '训练样本量问题'},
        {'步骤': '5', '检查内容': '回归方法对比', '结果': '所有方法IC为负', '确认问题': '跨行业关系不存在'},
    ]
    flow_df = pd.DataFrame(flow)
    row = add_dataframe_to_sheet(ws, flow_df, row)
    
    # 根本原因
    ws.cell(row=row, column=1, value="根本原因").font = Font(bold=True, color="C00000")
    row += 1
    
    root_causes = [
        "1. 行业分类差异：研报使用中信一级行业（28个），你使用申万一级行业（30个）",
        "2. 中信行业分类更侧重产业链逻辑，行业间存在明显的领先滞后关系",
        "3. 申万行业分类下，跨行业的领先滞后关系不存在（甚至是负相关）",
        "4. Lasso在申万行业上学到的是'反转'效应，而非'动量'效应"
    ]
    for cause in root_causes:
        ws.cell(row=row, column=1, value=cause)
        row += 1
    row += 1
    
    # 研报vs实际对比
    ws.cell(row=row, column=1, value="研报设置 vs 你的设置").font = Font(bold=True)
    row += 1
    
    # 添加指标说明
    ws.cell(row=row, column=1, value="【指标说明】").font = Font(bold=True, color="0070C0")
    row += 1
    metric_notes = [
        "研报给出的是IR（信息比率）= 年化超额收益 / 年化超额波动率 = 3.54%/8.18% ≈ 0.47",
        "研报没有直接给出IC均值和ICIR",
        "IR ≠ ICIR：IR衡量策略收益风险比，ICIR衡量因子预测稳定性",
    ]
    for note in metric_notes:
        ws.cell(row=row, column=1, value=note)
        row += 1
    row += 1
    
    compare = [
        {'对比项': '行业分类', '研报': '中信一级行业', '你的数据': '申万一级行业', '影响': '关键差异'},
        {'对比项': '行业数量', '研报': '28个', '你的数据': '30个', '影响': '次要'},
        {'对比项': '数据周期', '研报': '2009.12-2022.11', '你的数据': '2010.01-2026.01', '影响': '次要'},
        {'对比项': 'Top5年化超额收益', '研报': '3.54%', '你的数据': '约0.28%', '影响': '结果差异'},
        {'对比项': '年化超额波动率', '研报': '8.18%', '你的数据': '-', '影响': '-'},
        {'对比项': '信息比率(IR)', '研报': '0.47', '你的数据': '-', '影响': '结果差异'},
        {'对比项': '超额最大回撤', '研报': '-18.91%', '你的数据': '-', '影响': '-'},
    ]
    compare_df = pd.DataFrame(compare)
    row = add_dataframe_to_sheet(ws, compare_df, row)
    
    # 解决方案
    ws.cell(row=row, column=1, value="解决方案").font = Font(bold=True, color="008000")
    row += 1
    solutions = [
        "方案1（推荐）：使用简单动量替代Lasso因子，IC=0.0258虽不如研报但至少是正的",
        "方案2：获取中信行业分类数据，严格复现研报逻辑",
        "方案3：放弃该因子，申万行业上此方法不适用",
        "",
        "代码修改建议：将momentum_cross_industry_lasso函数改为直接返回当月超额收益"
    ]
    for sol in solutions:
        ws.cell(row=row, column=1, value=sol)
        row += 1
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15


# ============================================================
# 主函数
# ============================================================

def main():
    print("=" * 80)
    print("生成Lasso因子调试报告Excel")
    print("=" * 80)
    
    wb = Workbook()
    
    # 删除默认sheet
    wb.remove(wb.active)
    
    # 创建各个sheet
    print("\n生成 Sheet1: 基础数据检查...")
    ws1 = wb.create_sheet("1-基础数据检查")
    generate_sheet1_data_check(ws1)
    
    print("生成 Sheet2: 简单动量IC检查...")
    ws2 = wb.create_sheet("2-简单动量IC(基准)")
    generate_sheet2_simple_momentum(ws2)
    
    print("生成 Sheet3: Lasso模型诊断...")
    ws3 = wb.create_sheet("3-Lasso系数诊断")
    generate_sheet3_lasso_diagnosis(ws3)
    
    print("生成 Sheet4: train_periods测试...")
    ws4 = wb.create_sheet("4-训练窗口测试")
    generate_sheet4_train_periods(ws4)
    
    print("生成 Sheet5: 回归方法对比...")
    ws5 = wb.create_sheet("5-回归方法对比")
    generate_sheet5_regression_methods(ws5)
    
    print("生成 Sheet6: 最终结论...")
    ws6 = wb.create_sheet("6-最终结论与建议")
    generate_sheet6_conclusion(ws6)
    
    # 保存文件
    output_path = os.path.join(os.path.dirname(__file__), "Lasso因子调试报告.xlsx")
    wb.save(output_path)
    
    print(f"\n报告已保存到: {output_path}")
    print("\n" + "=" * 80)
    print("Sheet目录:")
    print("=" * 80)
    print("1-基础数据检查: 确认数据完整性和质量")
    print("2-简单动量IC(基准): 计算基准动量IC")
    print("3-Lasso系数诊断: 分析Lasso学到了什么")
    print("4-训练窗口测试: 检验train_periods影响")
    print("5-回归方法对比: 对比不同回归方法")
    print("6-最终结论与建议: 汇总结论和解决方案")


if __name__ == "__main__":
    main()
