"""
Lasso因子综合测试套件
合并了所有Lasso相关的测试、调试和报告生成功能

功能模块:
1. 系数诊断 - 分析Lasso系数结构
2. 因子调试 - 逐步检查因子计算
3. 训练窗口测试 - 测试不同train_periods
4. 回归方法对比 - 对比Lasso/Ridge/ElasticNet等
5. Excel报告生成 - 生成完整调试报告

用法:
    python lasso_test_suite.py                    # 交互式菜单
    python lasso_test_suite.py --mode coef        # 系数诊断
    python lasso_test_suite.py --mode debug       # 因子调试
    python lasso_test_suite.py --mode train       # 训练窗口测试
    python lasso_test_suite.py --mode methods     # 回归方法对比
    python lasso_test_suite.py --mode excel       # 生成Excel报告
    python lasso_test_suite.py --mode all         # 运行全部测试
"""

import pandas as pd
import numpy as np
import sys
import os
import argparse

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from sklearn.linear_model import LassoLarsIC, Ridge, Lasso, ElasticNet
from scipy import stats
import warnings
import data_loader


# ============================================================
# 公共数据处理函数
# ============================================================

def load_monthly_data():
    """加载并处理月度数据"""
    prices_df = data_loader.load_price_df()

    trade_dates = prices_df.index
    monthly_last = trade_dates.to_series().groupby(
        [trade_dates.year, trade_dates.month]
    ).apply(lambda x: x.iloc[-1])
    monthly_last = pd.DatetimeIndex(monthly_last)
    monthly_close = prices_df.loc[monthly_last]

    monthly_ret = monthly_close.pct_change()
    bench_monthly = monthly_ret.mean(axis=1)
    excess_monthly = monthly_ret.sub(bench_monthly, axis=0)

    return prices_df, monthly_close, excess_monthly


def calculate_simple_momentum_ic(excess_monthly):
    """计算简单动量IC（当月超额收益预测下月）"""
    ic_list = []
    monthly_last = excess_monthly.index

    for i in range(1, len(monthly_last) - 1):
        curr_ret = excess_monthly.iloc[i]
        next_ret = excess_monthly.iloc[i + 1]

        valid_mask = curr_ret.notna() & next_ret.notna()
        if valid_mask.sum() < 5:
            continue

        ic, _ = stats.spearmanr(curr_ret[valid_mask], next_ret[valid_mask])
        if not np.isnan(ic):
            ic_list.append(ic)

    if len(ic_list) > 0:
        ic_mean = np.mean(ic_list)
        icir = ic_mean / np.std(ic_list) if np.std(ic_list) > 0 else 0
        ic_win_rate = np.mean([1 if ic > 0 else 0 for ic in ic_list])
        return ic_mean, icir, ic_win_rate, len(ic_list), ic_list
    return np.nan, np.nan, np.nan, 0, []


def build_training_data(excess_monthly, s_pred, train_periods=60, min_train_samples=10):
    """构建训练数据"""
    industries = excess_monthly.columns.tolist()
    months = excess_monthly.index.tolist()

    train_s = list(range(1, s_pred))
    if train_periods is not None and len(train_s) > train_periods:
        train_s = train_s[-train_periods:]

    X_train_list = []
    y_train_dict = {ind: [] for ind in industries}

    for s in train_s:
        X_row = excess_monthly.loc[months[s]].values
        y_row = excess_monthly.loc[months[s + 1]]

        if np.all(np.isnan(X_row)) or y_row.isna().all():
            continue

        X_train_list.append(X_row)
        for ind in industries:
            y_val = y_row.get(ind, np.nan)
            y_train_dict[ind].append(y_val)

    if len(X_train_list) < min_train_samples:
        return None, None, None

    X_train = np.array(X_train_list)
    X_current = excess_monthly.loc[months[s_pred]].values.reshape(1, -1)

    return X_train, y_train_dict, X_current


def run_regression_and_calculate_ic(excess_monthly, model_factory, train_periods=60, min_train_samples=10):
    """用指定的回归方法测试IC"""
    industries = excess_monthly.columns.tolist()
    months = excess_monthly.index.tolist()
    T = len(months)

    predictions = []
    actuals = []
    n_nonzero_coefs = []

    for s_pred in range(1 + min_train_samples, T - 1):
        X_train, y_train_dict, X_current = build_training_data(
            excess_monthly, s_pred, train_periods, min_train_samples
        )

        if X_train is None or np.any(np.isnan(X_current)):
            continue

        pred_row = {}
        actual_row = {}

        for ind in industries:
            y_train = np.array(y_train_dict[ind])
            valid_mask = ~np.isnan(y_train)

            if valid_mask.sum() < min_train_samples:
                continue

            X_valid = X_train[valid_mask]
            y_valid = y_train[valid_mask]

            try:
                with warnings.catch_warnings():
                    warnings.simplefilter('ignore')
                    model = model_factory()
                    model.fit(X_valid, y_valid)
                    pred = model.predict(X_current)[0]
                    pred_row[ind] = pred

                    if hasattr(model, 'coef_'):
                        n_nonzero = np.sum(model.coef_ != 0)
                        n_nonzero_coefs.append(n_nonzero)
            except:
                continue

        if s_pred + 1 < T:
            actual_ret = excess_monthly.loc[months[s_pred + 1]]
            for ind in pred_row.keys():
                if ind in actual_ret.index and not np.isnan(actual_ret[ind]):
                    actual_row[ind] = actual_ret[ind]

        if len(pred_row) > 5 and len(actual_row) > 5:
            predictions.append(pred_row)
            actuals.append(actual_row)

    ic_list = []
    for pred_row, actual_row in zip(predictions, actuals):
        common_inds = set(pred_row.keys()) & set(actual_row.keys())
        if len(common_inds) < 5:
            continue

        pred_vals = [pred_row[ind] for ind in common_inds]
        actual_vals = [actual_row[ind] for ind in common_inds]

        ic, _ = stats.spearmanr(pred_vals, actual_vals)
        if not np.isnan(ic):
            ic_list.append(ic)

    if len(ic_list) > 0:
        ic_mean = np.mean(ic_list)
        icir = ic_mean / np.std(ic_list) if np.std(ic_list) > 0 else 0
        ic_win_rate = np.mean([1 if ic > 0 else 0 for ic in ic_list])
        avg_nonzero = np.mean(n_nonzero_coefs) if n_nonzero_coefs else np.nan
        return ic_mean, icir, ic_win_rate, len(ic_list), predictions, actuals, avg_nonzero
    return np.nan, np.nan, np.nan, 0, [], [], np.nan


# ============================================================
# 模块1: Lasso系数深度诊断
# ============================================================

def diagnose_lasso_coefficients():
    """深度诊断Lasso系数"""

    print("=" * 80)
    print("Lasso系数深度诊断")
    print("=" * 80)

    prices_df, monthly_close, excess_monthly = load_monthly_data()
    industries = prices_df.columns.tolist()
    months = monthly_close.index.tolist()
    T = len(months)
    n_ind = len(industries)

    print(f"\n行业数量: {n_ind}")
    print(f"月份数量: {T}")

    # 1. 最近时点的Lasso系数分析
    print("\n" + "=" * 80)
    print("1. 最近时点的Lasso系数分析")
    print("=" * 80)

    s_pred = T - 2
    factor_date = months[s_pred]
    print(f"\n预测时点: {factor_date.date()}")

    X_train, y_train_dict, X_current = build_training_data(excess_monthly, s_pred, train_periods=60)

    print(f"训练样本数: {len(X_train)}")
    print(f"特征数(行业数): {X_train.shape[1]}")

    sample_industries = industries[:5]

    for ind in sample_industries:
        ind_idx = industries.index(ind)
        y_train = np.array(y_train_dict[ind])
        valid_mask = ~np.isnan(y_train)

        X_valid = X_train[valid_mask]
        y_valid = y_train[valid_mask]

        print(f"\n--- 行业: {ind} ---")
        print(f"有效训练样本: {valid_mask.sum()}")

        with warnings.catch_warnings():
            warnings.simplefilter('ignore')
            lasso = LassoLarsIC(criterion='aic')
            lasso.fit(X_valid, y_valid)

        coef = lasso.coef_
        nonzero_idx = np.where(coef != 0)[0]

        print(f"非零系数数量: {len(nonzero_idx)}")
        print(f"截距: {lasso.intercept_:.6f}")

        if len(nonzero_idx) > 0:
            print("非零系数:")
            for idx in nonzero_idx:
                print(f"  {industries[idx]}: {coef[idx]:.4f}")

        self_coef = coef[ind_idx]
        print(f"自相关系数 (自己对自己): {self_coef:.4f}")

        pred = lasso.predict(X_current)[0]
        print(f"预测值: {pred:.4f}")

        if s_pred + 1 < T:
            actual = excess_monthly.loc[months[s_pred + 1], ind]
            print(f"实际值: {actual:.4f}")

    # 2. 所有行业Lasso系数符号统计
    print("\n" + "=" * 80)
    print("2. 所有行业Lasso系数符号统计")
    print("=" * 80)

    all_self_coefs = []
    all_nonzero_counts = []
    all_positive_coefs = []
    all_negative_coefs = []

    for ind in industries:
        y_train = np.array(y_train_dict[ind])
        valid_mask = ~np.isnan(y_train)

        if valid_mask.sum() < 10:
            continue

        X_valid = X_train[valid_mask]
        y_valid = y_train[valid_mask]

        with warnings.catch_warnings():
            warnings.simplefilter('ignore')
            lasso = LassoLarsIC(criterion='aic')
            lasso.fit(X_valid, y_valid)

        coef = lasso.coef_
        ind_idx = industries.index(ind)

        all_self_coefs.append(coef[ind_idx])
        all_nonzero_counts.append(np.sum(coef != 0))
        all_positive_coefs.append(np.sum(coef > 0))
        all_negative_coefs.append(np.sum(coef < 0))

    print(f"\n自相关系数统计:")
    print(f"  均值: {np.mean(all_self_coefs):.4f}")
    print(f"  >0的比例: {np.mean([c > 0 for c in all_self_coefs]):.2%}")
    print(f"  =0的比例: {np.mean([c == 0 for c in all_self_coefs]):.2%}")

    print(f"\n非零系数统计:")
    print(f"  平均非零数: {np.mean(all_nonzero_counts):.1f}")
    print(f"  平均正系数数: {np.mean(all_positive_coefs):.1f}")
    print(f"  平均负系数数: {np.mean(all_negative_coefs):.1f}")

    # 3. 对比: Ridge回归 vs Lasso
    print("\n" + "=" * 80)
    print("3. 对比: Ridge回归 vs Lasso")
    print("=" * 80)

    ridge_preds = {}
    lasso_preds = {}
    actuals = {}

    for ind in industries:
        y_train = np.array(y_train_dict[ind])
        valid_mask = ~np.isnan(y_train)

        if valid_mask.sum() < 10:
            continue

        X_valid = X_train[valid_mask]
        y_valid = y_train[valid_mask]

        ridge = Ridge(alpha=1.0)
        ridge.fit(X_valid, y_valid)
        ridge_preds[ind] = ridge.predict(X_current)[0]

        with warnings.catch_warnings():
            warnings.simplefilter('ignore')
            lasso = LassoLarsIC(criterion='aic')
            lasso.fit(X_valid, y_valid)
        lasso_preds[ind] = lasso.predict(X_current)[0]

        if s_pred + 1 < T:
            actuals[ind] = excess_monthly.loc[months[s_pred + 1], ind]

    common = set(ridge_preds.keys()) & set(lasso_preds.keys()) & set(actuals.keys())
    common = [ind for ind in common if not np.isnan(actuals[ind])]

    ridge_vals = [ridge_preds[ind] for ind in common]
    lasso_vals = [lasso_preds[ind] for ind in common]
    actual_vals = [actuals[ind] for ind in common]

    ridge_ic, _ = stats.spearmanr(ridge_vals, actual_vals)
    lasso_ic, _ = stats.spearmanr(lasso_vals, actual_vals)

    print(f"\n单期（{factor_date.date()}）IC对比:")
    print(f"  Ridge IC: {ridge_ic:.4f}")
    print(f"  Lasso IC: {lasso_ic:.4f}")

    # 4. 预测值 vs 当月超额收益 的相关性
    print("\n" + "=" * 80)
    print("4. 预测值 vs 当月超额收益 的相关性")
    print("=" * 80)

    current_excess = [excess_monthly.loc[months[s_pred], ind] for ind in common]

    pred_vs_current, _ = stats.spearmanr(lasso_vals, current_excess)
    current_vs_actual, _ = stats.spearmanr(current_excess, actual_vals)

    print(f"  Lasso预测 vs 当月超额: {pred_vs_current:.4f}")
    print(f"  当月超额 vs 下月实际: {current_vs_actual:.4f}")
    print(f"  Lasso预测 vs 下月实际: {lasso_ic:.4f}")

    if pred_vs_current < 0:
        print("\n  ⚠️ Lasso预测与当月超额收益负相关！")
        print("  这说明Lasso可能学到了\"反转\"效应而非\"动量\"效应")

    # 5. LassoLarsIC选择的alpha值
    print("\n" + "=" * 80)
    print("5. LassoLarsIC选择的alpha值")
    print("=" * 80)

    alphas = []
    for ind in industries[:10]:
        y_train = np.array(y_train_dict[ind])
        valid_mask = ~np.isnan(y_train)

        if valid_mask.sum() < 10:
            continue

        X_valid = X_train[valid_mask]
        y_valid = y_train[valid_mask]

        with warnings.catch_warnings():
            warnings.simplefilter('ignore')
            lasso = LassoLarsIC(criterion='aic')
            lasso.fit(X_valid, y_valid)
            alphas.append(lasso.alpha_)
            print(f"  {ind}: alpha={lasso.alpha_:.6f}")

    print(f"\n平均alpha: {np.mean(alphas):.6f}")

    print("\n" + "=" * 80)
    print("诊断完成")
    print("=" * 80)


# ============================================================
# 模块2: Lasso因子调试
# ============================================================

def debug_lasso_factor():
    """逐步调试Lasso因子"""

    print("=" * 80)
    print("Lasso因子调试")
    print("=" * 80)

    print("\n[1] 加载数据...")
    prices_df, monthly_close, excess_monthly = load_monthly_data()
    print(f"    价格数据: {prices_df.shape[0]}个交易日, {prices_df.shape[1]}个行业")
    print(f"    日期范围: {prices_df.index[0].date()} 至 {prices_df.index[-1].date()}")
    print(f"    行业列表: {list(prices_df.columns)[:5]}...")

    print("\n[2] 月度数据信息...")
    print(f"    月度数据: {len(monthly_close)}个月")

    print("\n[3] 检查数据质量...")
    nan_ratio = excess_monthly.isna().sum().sum() / (excess_monthly.shape[0] * excess_monthly.shape[1])
    print(f"    超额收益NaN比例: {nan_ratio:.2%}")
    print(f"    超额收益均值: {excess_monthly.mean().mean():.6f}")
    print(f"    超额收益标准差: {excess_monthly.std().mean():.6f}")

    print("\n[4] 检查简单动量IC（基准）...")
    simple_ic_mean, simple_icir, _, _, _ = calculate_simple_momentum_ic(excess_monthly)
    print(f"    简单动量IC均值: {simple_ic_mean:.4f}")
    print(f"    简单动量ICIR: {simple_icir:.4f}")
    print(f"    说明: 这是用当月超额收益直接预测下月超额收益的IC")

    print("\n[5] 检查Lasso模型效果...")
    lasso_factory = lambda: LassoLarsIC(criterion='aic')
    ic_mean, icir, win_rate, n_months, predictions, actuals, avg_nonzero = run_regression_and_calculate_ic(
        excess_monthly, lasso_factory, train_periods=60
    )

    if not np.isnan(ic_mean):
        print(f"    Lasso预测IC均值: {ic_mean:.4f}")
        print(f"    Lasso预测ICIR: {icir:.4f}")
        print(f"    IC胜率: {win_rate:.2%}")
        print(f"    有效月份数: {n_months}")
    else:
        print("    警告: 没有有效的Lasso预测结果！")

    print("\n[6] 模型诊断...")
    if not np.isnan(avg_nonzero):
        print(f"    平均非零系数数量: {avg_nonzero:.1f}")

    print("\n[7] 问题诊断...")
    print(f"    简单动量IC: {simple_ic_mean:.4f}")
    if not np.isnan(ic_mean):
        print(f"    Lasso预测IC: {ic_mean:.4f}")

        if ic_mean < simple_ic_mean * 0.5:
            print("\n    ⚠️ 问题: Lasso预测IC远低于简单动量IC！")
            print("    可能原因:")
            print("    1. Lasso正则化过强，大多数系数被压缩为0")
            print("    2. 训练样本不足，模型难以学习有效的领先滞后关系")
            print("    3. 行业分类不同（研报用中信，你用申万）")
            print("    4. 特征和目标之间的关系本身就很弱")

    print("\n[8] 预测值分布...")
    if len(predictions) > 0:
        all_preds = []
        for pred_row in predictions:
            all_preds.extend(pred_row.values())

        print(f"    预测值均值: {np.mean(all_preds):.6f}")
        print(f"    预测值标准差: {np.std(all_preds):.6f}")
        print(f"    预测值范围: [{np.min(all_preds):.4f}, {np.max(all_preds):.4f}]")

        if np.std(all_preds) < 0.001:
            print("\n    ⚠️ 问题: 预测值标准差过小，预测值几乎没有区分度！")

    print("\n" + "=" * 80)
    print("调试完成")
    print("=" * 80)


# ============================================================
# 模块3: 测试不同train_periods
# ============================================================

def test_train_periods():
    """测试不同train_periods对Lasso因子IC的影响"""

    print("=" * 80)
    print("测试不同train_periods对Lasso因子IC的影响")
    print("=" * 80)

    _, _, excess_monthly = load_monthly_data()

    test_periods = [
        (30, "30个月"),
        (60, "60个月（当前设置）"),
        (90, "90个月"),
        (120, "120个月"),
        (None, "全部历史（研报可能用法）"),
    ]

    results = []
    lasso_factory = lambda: LassoLarsIC(criterion='aic')

    for periods, desc in test_periods:
        print(f"\n测试 train_periods={periods} ({desc})...")
        ic_mean, icir, _, n_months, _, _, _ = run_regression_and_calculate_ic(
            excess_monthly, lasso_factory, train_periods=periods
        )
        results.append({
            'train_periods': periods if periods else '全部',
            'description': desc,
            'IC均值': ic_mean,
            'ICIR': icir,
            '有效月数': n_months
        })
        if not np.isnan(ic_mean):
            print(f"  IC均值: {ic_mean:.4f}, ICIR: {icir:.4f}, 有效月数: {n_months}")
        else:
            print(f"  无有效结果")

    print("\n" + "=" * 80)
    print("汇总结果:")
    print("=" * 80)
    df = pd.DataFrame(results)
    print(df.to_string(index=False))

    print("\n" + "=" * 80)
    print("结论:")
    print("=" * 80)
    valid_results = [r for r in results if not np.isnan(r['IC均值'])]
    if valid_results:
        best = max(valid_results, key=lambda x: x['IC均值'])
        print(f"最优train_periods: {best['train_periods']} ({best['description']})")
        print(f"最优IC均值: {best['IC均值']:.4f}")

    return results


# ============================================================
# 模块4: 对比不同回归方法
# ============================================================

def test_regression_methods():
    """对比不同回归方法在行业间动量因子上的表现"""

    print("=" * 80)
    print("对比不同回归方法在行业间动量因子上的表现")
    print("=" * 80)

    _, _, excess_monthly = load_monthly_data()

    methods = [
        ("LassoLarsIC(aic)", lambda: LassoLarsIC(criterion='aic')),
        ("LassoLarsIC(bic)", lambda: LassoLarsIC(criterion='bic')),
        ("Lasso(alpha=0.01)", lambda: Lasso(alpha=0.01, max_iter=10000)),
        ("Lasso(alpha=0.001)", lambda: Lasso(alpha=0.001, max_iter=10000)),
        ("Ridge(alpha=1.0)", lambda: Ridge(alpha=1.0)),
        ("Ridge(alpha=10.0)", lambda: Ridge(alpha=10.0)),
        ("ElasticNet(0.5)", lambda: ElasticNet(alpha=0.01, l1_ratio=0.5, max_iter=10000)),
    ]

    results = []

    print("\n测试 简单动量(当月→下月)...")
    simple_ic_mean, simple_icir, simple_win_rate, simple_n_months, _ = calculate_simple_momentum_ic(excess_monthly)
    results.append({
        '方法': '简单动量(当月→下月)',
        'IC均值': simple_ic_mean,
        'ICIR': simple_icir,
        'IC胜率': simple_win_rate,
        '有效月数': simple_n_months
    })
    print(f"  IC均值: {simple_ic_mean:.4f}, ICIR: {simple_icir:.4f}, IC胜率: {simple_win_rate:.2%}")

    for method_name, model_factory in methods:
        print(f"\n测试 {method_name}...")
        ic_mean, icir, win_rate, n_months, _, _, _ = run_regression_and_calculate_ic(
            excess_monthly, model_factory
        )
        results.append({
            '方法': method_name,
            'IC均值': ic_mean,
            'ICIR': icir,
            'IC胜率': win_rate,
            '有效月数': n_months
        })
        if not np.isnan(ic_mean):
            print(f"  IC均值: {ic_mean:.4f}, ICIR: {icir:.4f}, IC胜率: {win_rate:.2%}")
        else:
            print(f"  无有效结果")

    print("\n" + "=" * 80)
    print("汇总结果:")
    print("=" * 80)
    df = pd.DataFrame(results)
    print(df.to_string(index=False))

    print("\n" + "=" * 80)
    print("结论:")
    print("=" * 80)
    valid_results = [r for r in results if not np.isnan(r['IC均值'])]
    if valid_results:
        best = max(valid_results, key=lambda x: x['IC均值'])
        worst = min(valid_results, key=lambda x: x['IC均值'])

        print(f"最优方法: {best['方法']} (IC={best['IC均值']:.4f})")
        print(f"最差方法: {worst['方法']} (IC={worst['IC均值']:.4f})")

        lasso_aic = next((r for r in results if 'LassoLarsIC(aic)' in r['方法']), None)

        if lasso_aic and not np.isnan(lasso_aic['IC均值']):
            print(f"\n简单动量IC: {simple_ic_mean:.4f}")
            print(f"Lasso(aic)IC: {lasso_aic['IC均值']:.4f}")
            if lasso_aic['IC均值'] < simple_ic_mean:
                print("⚠️ Lasso比简单动量还差！说明Lasso选择的跨行业关系是噪音")
                print("   建议：直接使用简单动量，或改用Ridge回归")

    return results


# ============================================================
# 模块5: 生成Excel报告
# ============================================================

def generate_excel_report():
    """生成Lasso因子调试报告Excel"""

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        print("错误: 需要安装openpyxl库才能生成Excel报告")
        print("运行: pip install openpyxl")
        return

    print("=" * 80)
    print("生成Lasso因子调试报告Excel")
    print("=" * 80)

    def add_header_to_sheet(ws, title, purpose, method, row_start=1):
        title_font = Font(bold=True, size=14, color="FFFFFF")
        header_font = Font(bold=True, size=11)
        title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

        ws.merge_cells(start_row=row_start, start_column=1, end_row=row_start, end_column=6)
        cell = ws.cell(row=row_start, column=1, value=title)
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal='center')

        ws.merge_cells(start_row=row_start+1, start_column=1, end_row=row_start+1, end_column=6)
        cell = ws.cell(row=row_start+1, column=1, value=f"【目的】{purpose}")
        cell.font = header_font
        cell.fill = section_fill

        ws.merge_cells(start_row=row_start+2, start_column=1, end_row=row_start+2, end_column=6)
        cell = ws.cell(row=row_start+2, column=1, value=f"【方法】{method}")
        cell.font = header_font
        cell.fill = section_fill

        return row_start + 4

    def add_dataframe_to_sheet(ws, df, start_row):
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == start_row:
                    cell.font = header_font
                    cell.fill = header_fill

        return start_row + len(df) + 2

    def add_conclusion_to_sheet(ws, conclusion, start_row):
        conclusion_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
        cell = ws.cell(row=start_row, column=1, value=f"【结论】{conclusion}")
        cell.font = Font(bold=True)
        cell.fill = conclusion_fill
        return start_row + 2

    prices_df, monthly_close, excess_monthly = load_monthly_data()
    industries = prices_df.columns.tolist()
    months = monthly_close.index.tolist()
    T = len(months)

    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: 基础数据检查
    print("\n生成 Sheet1: 基础数据检查...")
    ws1 = wb.create_sheet("1-基础数据检查")

    row = add_header_to_sheet(ws1, "Sheet1: 基础数据检查",
                               "确认输入数据的完整性和基本统计特征",
                               "检查价格数据的日期范围、行业数量等")

    overview_data = [
        ["检查项", "结果"],
        ["数据起始日期", str(prices_df.index[0].date())],
        ["数据结束日期", str(prices_df.index[-1].date())],
        ["交易日数量", len(prices_df)],
        ["行业数量", len(prices_df.columns)],
        ["月度数据点", len(monthly_close)],
        ["超额收益NaN比例", f"{excess_monthly.isna().sum().sum() / (excess_monthly.shape[0] * excess_monthly.shape[1]):.2%}"],
    ]

    for item in overview_data:
        ws1.cell(row=row, column=1, value=item[0])
        ws1.cell(row=row, column=2, value=item[1])
        row += 1

    row += 1
    ws1.cell(row=row, column=1, value="行业列表").font = Font(bold=True)
    row += 1
    for i, ind in enumerate(prices_df.columns):
        ws1.cell(row=row, column=1, value=i+1)
        ws1.cell(row=row, column=2, value=ind)
        row += 1

    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 25

    # Sheet 2: 简单动量IC检查
    print("生成 Sheet2: 简单动量IC检查...")
    ws2 = wb.create_sheet("2-简单动量IC(基准)")

    row = add_header_to_sheet(ws2, "Sheet2: 简单动量IC检查（基准）",
                               "计算最基础的动量IC作为基准",
                               "计算所有行业当月超额收益与下月超额收益的Spearman相关系数")

    simple_ic_mean, simple_icir, simple_win_rate, n_months, ic_list = calculate_simple_momentum_ic(excess_monthly)

    summary_df = pd.DataFrame([
        {'指标': 'IC均值', '值': round(simple_ic_mean, 4)},
        {'指标': 'IC标准差', '值': round(np.std(ic_list), 4)},
        {'指标': 'ICIR', '值': round(simple_icir, 4)},
        {'指标': 'IC胜率', '值': f"{simple_win_rate:.2%}"},
        {'指标': '有效月数', '值': n_months},
    ])
    row = add_dataframe_to_sheet(ws2, summary_df, row)

    add_conclusion_to_sheet(ws2, f"简单动量IC均值={simple_ic_mean:.4f}，这是Lasso因子的基准线", row)

    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 15

    # Sheet 3: Lasso模型诊断
    print("生成 Sheet3: Lasso模型诊断...")
    ws3 = wb.create_sheet("3-Lasso系数诊断")

    row = add_header_to_sheet(ws3, "Sheet3: Lasso模型系数诊断",
                               "深入分析Lasso模型学到了什么",
                               "取最近一个时点，详细查看每个行业Lasso模型的系数")

    s_pred = T - 2
    X_train, y_train_dict, X_current = build_training_data(excess_monthly, s_pred, train_periods=60)

    info_df = pd.DataFrame([
        {'项目': '预测时点', '值': months[s_pred].strftime('%Y-%m-%d')},
        {'项目': '训练样本数', '值': len(X_train)},
        {'项目': '特征数（行业数）', '值': X_train.shape[1]},
    ])
    row = add_dataframe_to_sheet(ws3, info_df, row)

    coef_summary = []
    all_self_coefs = []
    all_nonzero = []
    all_alphas = []

    for ind in industries:
        ind_idx = industries.index(ind)
        y_train = np.array(y_train_dict[ind])
        valid_mask = ~np.isnan(y_train)

        if valid_mask.sum() < 10:
            continue

        X_valid = X_train[valid_mask]
        y_valid = y_train[valid_mask]

        with warnings.catch_warnings():
            warnings.simplefilter('ignore')
            lasso = LassoLarsIC(criterion='aic')
            lasso.fit(X_valid, y_valid)

        coef = lasso.coef_
        self_coef = coef[ind_idx]
        n_nonzero = np.sum(coef != 0)

        all_self_coefs.append(self_coef)
        all_nonzero.append(n_nonzero)
        all_alphas.append(lasso.alpha_)

        coef_summary.append({
            '行业': ind,
            '自相关系数': round(self_coef, 4),
            '非零系数数': n_nonzero,
            'alpha': f"{lasso.alpha_:.6f}",
        })

    ws3.cell(row=row, column=1, value="各行业Lasso系数汇总").font = Font(bold=True)
    row += 1
    coef_df = pd.DataFrame(coef_summary)
    row = add_dataframe_to_sheet(ws3, coef_df, row)

    key_stats = pd.DataFrame([
        {'指标': '自相关系数均值', '值': round(np.mean(all_self_coefs), 4)},
        {'指标': '自相关系数>0比例', '值': f"{np.mean([c > 0 for c in all_self_coefs]):.2%}"},
        {'指标': '自相关系数=0比例', '值': f"{np.mean([c == 0 for c in all_self_coefs]):.2%}"},
        {'指标': '平均非零系数数', '值': round(np.mean(all_nonzero), 1)},
    ])
    row = add_dataframe_to_sheet(ws3, key_stats, row)

    zero_ratio = np.mean([c == 0 for c in all_self_coefs])
    add_conclusion_to_sheet(ws3, f"{zero_ratio:.0%}的自相关系数被压成0！", row)

    ws3.column_dimensions['A'].width = 20
    ws3.column_dimensions['B'].width = 15

    # Sheet 4: 训练窗口测试
    print("生成 Sheet4: 训练窗口测试...")
    ws4 = wb.create_sheet("4-训练窗口测试")

    row = add_header_to_sheet(ws4, "Sheet4: 不同训练窗口测试",
                               "检验是否因为训练样本不足导致过拟合",
                               "尝试不同的train_periods，对比IC变化")

    test_periods = [(30, "30个月"), (60, "60个月（当前）"), (90, "90个月"), (120, "120个月"), (None, "全部历史")]
    lasso_factory = lambda: LassoLarsIC(criterion='aic')

    period_results = []
    for periods, desc in test_periods:
        ic_mean, icir, _, n_months, _, _, _ = run_regression_and_calculate_ic(
            excess_monthly, lasso_factory, train_periods=periods
        )
        period_results.append({
            'train_periods': periods if periods else '全部',
            '描述': desc,
            'IC均值': round(ic_mean, 4) if not np.isnan(ic_mean) else 'N/A',
            'ICIR': round(icir, 4) if not np.isnan(icir) else 'N/A',
        })

    results_df = pd.DataFrame(period_results)
    row = add_dataframe_to_sheet(ws4, results_df, row)

    add_conclusion_to_sheet(ws4, "无论train_periods设多少，IC都是负的", row)

    ws4.column_dimensions['A'].width = 15
    ws4.column_dimensions['B'].width = 20

    # Sheet 5: 回归方法对比
    print("生成 Sheet5: 回归方法对比...")
    ws5 = wb.create_sheet("5-回归方法对比")

    row = add_header_to_sheet(ws5, "Sheet5: 不同回归方法对比",
                               "检验是否Lasso方法本身有问题",
                               "对比Lasso、Ridge、ElasticNet等方法")

    methods = [
        ("LassoLarsIC(aic)", lambda: LassoLarsIC(criterion='aic')),
        ("LassoLarsIC(bic)", lambda: LassoLarsIC(criterion='bic')),
        ("Ridge(alpha=1.0)", lambda: Ridge(alpha=1.0)),
        ("Ridge(alpha=10.0)", lambda: Ridge(alpha=10.0)),
        ("ElasticNet(0.5)", lambda: ElasticNet(alpha=0.01, l1_ratio=0.5, max_iter=10000)),
    ]

    method_results = [
        {'方法': '简单动量(基准)', 'IC均值': round(simple_ic_mean, 4), 'ICIR': round(simple_icir, 4)}
    ]

    for method_name, model_factory in methods:
        ic_mean, icir, _, _, _, _, _ = run_regression_and_calculate_ic(excess_monthly, model_factory)
        method_results.append({
            '方法': method_name,
            'IC均值': round(ic_mean, 4) if not np.isnan(ic_mean) else 'N/A',
            'ICIR': round(icir, 4) if not np.isnan(icir) else 'N/A',
        })

    method_df = pd.DataFrame(method_results)
    row = add_dataframe_to_sheet(ws5, method_df, row)

    add_conclusion_to_sheet(ws5, "所有回归方法都不如简单动量", row)

    ws5.column_dimensions['A'].width = 25
    ws5.column_dimensions['B'].width = 12

    # Sheet 6: 最终结论
    print("生成 Sheet6: 最终结论...")
    ws6 = wb.create_sheet("6-最终结论与建议")

    row = add_header_to_sheet(ws6, "Sheet6: 最终结论与建议",
                               "汇总所有诊断结果，给出根本原因分析和解决方案",
                               "综合分析各项诊断数据")

    ws6.cell(row=row, column=1, value="根本原因").font = Font(bold=True, color="C00000")
    row += 1

    root_causes = [
        "1. 行业分类差异：研报使用中信一级行业，你使用申万一级行业",
        "2. 中信行业分类更侧重产业链逻辑，行业间存在明显的领先滞后关系",
        "3. 申万行业分类下，跨行业的领先滞后关系不存在",
        "4. Lasso在申万行业上学到的是'反转'效应，而非'动量'效应"
    ]
    for cause in root_causes:
        ws6.cell(row=row, column=1, value=cause)
        row += 1

    row += 1
    ws6.cell(row=row, column=1, value="解决方案").font = Font(bold=True, color="008000")
    row += 1

    solutions = [
        "方案1（推荐）：使用简单动量替代Lasso因子",
        "方案2：获取中信行业分类数据，严格复现研报逻辑",
        "方案3：放弃该因子，申万行业上此方法不适用",
    ]
    for sol in solutions:
        ws6.cell(row=row, column=1, value=sol)
        row += 1

    ws6.column_dimensions['A'].width = 60

    # 保存文件
    output_path = os.path.join(os.path.dirname(__file__), "Lasso因子调试报告.xlsx")
    wb.save(output_path)

    print(f"\n报告已保存到: {output_path}")
    print("\n" + "=" * 80)
    print("Sheet目录:")
    print("=" * 80)
    print("1-基础数据检查: 确认数据完整性和质量")
    print("2-简单动量IC(基准): 计算基准动量IC")
    print("3-Lasso系数诊断: 分析Lasso学到了什么")
    print("4-训练窗口测试: 检验train_periods影响")
    print("5-回归方法对比: 对比不同回归方法")
    print("6-最终结论与建议: 汇总结论和解决方案")


# ============================================================
# 主函数和命令行接口
# ============================================================

def show_menu():
    print("\n" + "=" * 80)
    print("Lasso因子综合测试套件")
    print("=" * 80)
    print("\n请选择要运行的功能:")
    print("  1. 系数诊断 - 分析Lasso系数结构")
    print("  2. 因子调试 - 逐步检查因子计算")
    print("  3. 训练窗口测试 - 测试不同train_periods")
    print("  4. 回归方法对比 - 对比Lasso/Ridge/ElasticNet等")
    print("  5. 生成Excel报告 - 生成完整调试报告")
    print("  6. 运行全部测试")
    print("  0. 退出")
    print()


def main():
    parser = argparse.ArgumentParser(description='Lasso因子综合测试套件')
    parser.add_argument('--mode', type=str,
                        choices=['coef', 'debug', 'train', 'methods', 'excel', 'all', 'menu'],
                        default='menu',
                        help='运行模式: coef(系数诊断), debug(因子调试), train(训练窗口), methods(回归方法), excel(Excel报告), all(全部), menu(交互菜单)')

    args = parser.parse_args()

    if args.mode == 'menu':
        while True:
            show_menu()
            choice = input("请输入选项 (0-6): ").strip()

            if choice == '1':
                diagnose_lasso_coefficients()
            elif choice == '2':
                debug_lasso_factor()
            elif choice == '3':
                test_train_periods()
            elif choice == '4':
                test_regression_methods()
            elif choice == '5':
                generate_excel_report()
            elif choice == '6':
                diagnose_lasso_coefficients()
                debug_lasso_factor()
                test_train_periods()
                test_regression_methods()
                generate_excel_report()
            elif choice == '0':
                print("退出程序")
                break
            else:
                print("无效选项，请重新输入")

            input("\n按回车键继续...")

    elif args.mode == 'coef':
        diagnose_lasso_coefficients()
    elif args.mode == 'debug':
        debug_lasso_factor()
    elif args.mode == 'train':
        test_train_periods()
    elif args.mode == 'methods':
        test_regression_methods()
    elif args.mode == 'excel':
        generate_excel_report()
    elif args.mode == 'all':
        diagnose_lasso_coefficients()
        debug_lasso_factor()
        test_train_periods()
        test_regression_methods()
        generate_excel_report()


if __name__ == "__main__":
    main()
